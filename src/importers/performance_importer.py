# -*- coding: utf-8 -*-

from __future__ import annotations

import math
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.config.model_config import load_model_version
from src.config.path_config import load_paths_config
from src.database.dataset_versions import get_dataset_version
from src.database.init_database import connect_database
from src.utils.logger import get_logger


logger = get_logger("importers.performance")


PERFORMANCE_SHEET_NAME = "Desempenho_2025_26"

REQUIRED_HEADERS = [
    "team_id",
    "source_league_id",
    "target_league_id",
    "season_label",
    "position",
    "played",
    "wins",
    "draws",
    "losses",
    "goals_for",
    "goals_against",
    "goal_difference",
    "points",
    "points_adjustment",
    "promoted",
    "promotion_method",
    "source_status",
    "data_confidence",
    "source_url",
    "accessed_at",
]

VALID_PROMOTION_METHODS = {
    "CHAMPION",
    "DIRECT",
    "PLAYOFF",
}

VALID_SOURCE_STATUSES = {
    "CONFIRMED",
    "COMPLETE",
    "PARTIAL",
    "CACHED",
    "MISSING",
    "CONFLICTING",
    "OUTDATED",
    "MANUAL_VALIDATED",
}


@dataclass
class PerformanceImportResult:
    inserted: int = 0
    updated: int = 0
    unchanged: int = 0
    skipped: int = 0
    errors: int = 0

    @property
    def processed(self) -> int:
        return (
            self.inserted
            + self.updated
            + self.unchanged
            + self.skipped
            + self.errors
        )


class PerformanceImportError(RuntimeError):
    """Erro ocorrido durante a importação dos desempenhos."""


def import_performance(
    dataset_path: str | Path | None = None,
    dataset_version: str | None = None,
    require_approved_dataset: bool = True,
    database_path: str | Path | None = None,
) -> PerformanceImportResult:
    """
    Importa a folha Desempenho_2025_26 para a SQLite.

    A operação é atómica:

    1. Lê todos os registos do Excel.
    2. Valida e prepara todos os registos sem gravar.
    3. Se existir algum erro, cancela a importação.
    4. Só grava quando todos os registos forem válidos.
    5. Qualquer erro durante a gravação provoca rollback total.
    """

    model_config = load_model_version()

    final_dataset_version = (
        dataset_version.strip()
        if dataset_version
        else str(
            model_config["dataset"]["expected_version"]
        )
    )

    final_dataset_path = _resolve_dataset_path(
        dataset_path
    )

    if require_approved_dataset:
        _assert_dataset_is_approved(
            dataset_version=final_dataset_version,
            dataset_path=final_dataset_path,
        )

    records = read_performance_from_excel(
        final_dataset_path
    )

    if not records:
        raise PerformanceImportError(
            "A folha Desempenho_2025_26 não contém registos."
        )

    result = PerformanceImportResult()
    connection = connect_database(database_path)

    logger.info(
        "A iniciar importação de desempenhos | "
        "dataset=%s | total=%s",
        final_dataset_version,
        len(records),
    )

    prepared_records: list[dict[str, Any]] = []
    preparation_errors: list[str] = []

    try:
        # ==========================================================
        # FASE 1 — Validar todos os registos sem gravar na base
        # ==========================================================

        for row_number, record in records:
            try:
                prepared = prepare_performance_record(
                    connection=connection,
                    record=record,
                    row_number=row_number,
                    dataset_version=final_dataset_version,
                )

                prepared_records.append(
                    prepared
                )

            except Exception as exc:
                result.errors += 1

                error_message = (
                    f"Linha {row_number} | "
                    f"team_id={record.get('team_id')} | "
                    f"erro={exc}"
                )

                preparation_errors.append(
                    error_message
                )

                logger.error(
                    "Erro ao preparar desempenho | %s",
                    error_message,
                )

        if preparation_errors:
            details = "\n".join(
                f" - {message}"
                for message in preparation_errors
            )

            raise PerformanceImportError(
                "A importação foi cancelada antes da gravação. "
                f"Foram encontradas {result.errors} linha(s) inválida(s):\n"
                f"{details}"
            )

        # ==========================================================
        # FASE 2 — Gravar apenas quando tudo estiver válido
        # ==========================================================

        try:
            connection.execute(
                "BEGIN IMMEDIATE"
            )

            for prepared in prepared_records:
                action = upsert_performance(
                    connection=connection,
                    performance=prepared,
                )

                if action == "INSERTED":
                    result.inserted += 1

                elif action == "UPDATED":
                    result.updated += 1

                elif action == "UNCHANGED":
                    result.unchanged += 1

                else:
                    result.skipped += 1

            validate_imported_performance(
                connection=connection,
                dataset_version=final_dataset_version,
                expected_total=len(records),
            )

            connection.commit()

        except Exception:
            connection.rollback()

            logger.exception(
                "Importação revertida integralmente | dataset=%s",
                final_dataset_version,
            )

            raise

    except PerformanceImportError:
        raise

    except sqlite3.Error as exc:
        raise PerformanceImportError(
            f"Erro SQLite durante a importação: {exc}"
        ) from exc

    except Exception as exc:
        raise PerformanceImportError(
            f"Erro durante a importação dos desempenhos: {exc}"
        ) from exc

    finally:
        connection.close()

    logger.info(
        "Importação de desempenhos concluída | "
        "inseridos=%s | atualizados=%s | "
        "inalterados=%s | ignorados=%s | erros=%s",
        result.inserted,
        result.updated,
        result.unchanged,
        result.skipped,
        result.errors,
    )

    return result


def read_performance_from_excel(
    dataset_path: str | Path,
) -> list[tuple[int, dict[str, Any]]]:
    """
    Lê os registos da folha Desempenho_2025_26.

    Devolve uma lista de pares:
        (número da linha, registo)
    """

    path = Path(
        dataset_path
    ).expanduser().resolve()

    if not path.exists():
        raise PerformanceImportError(
            f"O dataset não existe: {path}"
        )

    if not path.is_file():
        raise PerformanceImportError(
            f"O caminho não corresponde a um ficheiro: {path}"
        )

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    try:
        if PERFORMANCE_SHEET_NAME not in workbook.sheetnames:
            raise PerformanceImportError(
                f"Falta a folha obrigatória: "
                f"{PERFORMANCE_SHEET_NAME}"
            )

        worksheet = workbook[
            PERFORMANCE_SHEET_NAME
        ]

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        actual_headers = headers[
            :len(REQUIRED_HEADERS)
        ]

        if actual_headers != REQUIRED_HEADERS:
            raise PerformanceImportError(
                "Os cabeçalhos da folha "
                "Desempenho_2025_26 não correspondem "
                "ao formato esperado."
            )

        records: list[
            tuple[int, dict[str, Any]]
        ] = []

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):
            if all(
                value is None
                for value in row
            ):
                continue

            record = {
                str(headers[index]): row[index]
                for index in range(len(headers))
            }

            records.append(
                (
                    row_number,
                    record,
                )
            )

        return records

    finally:
        workbook.close()


def prepare_performance_record(
    connection: sqlite3.Connection,
    record: dict[str, Any],
    row_number: int,
    dataset_version: str,
) -> dict[str, Any]:
    """
    Limpa, valida e prepara um registo de desempenho.
    """

    team_id = clean_text(
        record.get("team_id")
    )

    source_league_id = clean_text(
        record.get("source_league_id")
    ).upper()

    target_league_id = clean_text(
        record.get("target_league_id")
    ).upper()

    season_label = clean_text(
        record.get("season_label")
    )

    position = to_non_negative_integer(
        value=record.get("position"),
        field_name="position",
        row_number=row_number,
        minimum=1,
    )

    played = to_non_negative_integer(
        value=record.get("played"),
        field_name="played",
        row_number=row_number,
    )

    wins = to_non_negative_integer(
        value=record.get("wins"),
        field_name="wins",
        row_number=row_number,
    )

    draws = to_non_negative_integer(
        value=record.get("draws"),
        field_name="draws",
        row_number=row_number,
    )

    losses = to_non_negative_integer(
        value=record.get("losses"),
        field_name="losses",
        row_number=row_number,
    )

    goals_for = to_non_negative_integer(
        value=record.get("goals_for"),
        field_name="goals_for",
        row_number=row_number,
    )

    goals_against = to_non_negative_integer(
        value=record.get("goals_against"),
        field_name="goals_against",
        row_number=row_number,
    )

    goal_difference = to_integer(
        value=record.get("goal_difference"),
        field_name="goal_difference",
        row_number=row_number,
    )

    points = to_integer(
        value=record.get("points"),
        field_name="points",
        row_number=row_number,
    )

    points_adjustment = to_integer(
        value=record.get("points_adjustment"),
        field_name="points_adjustment",
        row_number=row_number,
        default=0,
    )

    promoted = to_binary_integer(
        value=record.get("promoted"),
        field_name="promoted",
        row_number=row_number,
    )

    promotion_method = clean_optional_text(
        record.get("promotion_method")
    )

    source_status = clean_text(
        record.get("source_status")
    ).upper()

    data_confidence = to_float(
        value=record.get("data_confidence"),
        field_name="data_confidence",
        row_number=row_number,
    )

    source_url = clean_optional_text_preserve_case(
        record.get("source_url")
    )

    accessed_at = normalize_datetime_value(
        record.get("accessed_at")
    )

    if not team_id:
        raise PerformanceImportError(
            f"Linha {row_number}: team_id vazio."
        )

    if not source_league_id:
        raise PerformanceImportError(
            f"Linha {row_number}: source_league_id vazio."
        )

    if not target_league_id:
        raise PerformanceImportError(
            f"Linha {row_number}: target_league_id vazio."
        )

    team = connection.execute(
        """
        SELECT
            team_id,
            league_id,
            promoted,
            promotion_method
        FROM teams
        WHERE team_id = ?
        """,
        (team_id,),
    ).fetchone()

    if team is None:
        raise PerformanceImportError(
            f"Linha {row_number}: "
            f"equipa inexistente: {team_id}"
        )

    if target_league_id != team["league_id"]:
        raise PerformanceImportError(
            f"Linha {row_number}: "
            "target_league_id não corresponde "
            "à liga atual da equipa. "
            f"Esperado={team['league_id']}; "
            f"encontrado={target_league_id}"
        )

    if season_label != "2025/26":
        raise PerformanceImportError(
            f"Linha {row_number}: "
            "season_label deve ser 2025/26."
        )

    expected_played = (
        wins
        + draws
        + losses
    )

    if played != expected_played:
        raise PerformanceImportError(
            f"Linha {row_number}: "
            f"played={played}, mas "
            f"wins+draws+losses={expected_played}."
        )

    expected_goal_difference = (
        goals_for
        - goals_against
    )

    if goal_difference != expected_goal_difference:
        raise PerformanceImportError(
            f"Linha {row_number}: "
            "goal_difference incorreto. "
            f"Esperado={expected_goal_difference}; "
            f"encontrado={goal_difference}"
        )

    expected_points = (
        3 * wins
        + draws
        + points_adjustment
    )

    if points != expected_points:
        raise PerformanceImportError(
            f"Linha {row_number}: "
            f"points incorreto. "
            f"Esperado={expected_points}; "
            f"encontrado={points}"
        )

    if promoted != int(team["promoted"]):
        raise PerformanceImportError(
            f"Linha {row_number}: "
            "promoted não corresponde "
            "ao registo da equipa."
        )

    if promoted == 1:
        if promotion_method not in VALID_PROMOTION_METHODS:
            raise PerformanceImportError(
                f"Linha {row_number}: "
                "promotion_method inválido."
            )

        team_method = (
            str(
                team["promotion_method"]
            ).upper()
            if team["promotion_method"]
            else None
        )

        if team_method != promotion_method:
            raise PerformanceImportError(
                f"Linha {row_number}: "
                "promotion_method não corresponde "
                "ao registo da equipa."
            )

    else:
        promotion_method = None

    if source_status not in VALID_SOURCE_STATUSES:
        raise PerformanceImportError(
            f"Linha {row_number}: "
            f"source_status inválido: {source_status}"
        )

    if not 0 <= data_confidence <= 1:
        raise PerformanceImportError(
            f"Linha {row_number}: "
            "data_confidence deve estar entre 0 e 1."
        )

    return {
        "team_id": team_id,
        "source_league_id": source_league_id,
        "target_league_id": target_league_id,
        "season_label": season_label,
        "position": position,
        "played": played,
        "wins": wins,
        "draws": draws,
        "losses": losses,
        "goals_for": goals_for,
        "goals_against": goals_against,
        "goal_difference": goal_difference,
        "points": points,
        "points_adjustment": points_adjustment,
        "promoted": promoted,
        "promotion_method": promotion_method,
        "source_status": source_status,
        "data_confidence": data_confidence,
        "source_url": source_url,
        "accessed_at": accessed_at,
        "dataset_version": dataset_version,
    }


def upsert_performance(
    connection: sqlite3.Connection,
    performance: dict[str, Any],
) -> str:
    """
    Insere ou atualiza um registo de desempenho.

    Devolve:
        INSERTED
        UPDATED
        UNCHANGED
    """

    existing = connection.execute(
        """
        SELECT *
        FROM team_season_performance
        WHERE team_id = ?
          AND season_label = ?
          AND source_league_id = ?
        """,
        (
            performance["team_id"],
            performance["season_label"],
            performance["source_league_id"],
        ),
    ).fetchone()

    if existing is None:
        connection.execute(
            """
            INSERT INTO team_season_performance (
                team_id,
                source_league_id,
                target_league_id,
                season_label,
                position,
                played,
                wins,
                draws,
                losses,
                goals_for,
                goals_against,
                goal_difference,
                points,
                points_adjustment,
                promoted,
                promotion_method,
                source_status,
                data_confidence,
                source_url,
                accessed_at,
                dataset_version
            )
            VALUES (
                :team_id,
                :source_league_id,
                :target_league_id,
                :season_label,
                :position,
                :played,
                :wins,
                :draws,
                :losses,
                :goals_for,
                :goals_against,
                :goal_difference,
                :points,
                :points_adjustment,
                :promoted,
                :promotion_method,
                :source_status,
                :data_confidence,
                :source_url,
                :accessed_at,
                :dataset_version
            )
            """,
            performance,
        )

        logger.info(
            "Desempenho inserido | "
            "team_id=%s | época=%s",
            performance["team_id"],
            performance["season_label"],
        )

        return "INSERTED"

    if not performance_has_changes(
        existing=existing,
        new_values=performance,
    ):
        return "UNCHANGED"

    connection.execute(
        """
        UPDATE team_season_performance
        SET
            target_league_id = :target_league_id,
            position = :position,
            played = :played,
            wins = :wins,
            draws = :draws,
            losses = :losses,
            goals_for = :goals_for,
            goals_against = :goals_against,
            goal_difference = :goal_difference,
            points = :points,
            points_adjustment = :points_adjustment,
            promoted = :promoted,
            promotion_method = :promotion_method,
            source_status = :source_status,
            data_confidence = :data_confidence,
            source_url = :source_url,
            accessed_at = :accessed_at,
            dataset_version = :dataset_version,
            updated_at = CURRENT_TIMESTAMP
        WHERE team_id = :team_id
          AND season_label = :season_label
          AND source_league_id = :source_league_id
        """,
        performance,
    )

    logger.info(
        "Desempenho atualizado | "
        "team_id=%s | época=%s",
        performance["team_id"],
        performance["season_label"],
    )

    return "UPDATED"


def performance_has_changes(
    existing: sqlite3.Row,
    new_values: dict[str, Any],
) -> bool:
    """
    Confirma se existem alterações entre o registo atual
    e o novo registo.
    """

    fields = (
        "target_league_id",
        "position",
        "played",
        "wins",
        "draws",
        "losses",
        "goals_for",
        "goals_against",
        "goal_difference",
        "points",
        "points_adjustment",
        "promoted",
        "promotion_method",
        "source_status",
        "data_confidence",
        "source_url",
        "accessed_at",
        "dataset_version",
    )

    for field in fields:
        existing_value = existing[field]
        new_value = new_values[field]

        if field == "data_confidence":
            if abs(
                float(existing_value)
                - float(new_value)
            ) > 0.000001:
                return True

        elif existing_value != new_value:
            return True

    return False


def validate_imported_performance(
    connection: sqlite3.Connection,
    dataset_version: str,
    expected_total: int,
) -> None:
    """
    Confirma a coerência dos desempenhos gravados.
    """

    total = connection.execute(
        """
        SELECT COUNT(*) AS total
        FROM team_season_performance
        WHERE dataset_version = ?
        """,
        (dataset_version,),
    ).fetchone()["total"]

    if int(total) != expected_total:
        raise PerformanceImportError(
            "Total de desempenhos incorreto "
            "após importação. "
            f"Esperado={expected_total}; "
            f"encontrado={total}"
        )

    duplicate_rows = connection.execute(
        """
        SELECT
            team_id,
            season_label,
            source_league_id,
            COUNT(*) AS total
        FROM team_season_performance
        WHERE dataset_version = ?
        GROUP BY
            team_id,
            season_label,
            source_league_id
        HAVING COUNT(*) > 1
        """,
        (dataset_version,),
    ).fetchall()

    if duplicate_rows:
        raise PerformanceImportError(
            "Foram encontrados desempenhos duplicados."
        )


def list_imported_performance(
    dataset_version: str | None = None,
    database_path: str | Path | None = None,
) -> list[dict[str, Any]]:
    """
    Lista os desempenhos existentes na SQLite.
    """

    connection = connect_database(
        database_path
    )

    try:
        if dataset_version:
            rows = connection.execute(
                """
                SELECT
                    p.*,
                    t.team_name
                FROM team_season_performance p
                INNER JOIN teams t
                    ON t.team_id = p.team_id
                WHERE p.dataset_version = ?
                ORDER BY
                    p.target_league_id,
                    p.position
                """,
                (dataset_version,),
            ).fetchall()

        else:
            rows = connection.execute(
                """
                SELECT
                    p.*,
                    t.team_name
                FROM team_season_performance p
                INNER JOIN teams t
                    ON t.team_id = p.team_id
                ORDER BY
                    p.target_league_id,
                    p.position
                """
            ).fetchall()

        return [
            dict(row)
            for row in rows
        ]

    finally:
        connection.close()


def count_imported_performance(
    database_path: str | Path | None = None,
) -> dict[str, int]:
    """
    Conta os desempenhos por liga de destino.
    """

    connection = connect_database(
        database_path
    )

    try:
        rows = connection.execute(
            """
            SELECT
                target_league_id,
                COUNT(*) AS total
            FROM team_season_performance
            GROUP BY target_league_id
            ORDER BY target_league_id
            """
        ).fetchall()

        return {
            str(row["target_league_id"]): int(row["total"])
            for row in rows
        }

    finally:
        connection.close()


def clean_text(
    value: Any,
) -> str:
    if value is None:
        return ""

    return str(value).strip()


def clean_optional_text(
    value: Any,
) -> str | None:
    cleaned = clean_text(value)

    return (
        cleaned.upper()
        if cleaned
        else None
    )


def clean_optional_text_preserve_case(
    value: Any,
) -> str | None:
    cleaned = clean_text(value)

    return (
        cleaned
        if cleaned
        else None
    )


def to_integer(
    value: Any,
    field_name: str,
    row_number: int,
    default: int | None = None,
) -> int:
    if value is None or value == "":
        if default is not None:
            return default

        raise PerformanceImportError(
            f"Linha {row_number}: "
            f"{field_name} vazio."
        )

    try:
        number = float(value)

    except (TypeError, ValueError) as exc:
        raise PerformanceImportError(
            f"Linha {row_number}: "
            f"{field_name} inválido."
        ) from exc

    if not math.isfinite(number):
        raise PerformanceImportError(
            f"Linha {row_number}: "
            f"{field_name} inválido."
        )

    if not number.is_integer():
        raise PerformanceImportError(
            f"Linha {row_number}: "
            f"{field_name} deve ser inteiro."
        )

    return int(number)


def to_non_negative_integer(
    value: Any,
    field_name: str,
    row_number: int,
    minimum: int = 0,
) -> int:
    number = to_integer(
        value=value,
        field_name=field_name,
        row_number=row_number,
    )

    if number < minimum:
        raise PerformanceImportError(
            f"Linha {row_number}: "
            f"{field_name} deve ser igual "
            f"ou superior a {minimum}."
        )

    return number


def to_binary_integer(
    value: Any,
    field_name: str,
    row_number: int,
) -> int:
    number = to_integer(
        value=value,
        field_name=field_name,
        row_number=row_number,
    )

    if number not in {0, 1}:
        raise PerformanceImportError(
            f"Linha {row_number}: "
            f"{field_name} deve ser 0 ou 1."
        )

    return number


def to_float(
    value: Any,
    field_name: str,
    row_number: int,
) -> float:
    try:
        number = float(value)

    except (TypeError, ValueError) as exc:
        raise PerformanceImportError(
            f"Linha {row_number}: "
            f"{field_name} inválido."
        ) from exc

    if not math.isfinite(number):
        raise PerformanceImportError(
            f"Linha {row_number}: "
            f"{field_name} inválido."
        )

    return number


def normalize_datetime_value(
    value: Any,
) -> str | None:
    if value is None or value == "":
        return None

    if hasattr(value, "isoformat"):
        try:
            return value.isoformat(
                sep=" ",
                timespec="seconds",
            )

        except TypeError:
            return value.isoformat()

    return str(value).strip()


def _assert_dataset_is_approved(
    dataset_version: str,
    dataset_path: Path,
) -> None:
    """
    Confirma que o dataset está aprovado e que o caminho
    corresponde ao ficheiro registado.
    """

    dataset = get_dataset_version(
        dataset_version
    )

    if dataset is None:
        raise PerformanceImportError(
            "A versão do dataset não existe: "
            f"{dataset_version}"
        )

    if dataset.status != "APPROVED":
        raise PerformanceImportError(
            f"O dataset {dataset_version} "
            "não está aprovado. "
            f"Estado atual: {dataset.status}"
        )

    if not dataset.file_path:
        raise PerformanceImportError(
            "O dataset não possui caminho registado."
        )

    registered_path = Path(
        dataset.file_path
    ).expanduser().resolve()

    if registered_path != dataset_path:
        raise PerformanceImportError(
            "O ficheiro indicado não corresponde "
            "ao dataset registado."
        )


def _resolve_dataset_path(
    dataset_path: str | Path | None,
) -> Path:
    """
    Resolve o caminho absoluto do dataset.
    """

    paths = load_paths_config()

    if dataset_path is None:
        path = (
            paths["data"]["input"]
            / "FOOTWIN_Dataset_2026_27_V001.xlsx"
        )

    else:
        path = Path(
            dataset_path
        ).expanduser()

        if not path.is_absolute():
            path = (
                paths["project_root"]
                / path
            )

        path = path.resolve()

    if not path.exists():
        raise PerformanceImportError(
            f"O dataset não existe: {path}"
        )

    if not path.is_file():
        raise PerformanceImportError(
            f"O caminho não é um ficheiro: {path}"
        )

    return path
