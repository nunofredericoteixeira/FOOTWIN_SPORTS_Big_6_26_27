# -*- coding: utf-8 -*-

from __future__ import annotations

import re
import sqlite3
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.config.league_config import get_active_leagues
from src.config.model_config import load_model_version
from src.config.path_config import load_paths_config
from src.database.dataset_versions import get_dataset_version
from src.database.init_database import connect_database
from src.utils.logger import get_logger


logger = get_logger("importers.teams")


TEAMS_SHEET_NAME = "Equipas_2026_27"

REQUIRED_HEADERS = [
    "team_id",
    "team_name",
    "short_name",
    "normalized_name",
    "league_id",
    "country",
    "season_label",
    "promoted",
    "promotion_method",
    "previous_division",
    "active",
]


@dataclass
class TeamImportResult:
    inserted: int = 0
    updated: int = 0
    unchanged: int = 0
    skipped: int = 0
    errors: int = 0

    @property
    def processed(self) -> int:
        return (
            self.inserted
            + self.updated
            + self.unchanged
            + self.skipped
            + self.errors
        )


class TeamImportError(RuntimeError):
    """Erro ocorrido durante a importação das equipas."""


def import_teams(
    dataset_path: str | Path | None = None,
    dataset_version: str | None = None,
    require_approved_dataset: bool = True,
    database_path: str | Path | None = None,
) -> TeamImportResult:
    """
    Importa as equipas do Excel para a tabela teams.

    Por defeito, só permite importar datasets com estado APPROVED.
    """

    model_config = load_model_version()

    final_dataset_version = (
        dataset_version.strip()
        if dataset_version
        else str(
            model_config["dataset"]["expected_version"]
        )
    )

    final_dataset_path = _resolve_dataset_path(
        dataset_path=dataset_path,
    )

    if require_approved_dataset:
        _assert_dataset_is_approved(
            dataset_version=final_dataset_version,
            dataset_path=final_dataset_path,
        )

    records = read_teams_from_excel(
        dataset_path=final_dataset_path,
    )

    if not records:
        raise TeamImportError(
            "A folha Equipas_2026_27 não contém equipas."
        )

    configured_leagues = get_active_leagues()

    result = TeamImportResult()

    connection = connect_database(database_path)

    logger.info(
        "A iniciar importação de equipas | "
        "dataset=%s | total=%s",
        final_dataset_version,
        len(records),
    )

    try:
        with connection:
            for row_number, record in records:
                try:
                    prepared = prepare_team_record(
                        record=record,
                        row_number=row_number,
                        dataset_version=final_dataset_version,
                        configured_leagues=configured_leagues,
                    )

                    action = upsert_team(
                        connection=connection,
                        team=prepared,
                    )

                    if action == "INSERTED":
                        result.inserted += 1
                    elif action == "UPDATED":
                        result.updated += 1
                    elif action == "UNCHANGED":
                        result.unchanged += 1
                    else:
                        result.skipped += 1

                except Exception as exc:
                    result.errors += 1

                    logger.error(
                        "Erro ao importar equipa | "
                        "linha=%s | team_id=%s | erro=%s",
                        row_number,
                        record.get("team_id"),
                        exc,
                    )

        if result.errors > 0:
            raise TeamImportError(
                "A importação de equipas encontrou "
                f"{result.errors} erro(s)."
            )

        validate_imported_teams(
            connection=connection,
            configured_leagues=configured_leagues,
            expected_total=len(records),
        )

    except sqlite3.Error as exc:
        raise TeamImportError(
            f"Erro SQLite durante a importação: {exc}"
        ) from exc

    finally:
        connection.close()

    logger.info(
        "Importação de equipas concluída | "
        "inseridas=%s | atualizadas=%s | "
        "inalteradas=%s | ignoradas=%s | erros=%s",
        result.inserted,
        result.updated,
        result.unchanged,
        result.skipped,
        result.errors,
    )

    return result


def read_teams_from_excel(
    dataset_path: str | Path,
) -> list[tuple[int, dict[str, Any]]]:
    """
    Lê a folha Equipas_2026_27.

    Devolve uma lista de pares:
        (número da linha, registo)
    """

    path = Path(dataset_path).expanduser().resolve()

    if not path.exists():
        raise TeamImportError(
            f"O dataset não existe: {path}"
        )

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    try:
        if TEAMS_SHEET_NAME not in workbook.sheetnames:
            raise TeamImportError(
                f"Falta a folha obrigatória: {TEAMS_SHEET_NAME}"
            )

        worksheet = workbook[TEAMS_SHEET_NAME]

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        actual_headers = headers[:len(REQUIRED_HEADERS)]

        if actual_headers != REQUIRED_HEADERS:
            raise TeamImportError(
                "Os cabeçalhos da folha Equipas_2026_27 "
                "não correspondem ao formato esperado."
            )

        records: list[tuple[int, dict[str, Any]]] = []

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):
            if all(value is None for value in row):
                continue

            record = {
                str(headers[index]): row[index]
                for index in range(len(headers))
            }

            records.append(
                (row_number, record)
            )

        return records

    finally:
        workbook.close()


def prepare_team_record(
    record: dict[str, Any],
    row_number: int,
    dataset_version: str,
    configured_leagues: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """
    Limpa e valida um registo de equipa antes da importação.
    """

    team_id = clean_text(
        record.get("team_id")
    )

    team_name = clean_text(
        record.get("team_name")
    )

    short_name = clean_text(
        record.get("short_name")
    )

    league_id = clean_text(
        record.get("league_id")
    ).upper()

    country = clean_text(
        record.get("country")
    )

    season_label = clean_text(
        record.get("season_label")
    )

    previous_division = clean_text(
        record.get("previous_division")
    )

    promoted = to_binary_integer(
        record.get("promoted"),
        field_name="promoted",
        row_number=row_number,
    )

    active = to_binary_integer(
        record.get("active"),
        field_name="active",
        row_number=row_number,
    )

    promotion_method = clean_optional_text(
        record.get("promotion_method")
    )

    normalized_name = clean_text(
        record.get("normalized_name")
    )

    if not normalized_name and team_name:
        normalized_name = normalize_team_name(
            team_name
        )

    if not team_id:
        raise TeamImportError(
            f"Linha {row_number}: team_id vazio."
        )

    if not re.fullmatch(
        r"[A-Z0-9_]+",
        team_id,
    ):
        raise TeamImportError(
            f"Linha {row_number}: team_id inválido: {team_id}"
        )

    if not team_name:
        raise TeamImportError(
            f"Linha {row_number}: team_name vazio."
        )

    if not short_name:
        raise TeamImportError(
            f"Linha {row_number}: short_name vazio."
        )

    if not normalized_name:
        raise TeamImportError(
            f"Linha {row_number}: normalized_name vazio."
        )

    if league_id not in configured_leagues:
        raise TeamImportError(
            f"Linha {row_number}: liga inválida: {league_id}"
        )

    expected_country = str(
        configured_leagues[league_id]["country"]
    )

    if country != expected_country:
        raise TeamImportError(
            f"Linha {row_number}: país incorreto para "
            f"{league_id}. Esperado={expected_country}; "
            f"encontrado={country}"
        )

    expected_season = str(
        configured_leagues[league_id]["season_label"]
    )

    if season_label != expected_season:
        raise TeamImportError(
            f"Linha {row_number}: época incorreta. "
            f"Esperado={expected_season}; "
            f"encontrado={season_label}"
        )

    if not previous_division:
        raise TeamImportError(
            f"Linha {row_number}: previous_division vazio."
        )

    if promoted == 1:
        if promotion_method not in {
            "CHAMPION",
            "DIRECT",
            "PLAYOFF",
        }:
            raise TeamImportError(
                f"Linha {row_number}: equipa promovida "
                "sem promotion_method válido."
            )
    else:
        promotion_method = None

    return {
        "team_id": team_id,
        "team_name": team_name,
        "short_name": short_name,
        "normalized_name": normalized_name,
        "league_id": league_id,
        "country": country,
        "season_label": season_label,
        "promoted": promoted,
        "promotion_method": promotion_method,
        "previous_division": previous_division,
        "active": active,
        "dataset_version": dataset_version,
    }


def upsert_team(
    connection: sqlite3.Connection,
    team: dict[str, Any],
) -> str:
    """
    Insere ou atualiza uma equipa.

    Devolve:
    - INSERTED
    - UPDATED
    - UNCHANGED
    """

    existing = connection.execute(
        """
        SELECT
            team_id,
            team_name,
            short_name,
            normalized_name,
            league_id,
            country,
            season_label,
            promoted,
            promotion_method,
            previous_division,
            active,
            dataset_version
        FROM teams
        WHERE team_id = ?
        """,
        (team["team_id"],),
    ).fetchone()

    if existing is None:
        connection.execute(
            """
            INSERT INTO teams (
                team_id,
                team_name,
                short_name,
                normalized_name,
                league_id,
                country,
                season_label,
                promoted,
                promotion_method,
                previous_division,
                active,
                dataset_version
            )
            VALUES (
                :team_id,
                :team_name,
                :short_name,
                :normalized_name,
                :league_id,
                :country,
                :season_label,
                :promoted,
                :promotion_method,
                :previous_division,
                :active,
                :dataset_version
            )
            """,
            team,
        )

        logger.info(
            "Equipa inserida | %s | %s",
            team["team_id"],
            team["team_name"],
        )

        return "INSERTED"

    if not team_has_changes(
        existing=existing,
        new_values=team,
    ):
        return "UNCHANGED"

    connection.execute(
        """
        UPDATE teams
        SET
            team_name = :team_name,
            short_name = :short_name,
            normalized_name = :normalized_name,
            league_id = :league_id,
            country = :country,
            season_label = :season_label,
            promoted = :promoted,
            promotion_method = :promotion_method,
            previous_division = :previous_division,
            active = :active,
            dataset_version = :dataset_version,
            updated_at = CURRENT_TIMESTAMP
        WHERE team_id = :team_id
        """,
        team,
    )

    logger.info(
        "Equipa atualizada | %s | %s",
        team["team_id"],
        team["team_name"],
    )

    return "UPDATED"


def team_has_changes(
    existing: sqlite3.Row,
    new_values: dict[str, Any],
) -> bool:
    fields = (
        "team_name",
        "short_name",
        "normalized_name",
        "league_id",
        "country",
        "season_label",
        "promoted",
        "promotion_method",
        "previous_division",
        "active",
        "dataset_version",
    )

    for field in fields:
        if existing[field] != new_values[field]:
            return True

    return False


def validate_imported_teams(
    connection: sqlite3.Connection,
    configured_leagues: dict[str, dict[str, Any]],
    expected_total: int,
) -> None:
    """
    Confirma a coerência das equipas importadas.
    """

    total_active = connection.execute(
        """
        SELECT COUNT(*) AS total
        FROM teams
        WHERE active = 1
        """
    ).fetchone()["total"]

    if int(total_active) != expected_total:
        raise TeamImportError(
            "Total de equipas ativas incorreto após importação. "
            f"Esperado={expected_total}; "
            f"encontrado={total_active}"
        )

    rows = connection.execute(
        """
        SELECT
            league_id,
            COUNT(*) AS total
        FROM teams
        WHERE active = 1
        GROUP BY league_id
        """
    ).fetchall()

    imported_counts = {
        str(row["league_id"]): int(row["total"])
        for row in rows
    }

    for league_id, league in configured_leagues.items():
        expected_count = int(
            league["team_count"]
        )

        actual_count = imported_counts.get(
            league_id,
            0,
        )

        if expected_total == 114 and actual_count != expected_count:
            raise TeamImportError(
                f"Número de equipas incorreto em {league_id}. "
                f"Esperado={expected_count}; "
                f"encontrado={actual_count}"
            )


def list_imported_teams(
    league_id: str | None = None,
    database_path: str | Path | None = None,
) -> list[dict[str, Any]]:
    """
    Lista as equipas existentes na SQLite.
    """

    connection = connect_database(
        database_path
    )

    try:
        if league_id:
            rows = connection.execute(
                """
                SELECT
                    team_id,
                    team_name,
                    short_name,
                    normalized_name,
                    league_id,
                    country,
                    season_label,
                    promoted,
                    promotion_method,
                    previous_division,
                    active,
                    dataset_version
                FROM teams
                WHERE league_id = ?
                ORDER BY team_name
                """,
                (league_id.strip().upper(),),
            ).fetchall()
        else:
            rows = connection.execute(
                """
                SELECT
                    team_id,
                    team_name,
                    short_name,
                    normalized_name,
                    league_id,
                    country,
                    season_label,
                    promoted,
                    promotion_method,
                    previous_division,
                    active,
                    dataset_version
                FROM teams
                ORDER BY league_id, team_name
                """
            ).fetchall()

        return [
            dict(row)
            for row in rows
        ]

    finally:
        connection.close()


def count_imported_teams(
    database_path: str | Path | None = None,
) -> dict[str, int]:
    """
    Conta as equipas por liga.
    """

    connection = connect_database(
        database_path
    )

    try:
        rows = connection.execute(
            """
            SELECT
                league_id,
                COUNT(*) AS total
            FROM teams
            WHERE active = 1
            GROUP BY league_id
            ORDER BY league_id
            """
        ).fetchall()

        return {
            str(row["league_id"]): int(row["total"])
            for row in rows
        }

    finally:
        connection.close()


def normalize_team_name(
    value: str,
) -> str:
    """
    Normaliza um nome de equipa.

    Exemplo:
        Vitória SC -> vitoria_sc
    """

    normalized = unicodedata.normalize(
        "NFKD",
        value,
    )

    without_accents = "".join(
        character
        for character in normalized
        if not unicodedata.combining(character)
    )

    lowered = without_accents.lower().strip()

    cleaned = re.sub(
        r"[^a-z0-9]+",
        "_",
        lowered,
    )

    return cleaned.strip("_")


def clean_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def clean_optional_text(
    value: Any,
) -> str | None:
    cleaned = clean_text(value)

    return cleaned.upper() if cleaned else None


def to_binary_integer(
    value: Any,
    field_name: str,
    row_number: int,
) -> int:
    if value in {0, 0.0, "0", False}:
        return 0

    if value in {1, 1.0, "1", True}:
        return 1

    raise TeamImportError(
        f"Linha {row_number}: {field_name} deve ser 0 ou 1."
    )


def _assert_dataset_is_approved(
    dataset_version: str,
    dataset_path: Path,
) -> None:
    """
    Confirma que o dataset está aprovado e corresponde ao ficheiro registado.
    """

    dataset = get_dataset_version(
        dataset_version
    )

    if dataset is None:
        raise TeamImportError(
            "A versão do dataset ainda não está registada: "
            f"{dataset_version}"
        )

    if dataset.status != "APPROVED":
        raise TeamImportError(
            f"O dataset {dataset_version} não está aprovado. "
            f"Estado atual: {dataset.status}"
        )

    if not dataset.file_path:
        raise TeamImportError(
            "A versão do dataset não possui caminho registado."
        )

    registered_path = Path(
        dataset.file_path
    ).expanduser().resolve()

    if registered_path != dataset_path:
        raise TeamImportError(
            "O ficheiro indicado não corresponde ao ficheiro "
            "registado para esta versão do dataset."
        )


def _resolve_dataset_path(
    dataset_path: str | Path | None,
) -> Path:
    paths = load_paths_config()

    if dataset_path is None:
        path = (
            paths["data"]["input"]
            / "FOOTWIN_Dataset_2026_27_V001.xlsx"
        )
    else:
        path = Path(
            dataset_path
        ).expanduser()

        if not path.is_absolute():
            path = (
                paths["project_root"]
                / path
            )

        path = path.resolve()

    if not path.exists():
        raise TeamImportError(
            f"O dataset não existe: {path}"
        )

    if not path.is_file():
        raise TeamImportError(
            f"O caminho não é um ficheiro: {path}"
        )

    return path
