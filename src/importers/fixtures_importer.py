# -*- coding: utf-8 -*-

from __future__ import annotations

import math
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.config.model_config import load_model_version
from src.config.path_config import load_paths_config
from src.database.dataset_versions import get_dataset_version
from src.database.init_database import connect_database
from src.utils.logger import get_logger


logger = get_logger("importers.fixtures")


FIXTURES_SHEET_NAME = "Calendario_2026_27"

REQUIRED_HEADERS = [
    "match_id",
    "league_id",
    "season_label",
    "round_number",
    "match_date",
    "home_team_id",
    "away_team_id",
    "status",
    "home_goals",
    "away_goals",
    "schedule_type",
    "source_url",
]

VALID_MATCH_STATUSES = {
    "SCHEDULED",
    "PLAYED",
    "POSTPONED",
    "CANCELLED",
    "ABANDONED",
    "AWARDED",
}

VALID_SCHEDULE_TYPES = {
    "OFFICIAL",
    "SYNTHETIC",
}


@dataclass
class FixturesImportResult:
    inserted: int = 0
    updated: int = 0
    unchanged: int = 0
    skipped: int = 0
    errors: int = 0

    @property
    def processed(self) -> int:
        return (
            self.inserted
            + self.updated
            + self.unchanged
            + self.skipped
            + self.errors
        )


class FixturesImportError(RuntimeError):
    """Erro ocorrido durante a importação do calendário."""


def import_fixtures(
    dataset_path: str | Path | None = None,
    dataset_version: str | None = None,
    require_approved_dataset: bool = True,
    database_path: str | Path | None = None,
) -> FixturesImportResult:
    """
    Importa a folha Calendario_2026_27 para a SQLite.

    A operação é atómica:
    - todos os registos são validados primeiro;
    - só depois são gravados;
    - qualquer erro provoca rollback total.
    """

    model_config = load_model_version()

    final_dataset_version = (
        dataset_version.strip()
        if dataset_version
        else str(model_config["dataset"]["expected_version"])
    )

    final_dataset_path = _resolve_dataset_path(dataset_path)

    if require_approved_dataset:
        _assert_dataset_is_approved(
            dataset_version=final_dataset_version,
            dataset_path=final_dataset_path,
        )

    records = read_fixtures_from_excel(final_dataset_path)

    if not records:
        raise FixturesImportError(
            "A folha Calendario_2026_27 não contém jogos."
        )

    result = FixturesImportResult()
    connection = connect_database(database_path)

    logger.info(
        "A iniciar importação do calendário | dataset=%s | total=%s",
        final_dataset_version,
        len(records),
    )

    prepared_records: list[dict[str, Any]] = []
    preparation_errors: list[str] = []

    try:
        for row_number, record in records:
            try:
                prepared = prepare_fixture_record(
                    connection=connection,
                    record=record,
                    row_number=row_number,
                    dataset_version=final_dataset_version,
                )

                prepared_records.append(prepared)

            except Exception as exc:
                result.errors += 1

                error_message = (
                    f"Linha {row_number} | "
                    f"match_id={record.get('match_id')} | "
                    f"erro={exc}"
                )

                preparation_errors.append(error_message)

                logger.error(
                    "Erro ao preparar jogo | %s",
                    error_message,
                )

        _validate_duplicate_records(prepared_records)

        if preparation_errors:
            details = "\n".join(
                f" - {message}"
                for message in preparation_errors
            )

            raise FixturesImportError(
                "A importação foi cancelada antes da gravação. "
                f"Foram encontradas {result.errors} linha(s) inválida(s):\n"
                f"{details}"
            )

        try:
            connection.execute("BEGIN IMMEDIATE")

            for prepared in prepared_records:
                action = upsert_fixture(
                    connection=connection,
                    fixture=prepared,
                )

                if action == "INSERTED":
                    result.inserted += 1
                elif action == "UPDATED":
                    result.updated += 1
                elif action == "UNCHANGED":
                    result.unchanged += 1
                else:
                    result.skipped += 1

            validate_imported_fixtures(
                connection=connection,
                dataset_version=final_dataset_version,
                expected_total=len(records),
            )

            connection.commit()

        except Exception:
            connection.rollback()

            logger.exception(
                "Importação do calendário revertida integralmente | dataset=%s",
                final_dataset_version,
            )

            raise

    except FixturesImportError:
        raise

    except sqlite3.Error as exc:
        raise FixturesImportError(
            f"Erro SQLite durante a importação: {exc}"
        ) from exc

    except Exception as exc:
        raise FixturesImportError(
            f"Erro durante a importação do calendário: {exc}"
        ) from exc

    finally:
        connection.close()

    logger.info(
        "Importação do calendário concluída | "
        "inseridos=%s | atualizados=%s | "
        "inalterados=%s | ignorados=%s | erros=%s",
        result.inserted,
        result.updated,
        result.unchanged,
        result.skipped,
        result.errors,
    )

    return result


def read_fixtures_from_excel(
    dataset_path: str | Path,
) -> list[tuple[int, dict[str, Any]]]:
    """Lê os jogos da folha Calendario_2026_27."""

    path = Path(dataset_path).expanduser().resolve()

    if not path.exists():
        raise FixturesImportError(
            f"O dataset não existe: {path}"
        )

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    try:
        if FIXTURES_SHEET_NAME not in workbook.sheetnames:
            raise FixturesImportError(
                f"Falta a folha obrigatória: {FIXTURES_SHEET_NAME}"
            )

        worksheet = workbook[FIXTURES_SHEET_NAME]

        headers = [cell.value for cell in worksheet[1]]

        actual_headers = headers[:len(REQUIRED_HEADERS)]

        if actual_headers != REQUIRED_HEADERS:
            raise FixturesImportError(
                "Os cabeçalhos da folha Calendario_2026_27 "
                "não correspondem ao formato esperado."
            )

        records: list[tuple[int, dict[str, Any]]] = []

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):
            if all(value is None for value in row):
                continue

            record = {
                str(headers[index]): row[index]
                for index in range(len(headers))
            }

            records.append((row_number, record))

        return records

    finally:
        workbook.close()


def prepare_fixture_record(
    connection: sqlite3.Connection,
    record: dict[str, Any],
    row_number: int,
    dataset_version: str,
) -> dict[str, Any]:
    """Limpa, valida e prepara um jogo."""

    match_id = clean_text(record.get("match_id"))
    league_id = clean_text(record.get("league_id")).upper()
    season_label = clean_text(record.get("season_label"))

    round_number = to_optional_integer(
        value=record.get("round_number"),
        field_name="round_number",
        row_number=row_number,
        minimum=1,
    )

    match_date = normalize_datetime_value(
        record.get("match_date")
    )

    home_team_id = clean_text(
        record.get("home_team_id")
    )

    away_team_id = clean_text(
        record.get("away_team_id")
    )

    status = clean_text(
        record.get("status")
    ).upper()

    home_goals = to_optional_integer(
        value=record.get("home_goals"),
        field_name="home_goals",
        row_number=row_number,
        minimum=0,
    )

    away_goals = to_optional_integer(
        value=record.get("away_goals"),
        field_name="away_goals",
        row_number=row_number,
        minimum=0,
    )

    schedule_type = clean_text(
        record.get("schedule_type")
    ).upper()

    source_url = clean_optional_text(
        record.get("source_url")
    )

    if not match_id:
        raise FixturesImportError(
            f"Linha {row_number}: match_id vazio."
        )

    if not league_id:
        raise FixturesImportError(
            f"Linha {row_number}: league_id vazio."
        )

    league = connection.execute(
        """
        SELECT league_id, season_label
        FROM leagues
        WHERE league_id = ?
          AND active = 1
        """,
        (league_id,),
    ).fetchone()

    if league is None:
        raise FixturesImportError(
            f"Linha {row_number}: liga inexistente ou inativa: {league_id}"
        )

    if season_label != str(league["season_label"]):
        raise FixturesImportError(
            f"Linha {row_number}: época incorreta. "
            f"Esperado={league['season_label']}; "
            f"encontrado={season_label}"
        )

    if not home_team_id:
        raise FixturesImportError(
            f"Linha {row_number}: home_team_id vazio."
        )

    if not away_team_id:
        raise FixturesImportError(
            f"Linha {row_number}: away_team_id vazio."
        )

    if home_team_id == away_team_id:
        raise FixturesImportError(
            f"Linha {row_number}: uma equipa não pode jogar contra si própria."
        )

    home_team = connection.execute(
        """
        SELECT team_id, league_id
        FROM teams
        WHERE team_id = ?
          AND active = 1
        """,
        (home_team_id,),
    ).fetchone()

    if home_team is None:
        raise FixturesImportError(
            f"Linha {row_number}: equipa da casa inexistente: {home_team_id}"
        )

    away_team = connection.execute(
        """
        SELECT team_id, league_id
        FROM teams
        WHERE team_id = ?
          AND active = 1
        """,
        (away_team_id,),
    ).fetchone()

    if away_team is None:
        raise FixturesImportError(
            f"Linha {row_number}: equipa visitante inexistente: {away_team_id}"
        )

    if home_team["league_id"] != league_id:
        raise FixturesImportError(
            f"Linha {row_number}: equipa da casa pertence a "
            f"{home_team['league_id']} e não a {league_id}."
        )

    if away_team["league_id"] != league_id:
        raise FixturesImportError(
            f"Linha {row_number}: equipa visitante pertence a "
            f"{away_team['league_id']} e não a {league_id}."
        )

    if status not in VALID_MATCH_STATUSES:
        raise FixturesImportError(
            f"Linha {row_number}: estado inválido: {status}"
        )

    if schedule_type not in VALID_SCHEDULE_TYPES:
        raise FixturesImportError(
            f"Linha {row_number}: schedule_type inválido: {schedule_type}"
        )

    if status in {"PLAYED", "AWARDED"}:
        if home_goals is None or away_goals is None:
            raise FixturesImportError(
                f"Linha {row_number}: um jogo {status} "
                "deve possuir resultado."
            )
    else:
        if home_goals is not None or away_goals is not None:
            raise FixturesImportError(
                f"Linha {row_number}: jogos com estado {status} "
                "não devem possuir resultado."
            )

    return {
        "match_id": match_id,
        "league_id": league_id,
        "season_label": season_label,
        "round_number": round_number,
        "match_date": match_date,
        "home_team_id": home_team_id,
        "away_team_id": away_team_id,
        "status": status,
        "home_goals": home_goals,
        "away_goals": away_goals,
        "schedule_type": schedule_type,
        "source_url": source_url,
        "dataset_version": dataset_version,
    }


def _validate_duplicate_records(
    prepared_records: list[dict[str, Any]],
) -> None:
    """Deteta duplicações dentro do próprio Excel."""

    match_ids: set[str] = set()
    pairings: set[tuple[str, str, str]] = set()

    duplicate_match_ids: list[str] = []
    duplicate_pairings: list[str] = []

    for fixture in prepared_records:
        match_id = fixture["match_id"]

        if match_id in match_ids:
            duplicate_match_ids.append(match_id)

        match_ids.add(match_id)

        pairing = (
            fixture["league_id"],
            fixture["home_team_id"],
            fixture["away_team_id"],
        )

        if pairing in pairings:
            duplicate_pairings.append(
                f"{fixture['home_team_id']} vs "
                f"{fixture['away_team_id']}"
            )

        pairings.add(pairing)

    if duplicate_match_ids:
        raise FixturesImportError(
            "Existem match_id duplicados no Excel: "
            + ", ".join(sorted(set(duplicate_match_ids)))
        )

    if duplicate_pairings:
        raise FixturesImportError(
            "Existem emparelhamentos casa/fora duplicados no Excel: "
            + ", ".join(sorted(set(duplicate_pairings)))
        )


def upsert_fixture(
    connection: sqlite3.Connection,
    fixture: dict[str, Any],
) -> str:
    """Insere ou atualiza um jogo."""

    existing = connection.execute(
        """
        SELECT *
        FROM matches
        WHERE match_id = ?
        """,
        (fixture["match_id"],),
    ).fetchone()

    if existing is None:
        connection.execute(
            """
            INSERT INTO matches (
                match_id,
                league_id,
                season_label,
                round_number,
                match_date,
                home_team_id,
                away_team_id,
                status,
                home_goals,
                away_goals,
                schedule_type,
                source_url,
                dataset_version
            )
            VALUES (
                :match_id,
                :league_id,
                :season_label,
                :round_number,
                :match_date,
                :home_team_id,
                :away_team_id,
                :status,
                :home_goals,
                :away_goals,
                :schedule_type,
                :source_url,
                :dataset_version
            )
            """,
            fixture,
        )

        logger.info(
            "Jogo inserido | %s | %s vs %s",
            fixture["match_id"],
            fixture["home_team_id"],
            fixture["away_team_id"],
        )

        return "INSERTED"

    merged_fixture = dict(fixture)

    final_statuses = {
        "PLAYED",
        "AWARDED",
        "CANCELLED",
        "ABANDONED",
    }

    if existing["status"] in final_statuses:
        merged_fixture["status"] = existing["status"]
        merged_fixture["home_goals"] = existing["home_goals"]
        merged_fixture["away_goals"] = existing["away_goals"]

    if not fixture_has_changes(
        existing=existing,
        new_values=merged_fixture,
    ):
        return "UNCHANGED"

    connection.execute(
        """
        UPDATE matches
        SET
            league_id = :league_id,
            season_label = :season_label,
            round_number = :round_number,
            match_date = :match_date,
            home_team_id = :home_team_id,
            away_team_id = :away_team_id,
            status = :status,
            home_goals = :home_goals,
            away_goals = :away_goals,
            schedule_type = :schedule_type,
            source_url = :source_url,
            dataset_version = :dataset_version,
            updated_at = CURRENT_TIMESTAMP
        WHERE match_id = :match_id
        """,
        merged_fixture,
    )

    logger.info(
        "Jogo atualizado | %s | %s vs %s",
        merged_fixture["match_id"],
        merged_fixture["home_team_id"],
        merged_fixture["away_team_id"],
    )

    return "UPDATED"


def fixture_has_changes(
    existing: sqlite3.Row,
    new_values: dict[str, Any],
) -> bool:
    """Confirma se um jogo sofreu alterações."""

    fields = (
        "league_id",
        "season_label",
        "round_number",
        "match_date",
        "home_team_id",
        "away_team_id",
        "status",
        "home_goals",
        "away_goals",
        "schedule_type",
        "source_url",
        "dataset_version",
    )

    return any(
        existing[field] != new_values[field]
        for field in fields
    )


def validate_imported_fixtures(
    connection: sqlite3.Connection,
    dataset_version: str,
    expected_total: int,
) -> None:
    """Confirma o total de jogos importados."""

    total = connection.execute(
        """
        SELECT COUNT(*) AS total
        FROM matches
        WHERE dataset_version = ?
        """,
        (dataset_version,),
    ).fetchone()["total"]

    if int(total) != expected_total:
        raise FixturesImportError(
            "Total de jogos incorreto após importação. "
            f"Esperado={expected_total}; encontrado={total}"
        )

    duplicate_pairings = connection.execute(
        """
        SELECT
            league_id,
            home_team_id,
            away_team_id,
            COUNT(*) AS total
        FROM matches
        WHERE dataset_version = ?
        GROUP BY
            league_id,
            home_team_id,
            away_team_id
        HAVING COUNT(*) > 1
        """,
        (dataset_version,),
    ).fetchall()

    if duplicate_pairings:
        raise FixturesImportError(
            "Foram encontrados jogos casa/fora duplicados."
        )


def list_imported_fixtures(
    dataset_version: str | None = None,
    database_path: str | Path | None = None,
) -> list[dict[str, Any]]:
    """Lista os jogos importados."""

    connection = connect_database(database_path)

    try:
        if dataset_version:
            rows = connection.execute(
                """
                SELECT
                    m.*,
                    ht.team_name AS home_team_name,
                    at.team_name AS away_team_name
                FROM matches m
                INNER JOIN teams ht
                    ON ht.team_id = m.home_team_id
                INNER JOIN teams at
                    ON at.team_id = m.away_team_id
                WHERE m.dataset_version = ?
                ORDER BY
                    m.league_id,
                    m.round_number,
                    m.match_date,
                    m.match_id
                """,
                (dataset_version,),
            ).fetchall()
        else:
            rows = connection.execute(
                """
                SELECT
                    m.*,
                    ht.team_name AS home_team_name,
                    at.team_name AS away_team_name
                FROM matches m
                INNER JOIN teams ht
                    ON ht.team_id = m.home_team_id
                INNER JOIN teams at
                    ON at.team_id = m.away_team_id
                ORDER BY
                    m.league_id,
                    m.round_number,
                    m.match_date,
                    m.match_id
                """
            ).fetchall()

        return [dict(row) for row in rows]

    finally:
        connection.close()


def count_imported_fixtures(
    database_path: str | Path | None = None,
) -> dict[str, int]:
    """Conta os jogos por liga."""

    connection = connect_database(database_path)

    try:
        rows = connection.execute(
            """
            SELECT
                league_id,
                COUNT(*) AS total
            FROM matches
            GROUP BY league_id
            ORDER BY league_id
            """
        ).fetchall()

        return {
            str(row["league_id"]): int(row["total"])
            for row in rows
        }

    finally:
        connection.close()


def clean_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def clean_optional_text(value: Any) -> str | None:
    cleaned = clean_text(value)

    return cleaned if cleaned else None


def to_optional_integer(
    value: Any,
    field_name: str,
    row_number: int,
    minimum: int | None = None,
) -> int | None:
    if value is None or value == "":
        return None

    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise FixturesImportError(
            f"Linha {row_number}: {field_name} inválido."
        ) from exc

    if not math.isfinite(number) or not number.is_integer():
        raise FixturesImportError(
            f"Linha {row_number}: {field_name} deve ser inteiro."
        )

    integer = int(number)

    if minimum is not None and integer < minimum:
        raise FixturesImportError(
            f"Linha {row_number}: {field_name} deve ser "
            f"igual ou superior a {minimum}."
        )

    return integer


def normalize_datetime_value(value: Any) -> str | None:
    if value is None or value == "":
        return None

    if hasattr(value, "isoformat"):
        try:
            return value.isoformat(
                sep=" ",
                timespec="seconds",
            )
        except TypeError:
            return value.isoformat()

    return str(value).strip()


def _assert_dataset_is_approved(
    dataset_version: str,
    dataset_path: Path,
) -> None:
    dataset = get_dataset_version(dataset_version)

    if dataset is None:
        raise FixturesImportError(
            f"A versão do dataset não existe: {dataset_version}"
        )

    if dataset.status != "APPROVED":
        raise FixturesImportError(
            f"O dataset {dataset_version} não está aprovado. "
            f"Estado atual: {dataset.status}"
        )

    if not dataset.file_path:
        raise FixturesImportError(
            "O dataset não possui caminho registado."
        )

    registered_path = Path(
        dataset.file_path
    ).expanduser().resolve()

    if registered_path != dataset_path:
        raise FixturesImportError(
            "O ficheiro indicado não corresponde ao dataset registado."
        )


def _resolve_dataset_path(
    dataset_path: str | Path | None,
) -> Path:
    paths = load_paths_config()

    if dataset_path is None:
        path = (
            paths["data"]["input"]
            / "FOOTWIN_Dataset_2026_27_V001.xlsx"
        )
    else:
        path = Path(dataset_path).expanduser()

        if not path.is_absolute():
            path = paths["project_root"] / path

        path = path.resolve()

    if not path.exists():
        raise FixturesImportError(
            f"O dataset não existe: {path}"
        )

    if not path.is_file():
        raise FixturesImportError(
            f"O caminho não é um ficheiro: {path}"
        )

    return path
