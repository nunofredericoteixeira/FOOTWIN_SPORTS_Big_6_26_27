# -*- coding: utf-8 -*-

from __future__ import annotations

import math
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from src.config.league_config import get_active_leagues
from src.config.model_config import load_model_weights
from src.config.path_config import load_paths_config
from src.utils.logger import get_logger


logger = get_logger("validation.dataset")


EXPECTED_SHEETS = {
    "Resumo",
    "Ligas",
    "Equipas_2026_27",
    "Desempenho_2025_26",
    "Promovidas",
    "Calendario_2026_27",
    "Fontes",
    "Mapeamento_IDs",
    "Validacao",
}


EXPECTED_HEADERS: dict[str, list[str]] = {
    "Resumo": [
        "Indicador",
        "Valor",
        "Estado",
        "Observações",
    ],
    "Ligas": [
        "league_id",
        "league_name",
        "country",
        "country_code",
        "season_label",
        "team_count",
        "matches_per_team",
        "total_matches",
        "league_strength_factor",
        "relegation_places",
        "playoff_places",
        "active",
    ],
    "Equipas_2026_27": [
        "team_id",
        "team_name",
        "short_name",
        "normalized_name",
        "league_id",
        "country",
        "season_label",
        "promoted",
        "promotion_method",
        "previous_division",
        "active",
    ],
    "Desempenho_2025_26": [
        "team_id",
        "source_league_id",
        "target_league_id",
        "season_label",
        "position",
        "played",
        "wins",
        "draws",
        "losses",
        "goals_for",
        "goals_against",
        "goal_difference",
        "points",
        "points_adjustment",
        "promoted",
        "promotion_method",
        "source_status",
        "data_confidence",
        "source_url",
        "accessed_at",
    ],
    "Promovidas": [
        "team_id",
        "target_league_id",
        "source_league_id",
        "source_position",
        "promotion_method",
        "played",
        "points",
        "goals_for",
        "goals_against",
        "goal_difference",
        "promotion_factor",
        "source_status",
        "data_confidence",
        "source_url",
    ],
    "Calendario_2026_27": [
        "match_id",
        "league_id",
        "season_label",
        "round_number",
        "match_date",
        "home_team_id",
        "away_team_id",
        "status",
        "home_goals",
        "away_goals",
        "schedule_type",
        "source_url",
    ],
    "Fontes": [
        "source_record_id",
        "entity_type",
        "entity_id",
        "data_type",
        "source_name",
        "source_url",
        "season_label",
        "accessed_at",
        "source_status",
        "notes",
    ],
    "Mapeamento_IDs": [
        "source_name",
        "source_entity_id",
        "source_entity_name",
        "internal_entity_type",
        "internal_entity_id",
        "match_status",
        "confidence",
        "notes",
    ],
    "Validacao": [
        "issue_id",
        "severity",
        "entity_type",
        "entity_id",
        "field_name",
        "expected_value",
        "actual_value",
        "message",
        "resolved",
        "resolution_note",
    ],
}


@dataclass
class ValidationIssue:
    severity: str
    entity_type: str
    entity_id: str | None
    field_name: str | None
    expected_value: Any
    actual_value: Any
    message: str


@dataclass
class ValidationResult:
    dataset_path: Path
    issues: list[ValidationIssue] = field(default_factory=list)
    counts: dict[str, int] = field(default_factory=dict)

    @property
    def error_count(self) -> int:
        return sum(
            1
            for issue in self.issues
            if issue.severity == "ERROR"
        )

    @property
    def warning_count(self) -> int:
        return sum(
            1
            for issue in self.issues
            if issue.severity == "WARNING"
        )

    @property
    def info_count(self) -> int:
        return sum(
            1
            for issue in self.issues
            if issue.severity == "INFO"
        )

    @property
    def approved(self) -> bool:
        return self.error_count == 0

    @property
    def status(self) -> str:
        return "APPROVED" if self.approved else "REJECTED"


def validate_dataset(
    dataset_path: str | Path | None = None,
) -> ValidationResult:
    """
    Executa a validação estrutural e aritmética do dataset Excel.
    """

    paths = load_paths_config()

    if dataset_path is None:
        path = (
            paths["data"]["input"]
            / "FOOTWIN_Dataset_2026_27_V001.xlsx"
        )
    else:
        path = Path(dataset_path).expanduser().resolve()

    result = ValidationResult(dataset_path=path)

    if not path.exists():
        result.issues.append(
            ValidationIssue(
                severity="ERROR",
                entity_type="DATASET",
                entity_id=None,
                field_name="file_path",
                expected_value=str(path),
                actual_value=None,
                message="O ficheiro do dataset não existe.",
            )
        )
        return result

    logger.info(
        "A iniciar validação do dataset | ficheiro=%s",
        path,
    )

    workbook = load_workbook(
        filename=path,
        data_only=False,
        read_only=False,
    )

    _validate_sheet_names(workbook, result)
    _validate_headers(workbook, result)

    if result.error_count > 0:
        workbook.close()
        return result

    leagues = _read_sheet_records(workbook["Ligas"])
    teams = _read_sheet_records(workbook["Equipas_2026_27"])
    performance = _read_sheet_records(
        workbook["Desempenho_2025_26"]
    )
    promoted = _read_sheet_records(workbook["Promovidas"])
    fixtures = _read_sheet_records(
        workbook["Calendario_2026_27"]
    )

    result.counts = {
        "leagues": len(leagues),
        "teams": len(teams),
        "performance": len(performance),
        "promoted": len(promoted),
        "fixtures": len(fixtures),
    }

    _validate_leagues(leagues, result)
    _validate_teams(teams, result)
    _validate_performance(performance, teams, result)
    _validate_promoted(promoted, teams, result)
    _validate_fixtures(fixtures, teams, result)
    _validate_expected_totals(result)
    _validate_cross_references(
        teams=teams,
        performance=performance,
        promoted=promoted,
        fixtures=fixtures,
        result=result,
    )

    workbook.close()

    logger.info(
        "Validação concluída | estado=%s | erros=%s | avisos=%s",
        result.status,
        result.error_count,
        result.warning_count,
    )

    return result


def _validate_sheet_names(workbook, result: ValidationResult) -> None:
    existing = set(workbook.sheetnames)

    for sheet_name in sorted(EXPECTED_SHEETS - existing):
        result.issues.append(
            ValidationIssue(
                severity="ERROR",
                entity_type="WORKBOOK",
                entity_id=sheet_name,
                field_name="sheet_name",
                expected_value=sheet_name,
                actual_value=None,
                message=f"Falta a folha obrigatória '{sheet_name}'.",
            )
        )

    for sheet_name in sorted(existing - EXPECTED_SHEETS):
        result.issues.append(
            ValidationIssue(
                severity="WARNING",
                entity_type="WORKBOOK",
                entity_id=sheet_name,
                field_name="sheet_name",
                expected_value=None,
                actual_value=sheet_name,
                message=f"Foi encontrada uma folha adicional: '{sheet_name}'.",
            )
        )


def _validate_headers(workbook, result: ValidationResult) -> None:
    for sheet_name, expected_headers in EXPECTED_HEADERS.items():
        if sheet_name not in workbook.sheetnames:
            continue

        worksheet = workbook[sheet_name]

        actual_headers = [
            worksheet.cell(row=1, column=column).value
            for column in range(1, len(expected_headers) + 1)
        ]

        if actual_headers != expected_headers:
            result.issues.append(
                ValidationIssue(
                    severity="ERROR",
                    entity_type="SHEET",
                    entity_id=sheet_name,
                    field_name="headers",
                    expected_value=expected_headers,
                    actual_value=actual_headers,
                    message=(
                        f"Os cabeçalhos da folha '{sheet_name}' "
                        "não correspondem ao formato esperado."
                    ),
                )
            )


def _read_sheet_records(worksheet) -> list[dict[str, Any]]:
    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    records: list[dict[str, Any]] = []

    for row in worksheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        if all(value is None for value in row):
            continue

        record = {
            str(headers[index]): row[index]
            for index in range(len(headers))
        }

        records.append(record)

    return records


def _validate_leagues(
    leagues: list[dict[str, Any]],
    result: ValidationResult,
) -> None:
    configured = get_active_leagues()

    seen_ids: set[str] = set()

    for row_number, league in enumerate(leagues, start=2):
        league_id = _as_text(league.get("league_id"))

        if not league_id:
            _add_error(
                result,
                "LEAGUE",
                f"ROW_{row_number}",
                "league_id",
                "valor preenchido",
                league.get("league_id"),
                "A liga não possui league_id.",
            )
            continue

        if league_id in seen_ids:
            _add_error(
                result,
                "LEAGUE",
                league_id,
                "league_id",
                "valor único",
                league_id,
                "O league_id está duplicado.",
            )

        seen_ids.add(league_id)

        if league_id not in configured:
            _add_error(
                result,
                "LEAGUE",
                league_id,
                "league_id",
                sorted(configured),
                league_id,
                "A liga não existe na configuração oficial.",
            )
            continue

        expected = configured[league_id]

        comparisons = {
            "league_name": expected["name"],
            "country": expected["country"],
            "country_code": expected["country_code"],
            "season_label": expected["season_label"],
            "team_count": expected["team_count"],
            "matches_per_team": expected["matches_per_team"],
            "total_matches": expected["total_matches"],
        }

        for field_name, expected_value in comparisons.items():
            actual_value = league.get(field_name)

            if actual_value != expected_value:
                _add_error(
                    result,
                    "LEAGUE",
                    league_id,
                    field_name,
                    expected_value,
                    actual_value,
                    f"Valor incorreto em {field_name}.",
                )


def _validate_teams(
    teams: list[dict[str, Any]],
    result: ValidationResult,
) -> None:
    configured_leagues = get_active_leagues()

    team_ids: set[str] = set()
    normalized_per_league: set[tuple[str, str]] = set()
    counts_per_league: dict[str, int] = {}

    for row_number, team in enumerate(teams, start=2):
        team_id = _as_text(team.get("team_id"))
        league_id = _as_text(team.get("league_id"))
        normalized_name = _as_text(team.get("normalized_name"))

        if not team_id:
            _add_error(
                result,
                "TEAM",
                f"ROW_{row_number}",
                "team_id",
                "valor preenchido",
                team.get("team_id"),
                "A equipa não possui team_id.",
            )
            continue

        if team_id in team_ids:
            _add_error(
                result,
                "TEAM",
                team_id,
                "team_id",
                "valor único",
                team_id,
                "O team_id está duplicado.",
            )

        team_ids.add(team_id)

        if league_id not in configured_leagues:
            _add_error(
                result,
                "TEAM",
                team_id,
                "league_id",
                sorted(configured_leagues),
                league_id,
                "A equipa possui uma liga inválida.",
            )
        else:
            counts_per_league[league_id] = (
                counts_per_league.get(league_id, 0) + 1
            )

        if not _as_text(team.get("team_name")):
            _add_error(
                result,
                "TEAM",
                team_id,
                "team_name",
                "valor preenchido",
                team.get("team_name"),
                "O nome da equipa está vazio.",
            )

        if not normalized_name:
            _add_error(
                result,
                "TEAM",
                team_id,
                "normalized_name",
                "valor preenchido",
                team.get("normalized_name"),
                "O normalized_name está vazio.",
            )
        elif league_id:
            key = (league_id, normalized_name)

            if key in normalized_per_league:
                _add_error(
                    result,
                    "TEAM",
                    team_id,
                    "normalized_name",
                    "valor único na liga",
                    normalized_name,
                    "O normalized_name está duplicado na liga.",
                )

            normalized_per_league.add(key)

        promoted = _as_int(team.get("promoted"))

        if promoted not in {0, 1}:
            _add_error(
                result,
                "TEAM",
                team_id,
                "promoted",
                "0 ou 1",
                team.get("promoted"),
                "O campo promoted é inválido.",
            )

        promotion_method = _as_text(
            team.get("promotion_method")
        )

        if promoted == 1 and promotion_method not in {
            "CHAMPION",
            "DIRECT",
            "PLAYOFF",
        }:
            _add_error(
                result,
                "TEAM",
                team_id,
                "promotion_method",
                "CHAMPION, DIRECT ou PLAYOFF",
                promotion_method,
                "A equipa promovida não possui método válido.",
            )

        if promoted == 0 and promotion_method:
            _add_warning(
                result,
                "TEAM",
                team_id,
                "promotion_method",
                None,
                promotion_method,
                "Uma equipa não promovida possui método de promoção.",
            )

    for league_id, league in configured_leagues.items():
        actual_count = counts_per_league.get(league_id, 0)
        expected_count = int(league["team_count"])

        if actual_count != expected_count:
            _add_error(
                result,
                "LEAGUE",
                league_id,
                "team_count",
                expected_count,
                actual_count,
                "O número de equipas da liga está incorreto.",
            )


def _validate_performance(
    performance: list[dict[str, Any]],
    teams: list[dict[str, Any]],
    result: ValidationResult,
) -> None:
    team_ids = {
        _as_text(team.get("team_id"))
        for team in teams
        if _as_text(team.get("team_id"))
    }

    seen_team_ids: set[str] = set()
    positions_per_source: set[tuple[str, int]] = set()

    for row_number, row in enumerate(performance, start=2):
        team_id = _as_text(row.get("team_id"))
        source_league = _as_text(row.get("source_league_id"))

        if not team_id:
            _add_error(
                result,
                "PERFORMANCE",
                f"ROW_{row_number}",
                "team_id",
                "valor preenchido",
                row.get("team_id"),
                "O desempenho não possui team_id.",
            )
            continue

        if team_id not in team_ids:
            _add_error(
                result,
                "PERFORMANCE",
                team_id,
                "team_id",
                "equipa existente",
                team_id,
                "O desempenho refere uma equipa inexistente.",
            )

        if team_id in seen_team_ids:
            _add_error(
                result,
                "PERFORMANCE",
                team_id,
                "team_id",
                "um registo por equipa",
                team_id,
                "A equipa possui mais de um desempenho.",
            )

        seen_team_ids.add(team_id)

        played = _as_int(row.get("played"))
        wins = _as_int(row.get("wins"))
        draws = _as_int(row.get("draws"))
        losses = _as_int(row.get("losses"))
        goals_for = _as_int(row.get("goals_for"))
        goals_against = _as_int(row.get("goals_against"))
        goal_difference = _as_int(row.get("goal_difference"))
        points = _as_int(row.get("points"))
        points_adjustment = _as_int(
            row.get("points_adjustment"),
            default=0,
        )
        position = _as_int(row.get("position"))

        numeric_values = {
            "played": played,
            "wins": wins,
            "draws": draws,
            "losses": losses,
            "goals_for": goals_for,
            "goals_against": goals_against,
            "position": position,
        }

        for field_name, value in numeric_values.items():
            if value is None or value < 0:
                _add_error(
                    result,
                    "PERFORMANCE",
                    team_id,
                    field_name,
                    "número inteiro não negativo",
                    row.get(field_name),
                    f"O campo {field_name} é inválido.",
                )

        if None not in {played, wins, draws, losses}:
            expected_played = wins + draws + losses

            if played != expected_played:
                _add_error(
                    result,
                    "PERFORMANCE",
                    team_id,
                    "played",
                    expected_played,
                    played,
                    "Jogos não equivalem a vitórias + empates + derrotas.",
                )

        if None not in {goals_for, goals_against, goal_difference}:
            expected_goal_difference = goals_for - goals_against

            if goal_difference != expected_goal_difference:
                _add_error(
                    result,
                    "PERFORMANCE",
                    team_id,
                    "goal_difference",
                    expected_goal_difference,
                    goal_difference,
                    "A diferença de golos está incorreta.",
                )

        if None not in {wins, draws, points, points_adjustment}:
            expected_points = (
                3 * wins
                + draws
                + points_adjustment
            )

            if points != expected_points:
                _add_error(
                    result,
                    "PERFORMANCE",
                    team_id,
                    "points",
                    expected_points,
                    points,
                    "A pontuação está incorreta.",
                )

        if source_league and position is not None:
            position_key = (source_league, position)

            if position_key in positions_per_source:
                _add_error(
                    result,
                    "PERFORMANCE",
                    team_id,
                    "position",
                    "posição única por liga de origem",
                    position,
                    "A posição está duplicada na liga de origem.",
                )

            positions_per_source.add(position_key)

        confidence = _as_float(row.get("data_confidence"))

        if confidence is None or not 0 <= confidence <= 1:
            _add_error(
                result,
                "PERFORMANCE",
                team_id,
                "data_confidence",
                "valor entre 0 e 1",
                row.get("data_confidence"),
                "A confiança dos dados é inválida.",
            )


def _validate_promoted(
    promoted_rows: list[dict[str, Any]],
    teams: list[dict[str, Any]],
    result: ValidationResult,
) -> None:
    promoted_team_ids = {
        _as_text(team.get("team_id"))
        for team in teams
        if _as_int(team.get("promoted")) == 1
    }

    listed_promoted_ids: set[str] = set()

    for row_number, row in enumerate(promoted_rows, start=2):
        team_id = _as_text(row.get("team_id"))

        if not team_id:
            _add_error(
                result,
                "PROMOTED",
                f"ROW_{row_number}",
                "team_id",
                "valor preenchido",
                row.get("team_id"),
                "A linha de promovida não possui team_id.",
            )
            continue

        if team_id in listed_promoted_ids:
            _add_error(
                result,
                "PROMOTED",
                team_id,
                "team_id",
                "valor único",
                team_id,
                "A equipa promovida está repetida.",
            )

        listed_promoted_ids.add(team_id)

        if team_id not in promoted_team_ids:
            _add_error(
                result,
                "PROMOTED",
                team_id,
                "team_id",
                "equipa marcada como promovida",
                team_id,
                "A equipa não está marcada como promovida.",
            )

        method = _as_text(row.get("promotion_method"))

        if method not in {"CHAMPION", "DIRECT", "PLAYOFF"}:
            _add_error(
                result,
                "PROMOTED",
                team_id,
                "promotion_method",
                "CHAMPION, DIRECT ou PLAYOFF",
                method,
                "Método de promoção inválido.",
            )

    missing_promoted = promoted_team_ids - listed_promoted_ids

    for team_id in sorted(missing_promoted):
        _add_error(
            result,
            "PROMOTED",
            team_id,
            "team_id",
            "equipa presente na folha Promovidas",
            None,
            "A equipa promovida não aparece na folha Promovidas.",
        )


def _validate_fixtures(
    fixtures: list[dict[str, Any]],
    teams: list[dict[str, Any]],
    result: ValidationResult,
) -> None:
    team_lookup = {
        _as_text(team.get("team_id")): team
        for team in teams
        if _as_text(team.get("team_id"))
    }

    match_ids: set[str] = set()
    pairings: set[tuple[str, str, str]] = set()
    counts_per_league: dict[str, int] = {}

    for row_number, match in enumerate(fixtures, start=2):
        match_id = _as_text(match.get("match_id"))
        league_id = _as_text(match.get("league_id"))
        home_id = _as_text(match.get("home_team_id"))
        away_id = _as_text(match.get("away_team_id"))

        if not match_id:
            _add_error(
                result,
                "MATCH",
                f"ROW_{row_number}",
                "match_id",
                "valor preenchido",
                match.get("match_id"),
                "O jogo não possui match_id.",
            )
            continue

        if match_id in match_ids:
            _add_error(
                result,
                "MATCH",
                match_id,
                "match_id",
                "valor único",
                match_id,
                "O match_id está duplicado.",
            )

        match_ids.add(match_id)

        if home_id == away_id and home_id:
            _add_error(
                result,
                "MATCH",
                match_id,
                "away_team_id",
                "equipa diferente da equipa da casa",
                away_id,
                "Uma equipa não pode jogar contra si própria.",
            )

        for field_name, team_id in (
            ("home_team_id", home_id),
            ("away_team_id", away_id),
        ):
            if team_id not in team_lookup:
                _add_error(
                    result,
                    "MATCH",
                    match_id,
                    field_name,
                    "equipa existente",
                    team_id,
                    "O jogo refere uma equipa inexistente.",
                )

        if home_id in team_lookup:
            home_league = _as_text(
                team_lookup[home_id].get("league_id")
            )

            if home_league != league_id:
                _add_error(
                    result,
                    "MATCH",
                    match_id,
                    "home_team_id",
                    f"equipa da liga {league_id}",
                    home_league,
                    "A equipa da casa pertence a outra liga.",
                )

        if away_id in team_lookup:
            away_league = _as_text(
                team_lookup[away_id].get("league_id")
            )

            if away_league != league_id:
                _add_error(
                    result,
                    "MATCH",
                    match_id,
                    "away_team_id",
                    f"equipa da liga {league_id}",
                    away_league,
                    "A equipa visitante pertence a outra liga.",
                )

        if league_id:
            counts_per_league[league_id] = (
                counts_per_league.get(league_id, 0) + 1
            )

        pairing = (league_id, home_id, away_id)

        if pairing in pairings:
            _add_error(
                result,
                "MATCH",
                match_id,
                "home_team_id/away_team_id",
                "emparelhamento único",
                f"{home_id} vs {away_id}",
                "O mesmo jogo casa/fora está duplicado.",
            )

        pairings.add(pairing)

    configured = get_active_leagues()

    for league_id, league in configured.items():
        actual_count = counts_per_league.get(league_id, 0)
        expected_count = int(league["total_matches"])

        if fixtures and actual_count != expected_count:
            _add_error(
                result,
                "LEAGUE",
                league_id,
                "fixture_count",
                expected_count,
                actual_count,
                "O número de jogos da liga está incorreto.",
            )


def _validate_expected_totals(
    result: ValidationResult,
) -> None:
    validation_config = load_model_weights()["validation"]

    expected = {
        "leagues": int(validation_config["expected_leagues"]),
        "teams": int(validation_config["expected_teams"]),
        "performance": int(validation_config["expected_teams"]),
        "fixtures": int(validation_config["expected_matches"]),
    }

    for field_name, expected_value in expected.items():
        actual_value = result.counts.get(field_name, 0)

        if actual_value != expected_value:
            _add_error(
                result,
                "DATASET",
                None,
                field_name,
                expected_value,
                actual_value,
                f"Quantidade total incorreta em {field_name}.",
            )


def _validate_cross_references(
    teams: list[dict[str, Any]],
    performance: list[dict[str, Any]],
    promoted: list[dict[str, Any]],
    fixtures: list[dict[str, Any]],
    result: ValidationResult,
) -> None:
    team_ids = {
        _as_text(team.get("team_id"))
        for team in teams
        if _as_text(team.get("team_id"))
    }

    performance_ids = {
        _as_text(row.get("team_id"))
        for row in performance
        if _as_text(row.get("team_id"))
    }

    for missing_team_id in sorted(team_ids - performance_ids):
        _add_error(
            result,
            "TEAM",
            missing_team_id,
            "performance",
            "um registo de desempenho",
            None,
            "A equipa não possui desempenho da época anterior.",
        )

    fixture_team_ids: set[str] = set()

    for match in fixtures:
        home_id = _as_text(match.get("home_team_id"))
        away_id = _as_text(match.get("away_team_id"))

        if home_id:
            fixture_team_ids.add(home_id)

        if away_id:
            fixture_team_ids.add(away_id)

    if fixtures:
        for missing_team_id in sorted(team_ids - fixture_team_ids):
            _add_error(
                result,
                "TEAM",
                missing_team_id,
                "fixtures",
                "equipa presente no calendário",
                None,
                "A equipa não aparece no calendário.",
            )


def _add_error(
    result: ValidationResult,
    entity_type: str,
    entity_id: str | None,
    field_name: str | None,
    expected_value: Any,
    actual_value: Any,
    message: str,
) -> None:
    result.issues.append(
        ValidationIssue(
            severity="ERROR",
            entity_type=entity_type,
            entity_id=entity_id,
            field_name=field_name,
            expected_value=expected_value,
            actual_value=actual_value,
            message=message,
        )
    )


def _add_warning(
    result: ValidationResult,
    entity_type: str,
    entity_id: str | None,
    field_name: str | None,
    expected_value: Any,
    actual_value: Any,
    message: str,
) -> None:
    result.issues.append(
        ValidationIssue(
            severity="WARNING",
            entity_type=entity_type,
            entity_id=entity_id,
            field_name=field_name,
            expected_value=expected_value,
            actual_value=actual_value,
            message=message,
        )
    )


def _as_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _as_int(
    value: Any,
    default: int | None = None,
) -> int | None:
    if value is None or value == "":
        return default

    try:
        number = float(value)

        if not math.isfinite(number):
            return default

        if not number.is_integer():
            return default

        return int(number)

    except (TypeError, ValueError):
        return default


def _as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None

    try:
        number = float(value)

        if not math.isfinite(number):
            return None

        return number

    except (TypeError, ValueError):
        return None
