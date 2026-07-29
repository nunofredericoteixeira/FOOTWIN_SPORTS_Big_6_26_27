# -*- coding: utf-8 -*-

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config.path_config import load_paths_config
from src.utils.logger import get_logger
from src.validation.dataset_validator import ValidationResult


logger = get_logger("validation.report")


HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

ERROR_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE",
)

WARNING_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFEB9C",
)

INFO_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

APPROVED_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE",
)

REJECTED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE",
)


def create_validation_report(
    result: ValidationResult,
    output_path: str | Path | None = None,
) -> Path:
    """
    Cria um relatório Excel com o resultado completo da validação.
    """

    paths = load_paths_config()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if output_path is None:
        output_file = (
            paths["reports"]["validation"]
            / f"Validation_Report_{timestamp}.xlsx"
        )
    else:
        output_file = Path(output_path).expanduser().resolve()

    output_file.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Resumo"

    issues_sheet = workbook.create_sheet("Erros_Avisos")
    counts_sheet = workbook.create_sheet("Contagens")

    _populate_summary_sheet(
        worksheet=summary_sheet,
        result=result,
    )

    _populate_issues_sheet(
        worksheet=issues_sheet,
        result=result,
    )

    _populate_counts_sheet(
        worksheet=counts_sheet,
        result=result,
    )

    workbook.save(output_file)
    workbook.close()

    logger.info(
        "Relatório de validação criado | ficheiro=%s",
        output_file,
    )

    return output_file


def _populate_summary_sheet(
    worksheet,
    result: ValidationResult,
) -> None:
    headers = [
        "Indicador",
        "Valor",
    ]

    _write_headers(
        worksheet=worksheet,
        headers=headers,
    )

    rows = [
        ("Dataset", str(result.dataset_path)),
        ("Estado", result.status),
        ("Aprovado", "SIM" if result.approved else "NÃO"),
        ("Erros", result.error_count),
        ("Avisos", result.warning_count),
        ("Informações", result.info_count),
        (
            "Data do relatório",
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ),
    ]

    for row_number, values in enumerate(
        rows,
        start=2,
    ):
        worksheet.cell(
            row=row_number,
            column=1,
            value=values[0],
        )

        value_cell = worksheet.cell(
            row=row_number,
            column=2,
            value=values[1],
        )

        if values[0] == "Estado":
            if result.approved:
                value_cell.fill = APPROVED_FILL
            else:
                value_cell.fill = REJECTED_FILL

            value_cell.font = Font(bold=True)

    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 90

    worksheet.freeze_panes = "A2"


def _populate_issues_sheet(
    worksheet,
    result: ValidationResult,
) -> None:
    headers = [
        "Número",
        "Severidade",
        "Tipo de entidade",
        "ID da entidade",
        "Campo",
        "Valor esperado",
        "Valor encontrado",
        "Mensagem",
    ]

    _write_headers(
        worksheet=worksheet,
        headers=headers,
    )

    for row_number, issue in enumerate(
        result.issues,
        start=2,
    ):
        values = [
            row_number - 1,
            issue.severity,
            issue.entity_type,
            issue.entity_id,
            issue.field_name,
            _convert_value(issue.expected_value),
            _convert_value(issue.actual_value),
            issue.message,
        ]

        for column_number, value in enumerate(
            values,
            start=1,
        ):
            cell = worksheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        severity_cell = worksheet.cell(
            row=row_number,
            column=2,
        )

        if issue.severity == "ERROR":
            severity_cell.fill = ERROR_FILL
        elif issue.severity == "WARNING":
            severity_cell.fill = WARNING_FILL
        else:
            severity_cell.fill = INFO_FILL

    widths = {
        "A": 10,
        "B": 14,
        "C": 20,
        "D": 24,
        "E": 24,
        "F": 35,
        "G": 35,
        "H": 60,
    }

    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        f"A1:H{max(2, len(result.issues) + 1)}"
    )


def _populate_counts_sheet(
    worksheet,
    result: ValidationResult,
) -> None:
    headers = [
        "Tipo de registo",
        "Quantidade encontrada",
    ]

    _write_headers(
        worksheet=worksheet,
        headers=headers,
    )

    for row_number, (
        record_type,
        count,
    ) in enumerate(
        sorted(result.counts.items()),
        start=2,
    ):
        worksheet.cell(
            row=row_number,
            column=1,
            value=record_type,
        )

        worksheet.cell(
            row=row_number,
            column=2,
            value=count,
        )

    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 24

    worksheet.freeze_panes = "A2"


def _write_headers(
    worksheet,
    headers: list[str],
) -> None:
    for column_number, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=1,
            column=column_number,
            value=header,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.row_dimensions[1].height = 24


def _convert_value(value) -> str | int | float | None:
    """
    Converte estruturas complexas para texto compatível com Excel.
    """

    if value is None:
        return None

    if isinstance(value, (str, int, float)):
        return value

    if isinstance(value, (list, tuple, set)):
        return ", ".join(
            str(item)
            for item in value
        )

    if isinstance(value, dict):
        return "; ".join(
            f"{key}={item}"
            for key, item in value.items()
        )

    return str(value)
