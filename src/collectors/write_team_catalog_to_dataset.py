# -*- coding: utf-8 -*-

from __future__ import annotations

import re
import shutil
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from src.collectors.team_catalog import (
    EXPECTED_COUNTS,
    TEAM_CATALOG,
    TeamCatalogEntry,
    validate_team_catalog,
)


DEFAULT_DATASET_PATH = Path(
    "data/input/FOOTWIN_Dataset_2026_27_V001.xlsx"
)

SHEET_NAME = "Equipas_2026_27"

EXPECTED_HEADERS = [
    "team_id",
    "team_name",
    "short_name",
    "normalized_name",
    "league_id",
    "country",
    "season_label",
    "promoted",
    "promotion_method",
    "previous_division",
    "active",
]


@dataclass(frozen=True)
class TeamCatalogWriteResult:
    dataset_path: Path
    backup_path: Path
    rows_deleted: int
    rows_written: int
    league_counts: dict[str, int]


class TeamCatalogWriteError(RuntimeError):
    """Erro ao gravar o catálogo de equipas no dataset."""


def normalize_team_name(
    value: str,
) -> str:
    """
    Cria uma versão normalizada do nome da equipa.
    """

    normalized = unicodedata.normalize(
        "NFKD",
        value,
    )

    ascii_text = "".join(
        character
        for character in normalized
        if not unicodedata.combining(
            character
        )
    )

    ascii_text = ascii_text.casefold()

    ascii_text = re.sub(
        r"[^a-z0-9]+",
        " ",
        ascii_text,
    )

    return re.sub(
        r"\s+",
        " ",
        ascii_text,
    ).strip()


def build_backup_path(
    dataset_path: Path,
) -> Path:
    """
    Cria um nome único para o backup anterior à gravação.
    """

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    return dataset_path.with_name(
        f"{dataset_path.stem}"
        f"_BEFORE_TEAMS_"
        f"{timestamp}"
        f"{dataset_path.suffix}"
    )


def create_backup(
    dataset_path: Path,
) -> Path:
    """
    Cria uma cópia integral do dataset.
    """

    backup_path = build_backup_path(
        dataset_path
    )

    shutil.copy2(
        dataset_path,
        backup_path,
    )

    if not backup_path.exists():
        raise TeamCatalogWriteError(
            "O backup do dataset não foi criado."
        )

    if backup_path.stat().st_size <= 0:
        raise TeamCatalogWriteError(
            "O backup criado está vazio."
        )

    return backup_path


def validate_headers(
    worksheet,
) -> None:
    """
    Confirma os cabeçalhos da folha de equipas.
    """

    headers = [
        worksheet.cell(
            row=1,
            column=column,
        ).value
        for column in range(
            1,
            len(EXPECTED_HEADERS) + 1,
        )
    ]

    if headers != EXPECTED_HEADERS:
        raise TeamCatalogWriteError(
            "Os cabeçalhos da folha "
            f"{SHEET_NAME} não correspondem "
            "ao formato esperado.\n"
            f"Esperado: {EXPECTED_HEADERS}\n"
            f"Encontrado: {headers}"
        )


def clear_existing_rows(
    worksheet,
) -> int:
    """
    Remove todos os registos existentes abaixo do cabeçalho.
    """

    if worksheet.max_row <= 1:
        return 0

    rows_to_delete = (
        worksheet.max_row - 1
    )

    worksheet.delete_rows(
        2,
        rows_to_delete,
    )

    return rows_to_delete


def build_excel_row(
    entry: TeamCatalogEntry,
) -> list:
    """
    Converte uma equipa no formato esperado pelo Excel.
    """

    return [
        entry.team_id,
        entry.team_name,
        entry.short_name,
        normalize_team_name(
            entry.team_name
        ),
        entry.league_id,
        entry.country,
        "2026/27",
        entry.promoted,
        entry.promotion_method,
        entry.previous_division,
        1,
    ]


def write_catalog_rows(
    worksheet,
) -> dict[str, int]:
    """
    Grava as 114 equipas no Excel.
    """

    current_row = 2
    league_counts: dict[str, int] = {}

    league_order = (
        "ENG1",
        "ESP1",
        "ITA1",
        "GER1",
        "FRA1",
        "POR1",
    )

    for league_id in league_order:
        entries = TEAM_CATALOG[
            league_id
        ]

        league_counts[
            league_id
        ] = len(entries)

        for entry in entries:
            values = build_excel_row(
                entry
            )

            for column, value in enumerate(
                values,
                start=1,
            ):
                worksheet.cell(
                    row=current_row,
                    column=column,
                    value=value,
                )

            current_row += 1

    return league_counts


def validate_written_rows(
    worksheet,
) -> None:
    """
    Valida os registos depois da gravação.
    """

    rows = list(
        worksheet.iter_rows(
            min_row=2,
            max_col=len(
                EXPECTED_HEADERS
            ),
            values_only=True,
        )
    )

    non_empty_rows = [
        row
        for row in rows
        if any(
            value is not None
            for value in row
        )
    ]

    if len(non_empty_rows) != 114:
        raise TeamCatalogWriteError(
            "Quantidade gravada incorreta: "
            f"{len(non_empty_rows)}. "
            "Esperadas: 114."
        )

    team_ids = [
        str(row[0]).strip()
        for row in non_empty_rows
    ]

    if len(team_ids) != len(
        set(team_ids)
    ):
        raise TeamCatalogWriteError(
            "Foram gravados team_id duplicados."
        )

    league_counts: dict[str, int] = {}

    for row in non_empty_rows:
        league_id = str(
            row[4]
        ).strip()

        league_counts[
            league_id
        ] = (
            league_counts.get(
                league_id,
                0,
            )
            + 1
        )

        if row[6] != "2026/27":
            raise TeamCatalogWriteError(
                f"Época incorreta em "
                f"{row[0]}: {row[6]}"
            )

        if row[10] != 1:
            raise TeamCatalogWriteError(
                f"Estado active incorreto em "
                f"{row[0]}: {row[10]}"
            )

    if league_counts != EXPECTED_COUNTS:
        raise TeamCatalogWriteError(
            "Contagens por liga incorretas.\n"
            f"Esperado: {EXPECTED_COUNTS}\n"
            f"Encontrado: {league_counts}"
        )


def write_team_catalog_to_dataset(
    dataset_path: str | Path = (
        DEFAULT_DATASET_PATH
    ),
) -> TeamCatalogWriteResult:
    """
    Grava o catálogo completo na folha Equipas_2026_27.
    """

    validate_team_catalog()

    final_dataset_path = Path(
        dataset_path
    ).expanduser().resolve()

    if not final_dataset_path.exists():
        raise TeamCatalogWriteError(
            "O dataset não existe: "
            f"{final_dataset_path}"
        )

    if not final_dataset_path.is_file():
        raise TeamCatalogWriteError(
            "O caminho não corresponde "
            "a um ficheiro: "
            f"{final_dataset_path}"
        )

    backup_path = create_backup(
        final_dataset_path
    )

    workbook = load_workbook(
        final_dataset_path
    )

    try:
        if SHEET_NAME not in (
            workbook.sheetnames
        ):
            raise TeamCatalogWriteError(
                "Falta a folha obrigatória: "
                f"{SHEET_NAME}"
            )

        worksheet = workbook[
            SHEET_NAME
        ]

        validate_headers(
            worksheet
        )

        rows_deleted = (
            clear_existing_rows(
                worksheet
            )
        )

        league_counts = (
            write_catalog_rows(
                worksheet
            )
        )

        validate_written_rows(
            worksheet
        )

        workbook.save(
            final_dataset_path
        )

    except Exception:
        workbook.close()

        shutil.copy2(
            backup_path,
            final_dataset_path,
        )

        raise

    finally:
        try:
            workbook.close()
        except Exception:
            pass

    if not final_dataset_path.exists():
        raise TeamCatalogWriteError(
            "O dataset desapareceu "
            "após a gravação."
        )

    return TeamCatalogWriteResult(
        dataset_path=(
            final_dataset_path
        ),
        backup_path=backup_path,
        rows_deleted=rows_deleted,
        rows_written=114,
        league_counts=(
            league_counts
        ),
    )


def print_result(
    result: TeamCatalogWriteResult,
) -> None:
    print()
    print("=" * 100)
    print(
        "✅ CATÁLOGO DE EQUIPAS "
        "GRAVADO NO DATASET"
    )
    print("=" * 100)

    print(
        f"Dataset:          "
        f"{result.dataset_path}"
    )

    print(
        f"Backup:           "
        f"{result.backup_path}"
    )

    print(
        f"Linhas removidas: "
        f"{result.rows_deleted}"
    )

    print(
        f"Linhas gravadas:  "
        f"{result.rows_written}"
    )

    print("-" * 100)

    for league_id, count in (
        result.league_counts.items()
    ):
        print(
            f"{league_id}: "
            f"{count} equipas"
        )

    print("=" * 100)


def main() -> int:
    try:
        result = (
            write_team_catalog_to_dataset()
        )

        print_result(
            result
        )

        return 0

    except Exception as exc:
        print()
        print("=" * 100)
        print(
            "❌ ERRO AO GRAVAR "
            "O CATÁLOGO DE EQUIPAS"
        )
        print("=" * 100)
        print(exc)
        print("=" * 100)

        return 1


if __name__ == "__main__":
    raise SystemExit(
        main()
    )
