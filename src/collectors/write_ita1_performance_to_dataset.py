# -*- coding: utf-8 -*-

from __future__ import annotations

import json
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook


DATASET_PATH = Path(
    "data/input/FOOTWIN_Dataset_2026_27_V001.xlsx"
)

PERFORMANCE_SHEET_NAME = "Desempenho_2025_26"
TEAMS_SHEET_NAME = "Equipas_2026_27"

SOURCE_URL = (
    "https://en.legaseriea.it/serie-a/standings"
)

SEASON_ID = (
    "serie-a::Football_Season::"
    "5f0e080fc3a44073984b75b3a8e06a8a"
)

ACTION_ID = (
    "7f17e9c8520173649139d9ccfbd370bf58e8779179"
)

SOURCE_LEAGUE_ID = "ITA1"
TARGET_LEAGUE_ID = "ITA1"
SEASON_LABEL = "2025/26"

EXCLUDED_RELEGATED_TEAMS = {
    "cremonese",
    "hellas verona",
    "pisa",
}

PROMOTED_TARGET_TEAMS = {
    "frosinone",
    "monza",
    "venezia",
}


class ItalianPerformanceCollectorError(
    RuntimeError
):
    """Erro na recolha ou escrita da performance ITA1."""


def normalize_name(
    value: Any,
) -> str:
    return (
        str(value or "")
        .strip()
        .casefold()
    )


def request_standings() -> dict[str, Any]:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 "
            "(Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 "
            "(KHTML, like Gecko) "
            "Chrome/150.0.0.0 "
            "Safari/537.36"
        ),
        "Accept": "text/x-component",
        "Content-Type": (
            "text/plain;charset=UTF-8"
        ),
        "Next-Action": ACTION_ID,
        "Origin": "https://en.legaseriea.it",
        "Referer": SOURCE_URL,
    }

    response = requests.post(
        SOURCE_URL,
        headers=headers,
        data=json.dumps(
            [SEASON_ID]
        ),
        timeout=60,
    )

    response.raise_for_status()

    for line in response.text.splitlines():
        if line.startswith("1:"):
            payload = json.loads(
                line[2:]
            )

            if not isinstance(
                payload,
                dict,
            ):
                raise ItalianPerformanceCollectorError(
                    "Payload da classificação inválido."
                )

            return payload

    raise ItalianPerformanceCollectorError(
        "Não foi encontrado o payload "
        "da classificação ITA1."
    )


def parse_standings(
    payload: dict[str, Any],
) -> list[dict[str, Any]]:
    teams = payload.get(
        "teams",
        [],
    )

    if len(teams) != 20:
        raise ItalianPerformanceCollectorError(
            "Número inesperado de equipas "
            f"na Serie A: {len(teams)}"
        )

    standings: list[
        dict[str, Any]
    ] = []

    for team in teams:
        stats = {
            str(item.get("statsId")): (
                item.get("statsValue")
            )
            for item in team.get(
                "stats",
                [],
            )
        }

        official_name = str(
            team.get("officialName")
            or team.get("shortName")
            or ""
        ).strip()

        record = {
            "source_team_name": official_name,
            "normalized_name": normalize_name(
                official_name
            ),
            "position": int(
                stats["rank"]
            ),
            "played": int(
                stats["matches-played"]
            ),
            "wins": int(
                stats["win"]
            ),
            "draws": int(
                stats["draw"]
            ),
            "losses": int(
                stats["lose"]
            ),
            "goals_for": int(
                stats["goals-for"]
            ),
            "goals_against": int(
                stats["goals-against"]
            ),
            "goal_difference": int(
                stats["goal-difference"]
            ),
            "points": int(
                stats["points"]
            ),
        }

        if (
            record["played"]
            != record["wins"]
            + record["draws"]
            + record["losses"]
        ):
            raise ItalianPerformanceCollectorError(
                "Totais de jogos incoerentes para "
                f"{official_name}."
            )

        if (
            record["goal_difference"]
            != record["goals_for"]
            - record["goals_against"]
        ):
            raise ItalianPerformanceCollectorError(
                "Diferença de golos incoerente para "
                f"{official_name}."
            )

        expected_points = (
            3 * record["wins"]
            + record["draws"]
        )

        if record["points"] != expected_points:
            raise ItalianPerformanceCollectorError(
                "Pontuação incoerente para "
                f"{official_name}."
            )

        standings.append(
            record
        )

    return standings


def load_target_teams(
    workbook,
) -> dict[str, dict[str, Any]]:
    worksheet = workbook[
        TEAMS_SHEET_NAME
    ]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    indexes = {
        str(header): index
        for index, header in enumerate(
            headers
        )
        if header is not None
    }

    required = {
        "team_id",
        "team_name",
        "short_name",
        "normalized_name",
        "league_id",
        "promoted",
        "promotion_method",
    }

    missing = (
        required
        - set(indexes)
    )

    if missing:
        raise ItalianPerformanceCollectorError(
            "Faltam colunas na folha de equipas: "
            f"{sorted(missing)}"
        )

    teams: dict[
        str,
        dict[str, Any]
    ] = {}

    for row in worksheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        league_id = str(
            row[indexes["league_id"]]
            or ""
        ).strip().upper()

        if league_id != TARGET_LEAGUE_ID:
            continue

        normalized_candidates = {
            normalize_name(
                row[indexes["team_name"]]
            ),
            normalize_name(
                row[indexes["short_name"]]
            ),
            normalize_name(
                row[indexes["normalized_name"]]
            ),
        }

        record = {
            "team_id": row[
                indexes["team_id"]
            ],
            "promoted": int(
                row[indexes["promoted"]]
                or 0
            ),
            "promotion_method": (
                str(
                    row[
                        indexes[
                            "promotion_method"
                        ]
                    ]
                ).strip().upper()
                if row[
                    indexes[
                        "promotion_method"
                    ]
                ]
                else None
            ),
        }

        for candidate in normalized_candidates:
            if candidate:
                teams[candidate] = record

    return teams


def remove_existing_ita1_rows(
    worksheet,
    headers: list[Any],
) -> int:
    indexes = {
        str(header): index
        for index, header in enumerate(
            headers
        )
        if header is not None
    }

    rows_to_delete: list[int] = []

    for row_number in range(
        worksheet.max_row,
        1,
        -1,
    ):
        target_value = worksheet.cell(
            row=row_number,
            column=(
                indexes[
                    "target_league_id"
                ]
                + 1
            ),
        ).value

        if (
            str(target_value or "")
            .strip()
            .upper()
            == TARGET_LEAGUE_ID
        ):
            rows_to_delete.append(
                row_number
            )

    for row_number in rows_to_delete:
        worksheet.delete_rows(
            row_number,
            1,
        )

    return len(rows_to_delete)


def write_performance(
    standings: list[dict[str, Any]],
) -> None:
    if not DATASET_PATH.exists():
        raise ItalianPerformanceCollectorError(
            f"Dataset inexistente: {DATASET_PATH}"
        )

    backup_path = DATASET_PATH.with_name(
        f"{DATASET_PATH.stem}_BACKUP_BEFORE_ITA1"
        f"{DATASET_PATH.suffix}"
    )

    shutil.copy2(
        DATASET_PATH,
        backup_path,
    )

    workbook = load_workbook(
        DATASET_PATH
    )

    try:
        if PERFORMANCE_SHEET_NAME not in (
            workbook.sheetnames
        ):
            raise ItalianPerformanceCollectorError(
                "Falta a folha "
                f"{PERFORMANCE_SHEET_NAME}."
            )

        target_teams = load_target_teams(
            workbook
        )

        worksheet = workbook[
            PERFORMANCE_SHEET_NAME
        ]

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        header_indexes = {
            str(header): index
            for index, header in enumerate(
                headers
            )
            if header is not None
        }

        required_headers = {
            "team_id",
            "source_league_id",
            "target_league_id",
            "season_label",
            "position",
            "played",
            "wins",
            "draws",
            "losses",
            "goals_for",
            "goals_against",
            "goal_difference",
            "points",
            "points_adjustment",
            "promoted",
            "promotion_method",
            "source_status",
            "data_confidence",
            "source_url",
            "accessed_at",
        }

        missing_headers = (
            required_headers
            - set(header_indexes)
        )

        if missing_headers:
            raise ItalianPerformanceCollectorError(
                "Faltam colunas na folha de "
                "performance: "
                f"{sorted(missing_headers)}"
            )

        deleted = remove_existing_ita1_rows(
            worksheet,
            headers,
        )

        accessed_at = datetime.now(
            timezone.utc
        ).replace(
            microsecond=0
        ).isoformat()

        written = 0
        skipped_relegated = 0
        missing_target: list[str] = []

        for performance in standings:
            normalized_name = performance[
                "normalized_name"
            ]

            if normalized_name in (
                EXCLUDED_RELEGATED_TEAMS
            ):
                skipped_relegated += 1
                continue

            target_team = target_teams.get(
                normalized_name
            )

            if target_team is None:
                missing_target.append(
                    performance[
                        "source_team_name"
                    ]
                )
                continue

            if target_team["promoted"] != 0:
                raise ItalianPerformanceCollectorError(
                    "Uma equipa recolhida da ITA1 "
                    "está marcada como promovida: "
                    f"{performance['source_team_name']}"
                )

            row_data = {
                "team_id": (
                    target_team["team_id"]
                ),
                "source_league_id": (
                    SOURCE_LEAGUE_ID
                ),
                "target_league_id": (
                    TARGET_LEAGUE_ID
                ),
                "season_label": (
                    SEASON_LABEL
                ),
                "position": (
                    performance["position"]
                ),
                "played": (
                    performance["played"]
                ),
                "wins": (
                    performance["wins"]
                ),
                "draws": (
                    performance["draws"]
                ),
                "losses": (
                    performance["losses"]
                ),
                "goals_for": (
                    performance["goals_for"]
                ),
                "goals_against": (
                    performance["goals_against"]
                ),
                "goal_difference": (
                    performance[
                        "goal_difference"
                    ]
                ),
                "points": (
                    performance["points"]
                ),
                "points_adjustment": 0,
                "promoted": 0,
                "promotion_method": None,
                "source_status": "CONFIRMED",
                "data_confidence": 1.0,
                "source_url": SOURCE_URL,
                "accessed_at": accessed_at,
            }

            worksheet.append(
                [
                    row_data.get(
                        str(header)
                    )
                    for header in headers
                ]
            )

            written += 1

        if missing_target:
            raise ItalianPerformanceCollectorError(
                "Não foi possível mapear estas "
                "equipas da ITA1: "
                + ", ".join(
                    missing_target
                )
            )

        if written != 17:
            raise ItalianPerformanceCollectorError(
                "Esperavam-se 17 equipas ITA1 "
                f"gravadas, mas foram gravadas {written}."
            )

        promoted_names_found = {
            name
            for name in PROMOTED_TARGET_TEAMS
            if name in target_teams
        }

        if promoted_names_found != (
            PROMOTED_TARGET_TEAMS
        ):
            raise ItalianPerformanceCollectorError(
                "Não foram encontradas todas as "
                "promovidas ITA1 no dataset."
            )

        workbook.save(
            DATASET_PATH
        )

        print()
        print("=" * 100)
        print(
            "ITA1 — PERFORMANCE GRAVADA NO DATASET"
        )
        print("=" * 100)
        print(
            f"Dataset: {DATASET_PATH.resolve()}"
        )
        print(
            f"Backup:  {backup_path.resolve()}"
        )
        print(
            f"Linhas ITA1 removidas: {deleted}"
        )
        print(
            f"Equipas ITA1 gravadas: {written}"
        )
        print(
            "Relegadas ignoradas: "
            f"{skipped_relegated}"
        )
        print(
            "Promovidas pendentes da ITA2: "
            "Frosinone, Monza, Venezia"
        )
        print("=" * 100)

    except Exception:
        workbook.close()

        shutil.copy2(
            backup_path,
            DATASET_PATH,
        )

        raise

    else:
        workbook.close()


def main() -> int:
    print()
    print("=" * 100)
    print(
        "FOOTWIN SPORTS — RECOLHA ITA1 2025/26"
    )
    print("=" * 100)

    payload = request_standings()

    competition = payload.get(
        "competition",
        {},
    )

    print(
        "Competição: "
        f"{competition.get('name')}"
    )
    print(
        "Época: "
        f"{competition.get('seasonName')}"
    )

    standings = parse_standings(
        payload
    )

    print(
        f"Equipas recolhidas: {len(standings)}"
    )

    write_performance(
        standings
    )

    print()
    print(
        "✅ ITA1 concluída com sucesso."
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(
        main()
    )
