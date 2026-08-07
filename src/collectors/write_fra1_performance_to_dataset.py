# -*- coding: utf-8 -*-

from __future__ import annotations

import json
import re
import shutil
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DATASET_PATH = Path(
    "data/input/FOOTWIN_Dataset_2026_27_V001.xlsx"
)

FRA1_JSON_PATH = Path(
    "data/raw/performance_pages/fra_api_standings/"
    "02_FRA1_championship-standings_1_general__season-2025.txt"
)

FRA2_JSON_PATH = Path(
    "data/raw/performance_pages/fra_api_standings/"
    "08_FRA2_championship-standings_4_general__season-2025.txt"
)

TEAMS_SHEET_NAME = "Equipas_2026_27"
PROMOTED_SHEET_NAME = "Promovidas"
PERFORMANCE_SHEET_NAME = "Desempenho_2025_26"

TARGET_LEAGUE_ID = "FRA1"
SEASON_LABEL = "2025/26"

FRA1_SOURCE_URL = (
    "https://ma-api.ligue1.fr/"
    "championship-standings/1/general?season=2025"
)

FRA2_SOURCE_URL = (
    "https://ma-api.ligue1.fr/"
    "championship-standings/4/general?season=2025"
)

RELEGATED_FRA1_TEAMS = {
    "fc nantes",
    "fc metz",
}

PROMOTED_TEAMS = {
    "estac troyes": {
        "team_id": "FRA1_TROYES",
        "promotion_method": "CHAMPION",
    },
    "le mans fc": {
        "team_id": "FRA1_LE_MANS",
        "promotion_method": "DIRECT",
    },
}


class FrenchPerformanceCollectorError(
    RuntimeError
):
    """Erro na recolha ou escrita da performance FRA1."""


def normalize_text(
    value: Any,
) -> str:
    text = str(value or "").strip().casefold()

    text = unicodedata.normalize(
        "NFKD",
        text,
    )

    text = "".join(
        character
        for character in text
        if not unicodedata.combining(character)
    )

    text = re.sub(
        r"[^a-z0-9]+",
        " ",
        text,
    )

    return re.sub(
        r"\s+",
        " ",
        text,
    ).strip()


def load_json(
    path: Path,
) -> dict[str, Any]:
    if not path.exists():
        raise FrenchPerformanceCollectorError(
            f"JSON inexistente: {path}"
        )

    try:
        data = json.loads(
            path.read_text(
                encoding="utf-8"
            )
        )
    except json.JSONDecodeError as exc:
        raise FrenchPerformanceCollectorError(
            f"JSON inválido: {path}"
        ) from exc

    if not isinstance(
        data,
        dict,
    ):
        raise FrenchPerformanceCollectorError(
            f"Conteúdo JSON inválido: {path}"
        )

    return data


def parse_standings(
    path: Path,
    expected_teams: int,
) -> list[dict[str, Any]]:
    data = load_json(
        path
    )

    season = data.get(
        "season"
    )

    if season != 2025:
        raise FrenchPerformanceCollectorError(
            f"Época inesperada em {path.name}: {season}"
        )

    raw_standings = data.get(
        "standings",
        {}
    )

    if not isinstance(
        raw_standings,
        dict,
    ):
        raise FrenchPerformanceCollectorError(
            f"Standings inválidos em {path}."
        )

    if len(raw_standings) != expected_teams:
        raise FrenchPerformanceCollectorError(
            f"Esperavam-se {expected_teams} equipas em "
            f"{path.name}, mas foram encontradas "
            f"{len(raw_standings)}."
        )

    standings: list[
        dict[str, Any]
    ] = []

    for key, row in raw_standings.items():
        identity = row.get(
            "clubIdentity",
            {}
        )

        raw_team_name = str(
            identity.get("name")
            or identity.get("officialName")
            or identity.get("shortName")
            or ""
        ).strip()

        if not raw_team_name:
            raise FrenchPerformanceCollectorError(
                f"Foi encontrada uma equipa sem nome em "
                f"{path.name}."
            )

        try:
            position = int(
                row.get(
                    "rank",
                    key,
                )
            )
            played = int(
                row["played"]
            )
            wins = int(
                row["wins"]
            )
            draws = int(
                row["draws"]
            )
            losses = int(
                row["losses"]
            )
            goals_for = int(
                row["forGoals"]
            )
            goals_against = int(
                row["againstGoals"]
            )
            goal_difference = int(
                row["goalsDifference"]
            )
            points = int(
                row["points"]
            )
        except (
            KeyError,
            TypeError,
            ValueError,
        ) as exc:
            raise FrenchPerformanceCollectorError(
                "Dados inválidos para "
                f"{raw_team_name}."
            ) from exc

        if (
            played
            != wins
            + draws
            + losses
        ):
            raise FrenchPerformanceCollectorError(
                "Totais de jogos incoerentes para "
                f"{raw_team_name}."
            )

        if (
            goal_difference
            != goals_for
            - goals_against
        ):
            raise FrenchPerformanceCollectorError(
                "Diferença de golos incoerente para "
                f"{raw_team_name}."
            )

        if (
            points
            != 3 * wins
            + draws
        ):
            raise FrenchPerformanceCollectorError(
                "Pontuação incoerente para "
                f"{raw_team_name}."
            )

        names = {
            raw_team_name,
            identity.get("officialName"),
            identity.get("shortName"),
            identity.get("displayName"),
            identity.get("businessName"),
        }

        lookup_names = {
            normalize_text(name)
            for name in names
            if name
        }

        standings.append(
            {
                "source_team_name": raw_team_name,
                "lookup_names": lookup_names,
                "position": position,
                "played": played,
                "wins": wins,
                "draws": draws,
                "losses": losses,
                "goals_for": goals_for,
                "goals_against": goals_against,
                "goal_difference": goal_difference,
                "points": points,
            }
        )

    expected_positions = set(
        range(
            1,
            expected_teams + 1,
        )
    )

    actual_positions = {
        record["position"]
        for record in standings
    }

    if actual_positions != expected_positions:
        raise FrenchPerformanceCollectorError(
            f"As posições de {path.name} não são completas."
        )

    return sorted(
        standings,
        key=lambda record: record["position"],
    )


def load_target_teams(
    workbook,
) -> dict[str, dict[str, Any]]:
    worksheet = workbook[
        TEAMS_SHEET_NAME
    ]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    indexes = {
        str(header): index
        for index, header in enumerate(
            headers
        )
        if header is not None
    }

    required = {
        "team_id",
        "team_name",
        "short_name",
        "normalized_name",
        "league_id",
        "promoted",
        "promotion_method",
    }

    missing = (
        required
        - set(indexes)
    )

    if missing:
        raise FrenchPerformanceCollectorError(
            "Faltam colunas na folha de equipas: "
            f"{sorted(missing)}"
        )

    teams: dict[
        str,
        dict[str, Any]
    ] = {}

    for row in worksheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        league_id = str(
            row[indexes["league_id"]]
            or ""
        ).strip().upper()

        if league_id != TARGET_LEAGUE_ID:
            continue

        keys = {
            normalize_text(
                row[indexes["team_name"]]
            ),
            normalize_text(
                row[indexes["short_name"]]
            ),
            normalize_text(
                row[indexes["normalized_name"]]
            ),
        }

        record = {
            "team_id": str(
                row[indexes["team_id"]]
                or ""
            ).strip(),
            "team_name": str(
                row[indexes["team_name"]]
                or ""
            ).strip(),
            "promoted": int(
                row[indexes["promoted"]]
                or 0
            ),
            "promotion_method": (
                str(
                    row[
                        indexes[
                            "promotion_method"
                        ]
                    ]
                    or ""
                )
                .strip()
                .upper()
                or None
            ),
        }

        for key in keys:
            if key:
                teams[key] = record

    aliases = {
        "havre athletic club": "le havre ac",
        "havre ac": "le havre ac",
        "le havre": "le havre ac",
        "estac troyes": "estac troyes",
        "troyes": "estac troyes",
        "le mans fc": "le mans fc",
    }

    for alias, target_key in aliases.items():
        if target_key in teams:
            teams[alias] = teams[
                target_key
            ]

    return teams


def find_target_team(
    performance: dict[str, Any],
    target_teams: dict[str, dict[str, Any]],
) -> dict[str, Any] | None:
    for lookup_name in performance[
        "lookup_names"
    ]:
        target_team = target_teams.get(
            lookup_name
        )

        if target_team is not None:
            return target_team

    return None


def remove_existing_rows(
    worksheet,
    headers: list[Any],
) -> int:
    indexes = {
        str(header): index
        for index, header in enumerate(
            headers
        )
        if header is not None
    }

    rows_to_delete: list[int] = []

    for row_number in range(
        worksheet.max_row,
        1,
        -1,
    ):
        target_league_id = worksheet.cell(
            row=row_number,
            column=(
                indexes[
                    "target_league_id"
                ]
                + 1
            ),
        ).value

        if (
            str(target_league_id or "")
            .strip()
            .upper()
            == TARGET_LEAGUE_ID
        ):
            rows_to_delete.append(
                row_number
            )

    for row_number in rows_to_delete:
        worksheet.delete_rows(
            row_number,
            1,
        )

    return len(rows_to_delete)


def update_promoted_sheet(
    workbook,
    promoted_records: list[dict[str, Any]],
) -> None:
    worksheet = workbook[
        PROMOTED_SHEET_NAME
    ]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    indexes = {
        str(header): index
        for index, header in enumerate(
            headers
        )
        if header is not None
    }

    by_team_id = {
        record["team_id"]: record
        for record in promoted_records
    }

    updated = 0

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        team_id = str(
            worksheet.cell(
                row=row_number,
                column=(
                    indexes["team_id"]
                    + 1
                ),
            ).value
            or ""
        ).strip()

        target_league_id = str(
            worksheet.cell(
                row=row_number,
                column=(
                    indexes[
                        "target_league_id"
                    ]
                    + 1
                ),
            ).value
            or ""
        ).strip().upper()

        if target_league_id != TARGET_LEAGUE_ID:
            continue

        record = by_team_id.get(
            team_id
        )

        if record is None:
            continue

        values = {
            "source_league_id": "FRA2",
            "source_position": record[
                "position"
            ],
            "promotion_method": record[
                "promotion_method"
            ],
            "played": record["played"],
            "points": record["points"],
            "goals_for": record[
                "goals_for"
            ],
            "goals_against": record[
                "goals_against"
            ],
            "goal_difference": record[
                "goal_difference"
            ],
            "source_status": "CONFIRMED",
            "data_confidence": 1.0,
            "source_url": FRA2_SOURCE_URL,
        }

        for field, value in values.items():
            if field not in indexes:
                continue

            worksheet.cell(
                row=row_number,
                column=indexes[field] + 1,
                value=value,
            )

        updated += 1

    if updated != 2:
        raise FrenchPerformanceCollectorError(
            "Esperavam-se 2 promovidas atualizadas, "
            f"mas foram atualizadas {updated}."
        )


def write_performance(
    fra1_standings: list[dict[str, Any]],
    fra2_standings: list[dict[str, Any]],
) -> None:
    if not DATASET_PATH.exists():
        raise FrenchPerformanceCollectorError(
            f"Dataset inexistente: {DATASET_PATH}"
        )

    backup_path = DATASET_PATH.with_name(
        f"{DATASET_PATH.stem}"
        "_BACKUP_BEFORE_FRA1"
        f"{DATASET_PATH.suffix}"
    )

    shutil.copy2(
        DATASET_PATH,
        backup_path,
    )

    workbook = load_workbook(
        DATASET_PATH
    )

    try:
        for required_sheet in (
            TEAMS_SHEET_NAME,
            PROMOTED_SHEET_NAME,
            PERFORMANCE_SHEET_NAME,
        ):
            if required_sheet not in workbook.sheetnames:
                raise FrenchPerformanceCollectorError(
                    f"Falta a folha {required_sheet}."
                )

        target_teams = load_target_teams(
            workbook
        )

        worksheet = workbook[
            PERFORMANCE_SHEET_NAME
        ]

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        required_headers = {
            "team_id",
            "source_league_id",
            "target_league_id",
            "season_label",
            "position",
            "played",
            "wins",
            "draws",
            "losses",
            "goals_for",
            "goals_against",
            "goal_difference",
            "points",
            "points_adjustment",
            "promoted",
            "promotion_method",
            "source_status",
            "data_confidence",
            "source_url",
            "accessed_at",
        }

        existing_headers = {
            str(header)
            for header in headers
            if header is not None
        }

        missing_headers = (
            required_headers
            - existing_headers
        )

        if missing_headers:
            raise FrenchPerformanceCollectorError(
                "Faltam colunas na folha de performance: "
                f"{sorted(missing_headers)}"
            )

        removed = remove_existing_rows(
            worksheet,
            headers,
        )

        accessed_at = datetime.now(
            timezone.utc
        ).replace(
            microsecond=0
        ).isoformat()

        permanent_written = 0
        relegated_skipped = 0
        promoted_written = 0
        missing_target: list[str] = []

        for performance in fra1_standings:
            if (
                performance["lookup_names"]
                & RELEGATED_FRA1_TEAMS
            ):
                relegated_skipped += 1
                continue

            target_team = find_target_team(
                performance,
                target_teams,
            )

            if target_team is None:
                missing_target.append(
                    performance[
                        "source_team_name"
                    ]
                )
                continue

            if target_team["promoted"] != 0:
                raise FrenchPerformanceCollectorError(
                    "Uma equipa da FRA1 anterior está "
                    "marcada como promovida: "
                    f"{performance['source_team_name']}"
                )

            row_data = {
                "team_id": target_team[
                    "team_id"
                ],
                "source_league_id": "FRA1",
                "target_league_id": "FRA1",
                "season_label": SEASON_LABEL,
                "position": performance[
                    "position"
                ],
                "played": performance[
                    "played"
                ],
                "wins": performance["wins"],
                "draws": performance[
                    "draws"
                ],
                "losses": performance[
                    "losses"
                ],
                "goals_for": performance[
                    "goals_for"
                ],
                "goals_against": performance[
                    "goals_against"
                ],
                "goal_difference": performance[
                    "goal_difference"
                ],
                "points": performance[
                    "points"
                ],
                "points_adjustment": 0,
                "promoted": 0,
                "promotion_method": None,
                "source_status": "CONFIRMED",
                "data_confidence": 1.0,
                "source_url": FRA1_SOURCE_URL,
                "accessed_at": accessed_at,
            }

            worksheet.append(
                [
                    row_data.get(
                        str(header)
                    )
                    for header in headers
                ]
            )

            permanent_written += 1

        promoted_records: list[
            dict[str, Any]
        ] = []

        for performance in fra2_standings:
            promoted_config = None

            for lookup_name in performance[
                "lookup_names"
            ]:
                promoted_config = (
                    PROMOTED_TEAMS.get(
                        lookup_name
                    )
                )

                if promoted_config is not None:
                    break

            if promoted_config is None:
                continue

            target_team = find_target_team(
                performance,
                target_teams,
            )

            if target_team is None:
                raise FrenchPerformanceCollectorError(
                    "Não foi possível mapear a promovida "
                    f"{performance['source_team_name']}."
                )

            if (
                target_team["team_id"]
                != promoted_config["team_id"]
            ):
                raise FrenchPerformanceCollectorError(
                    "team_id inesperado para "
                    f"{performance['source_team_name']}: "
                    f"{target_team['team_id']}."
                )

            if target_team["promoted"] != 1:
                raise FrenchPerformanceCollectorError(
                    "A equipa promovida não está marcada "
                    f"como promovida: "
                    f"{performance['source_team_name']}."
                )

            expected_method = promoted_config[
                "promotion_method"
            ]

            if (
                target_team["promotion_method"]
                != expected_method
            ):
                raise FrenchPerformanceCollectorError(
                    "Método de promoção inesperado para "
                    f"{performance['source_team_name']}. "
                    f"Dataset={target_team['promotion_method']} | "
                    f"Esperado={expected_method}"
                )

            row_data = {
                "team_id": target_team[
                    "team_id"
                ],
                "source_league_id": "FRA2",
                "target_league_id": "FRA1",
                "season_label": SEASON_LABEL,
                "position": performance[
                    "position"
                ],
                "played": performance[
                    "played"
                ],
                "wins": performance["wins"],
                "draws": performance[
                    "draws"
                ],
                "losses": performance[
                    "losses"
                ],
                "goals_for": performance[
                    "goals_for"
                ],
                "goals_against": performance[
                    "goals_against"
                ],
                "goal_difference": performance[
                    "goal_difference"
                ],
                "points": performance[
                    "points"
                ],
                "points_adjustment": 0,
                "promoted": 1,
                "promotion_method": expected_method,
                "source_status": "CONFIRMED",
                "data_confidence": 1.0,
                "source_url": FRA2_SOURCE_URL,
                "accessed_at": accessed_at,
            }

            worksheet.append(
                [
                    row_data.get(
                        str(header)
                    )
                    for header in headers
                ]
            )

            promoted_records.append(
                {
                    **performance,
                    "team_id": target_team[
                        "team_id"
                    ],
                    "promotion_method": (
                        expected_method
                    ),
                }
            )

            promoted_written += 1

        if missing_target:
            raise FrenchPerformanceCollectorError(
                "Não foi possível mapear estas equipas: "
                + ", ".join(
                    missing_target
                )
            )

        if permanent_written != 16:
            raise FrenchPerformanceCollectorError(
                "Esperavam-se 16 equipas permanentes "
                f"gravadas, mas foram gravadas "
                f"{permanent_written}."
            )

        if relegated_skipped != 2:
            raise FrenchPerformanceCollectorError(
                "Esperavam-se 2 despromovidas ignoradas, "
                f"mas foram ignoradas "
                f"{relegated_skipped}."
            )

        if promoted_written != 2:
            raise FrenchPerformanceCollectorError(
                "Esperavam-se 2 promovidas gravadas, "
                f"mas foram gravadas "
                f"{promoted_written}."
            )

        update_promoted_sheet(
            workbook,
            promoted_records,
        )

        workbook.save(
            DATASET_PATH
        )

        print()
        print("=" * 100)
        print(
            "FRA1 — PERFORMANCE GRAVADA NO DATASET"
        )
        print("=" * 100)
        print(
            f"Dataset: {DATASET_PATH.resolve()}"
        )
        print(
            f"Backup:  {backup_path.resolve()}"
        )
        print(
            f"Linhas FRA1 removidas: {removed}"
        )
        print(
            "Equipas permanentes gravadas: "
            f"{permanent_written}"
        )
        print(
            "Despromovidas ignoradas: "
            f"{relegated_skipped}"
        )
        print(
            "Promovidas FRA2 gravadas: "
            f"{promoted_written}"
        )
        print(
            "Total FRA1 2026/27 gravado: "
            f"{permanent_written + promoted_written}"
        )
        print("=" * 100)

    except Exception:
        workbook.close()

        shutil.copy2(
            backup_path,
            DATASET_PATH,
        )

        raise

    else:
        workbook.close()


def main() -> int:
    print()
    print("=" * 100)
    print(
        "FOOTWIN SPORTS — RECOLHA FRA1 2025/26"
    )
    print("=" * 100)

    fra1_standings = parse_standings(
        FRA1_JSON_PATH,
        expected_teams=18,
    )

    fra2_standings = parse_standings(
        FRA2_JSON_PATH,
        expected_teams=18,
    )

    print(
        "Equipas FRA1 recolhidas: "
        f"{len(fra1_standings)}"
    )

    print(
        "Equipas FRA2 recolhidas: "
        f"{len(fra2_standings)}"
    )

    write_performance(
        fra1_standings,
        fra2_standings,
    )

    print()
    print(
        "✅ FRA1 concluída com sucesso."
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(
        main()
    )
