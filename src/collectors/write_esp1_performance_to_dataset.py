# -*- coding: utf-8 -*-

from __future__ import annotations

import json
import re
import shutil
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DATASET_PATH = Path(
    "data/input/FOOTWIN_Dataset_2026_27_V001.xlsx"
)

ESP1_JSON_PATH = Path(
    "data/raw/performance_pages/"
    "ESP1_2025_26_laliga-easports-2025.json"
)

ESP2_JSON_PATH = Path(
    "data/raw/performance_pages/"
    "ESP2_2025_26_laliga-hypermotion-2025.json"
)

TEAMS_SHEET_NAME = "Equipas_2026_27"
PROMOTED_SHEET_NAME = "Promovidas"
PERFORMANCE_SHEET_NAME = "Desempenho_2025_26"

TARGET_LEAGUE_ID = "ESP1"
SEASON_LABEL = "2025/26"

ESP1_SOURCE_URL = (
    "https://apim.laliga.com/webview/"
    "api/web/subscriptions/"
    "laliga-easports-2025/standing"
)

ESP2_SOURCE_URL = (
    "https://apim.laliga.com/webview/"
    "api/web/subscriptions/"
    "laliga-hypermotion-2025/standing"
)

RELEGATED_ESP1_TEAMS = {
    "real oviedo",
    "girona fc",
    "rcd mallorca",
}

PROMOTED_TEAMS = {
    "r racing club": {
        "team_id": "ESP1_RACING",
        "promotion_method": "CHAMPION",
    },
    "rc deportivo": {
        "team_id": "ESP1_DEPORTIVO",
        "promotion_method": "DIRECT",
    },
    "malaga cf": {
        "team_id": "ESP1_MALAGA",
        "promotion_method": "PLAYOFF",
    },
}


class SpanishPerformanceCollectorError(
    RuntimeError
):
    """Erro na recolha ou escrita da performance ESP1."""


def normalize_text(
    value: Any,
) -> str:
    text = str(value or "").strip().casefold()

    text = unicodedata.normalize(
        "NFKD",
        text,
    )

    text = "".join(
        character
        for character in text
        if not unicodedata.combining(character)
    )

    text = re.sub(
        r"[^a-z0-9]+",
        " ",
        text,
    )

    return re.sub(
        r"\s+",
        " ",
        text,
    ).strip()


def load_json(
    path: Path,
) -> dict[str, Any]:
    if not path.exists():
        raise SpanishPerformanceCollectorError(
            f"JSON inexistente: {path}"
        )

    try:
        data = json.loads(
            path.read_text(
                encoding="utf-8"
            )
        )
    except json.JSONDecodeError as exc:
        raise SpanishPerformanceCollectorError(
            f"JSON inválido: {path}"
        ) from exc

    if not isinstance(
        data,
        dict,
    ):
        raise SpanishPerformanceCollectorError(
            f"Conteúdo JSON inválido: {path}"
        )

    return data


def parse_standings(
    path: Path,
    expected_teams: int,
) -> list[dict[str, Any]]:
    data = load_json(
        path
    )

    raw_standings = data.get(
        "standings",
        []
    )

    if not isinstance(
        raw_standings,
        list,
    ):
        raise SpanishPerformanceCollectorError(
            f"Standings inválidos em {path}."
        )

    if len(raw_standings) != expected_teams:
        raise SpanishPerformanceCollectorError(
            f"Esperavam-se {expected_teams} equipas em "
            f"{path.name}, mas foram encontradas "
            f"{len(raw_standings)}."
        )

    standings: list[
        dict[str, Any]
    ] = []

    for row in raw_standings:
        team = row.get(
            "team",
            {}
        )

        raw_team_name = str(
            team.get("nickname")
            or team.get("name")
            or ""
        ).strip()

        if not raw_team_name:
            raise SpanishPerformanceCollectorError(
                f"Foi encontrada uma equipa sem nome em "
                f"{path.name}."
            )

        try:
            position = int(
                row["position"]
            )
            played = int(
                row["played"]
            )
            wins = int(
                row["won"]
            )
            draws = int(
                row["drawn"]
            )
            losses = int(
                row["lost"]
            )
            goals_for = int(
                row["goals_for"]
            )
            goals_against = int(
                row["goals_against"]
            )
            points = int(
                row["points"]
            )
        except (
            KeyError,
            TypeError,
            ValueError,
        ) as exc:
            raise SpanishPerformanceCollectorError(
                "Dados inválidos para "
                f"{raw_team_name}."
            ) from exc

        goal_difference = (
            goals_for
            - goals_against
        )

        if (
            played
            != wins
            + draws
            + losses
        ):
            raise SpanishPerformanceCollectorError(
                "Totais de jogos incoerentes para "
                f"{raw_team_name}."
            )

        if (
            points
            != 3 * wins
            + draws
        ):
            raise SpanishPerformanceCollectorError(
                "Pontuação incoerente para "
                f"{raw_team_name}."
            )

        standings.append(
            {
                "source_team_name": raw_team_name,
                "lookup_name": normalize_text(
                    raw_team_name
                ),
                "position": position,
                "played": played,
                "wins": wins,
                "draws": draws,
                "losses": losses,
                "goals_for": goals_for,
                "goals_against": goals_against,
                "goal_difference": goal_difference,
                "points": points,
            }
        )

    expected_positions = set(
        range(
            1,
            expected_teams + 1,
        )
    )

    actual_positions = {
        record["position"]
        for record in standings
    }

    if actual_positions != expected_positions:
        raise SpanishPerformanceCollectorError(
            f"As posições de {path.name} não são completas."
        )

    return standings


def load_target_teams(
    workbook,
) -> dict[str, dict[str, Any]]:
    worksheet = workbook[
        TEAMS_SHEET_NAME
    ]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    indexes = {
        str(header): index
        for index, header in enumerate(
            headers
        )
        if header is not None
    }

    required = {
        "team_id",
        "team_name",
        "short_name",
        "normalized_name",
        "league_id",
        "promoted",
        "promotion_method",
    }

    missing = (
        required
        - set(indexes)
    )

    if missing:
        raise SpanishPerformanceCollectorError(
            "Faltam colunas na folha de equipas: "
            f"{sorted(missing)}"
        )

    teams: dict[
        str,
        dict[str, Any]
    ] = {}

    for row in worksheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        league_id = str(
            row[indexes["league_id"]]
            or ""
        ).strip().upper()

        if league_id != TARGET_LEAGUE_ID:
            continue

        keys = {
            normalize_text(
                row[indexes["team_name"]]
            ),
            normalize_text(
                row[indexes["short_name"]]
            ),
            normalize_text(
                row[indexes["normalized_name"]]
            ),
        }

        record = {
            "team_id": str(
                row[indexes["team_id"]]
                or ""
            ).strip(),
            "team_name": str(
                row[indexes["team_name"]]
                or ""
            ).strip(),
            "promoted": int(
                row[indexes["promoted"]]
                or 0
            ),
            "promotion_method": (
                str(
                    row[
                        indexes[
                            "promotion_method"
                        ]
                    ]
                    or ""
                )
                .strip()
                .upper()
                or None
            ),
        }

        for key in keys:
            if key:
                teams[key] = record

    aliases = {
        "atletico de madrid": "atletico de madrid",
        "deportivo alaves": "deportivo alaves",
        "rcd espanyol de barcelona": (
            "rcd espanyol de barcelona"
        ),
        "ca osasuna": "ca osasuna",
        "r racing club": "r racing club",
        "rc deportivo": "rc deportivo",
        "malaga cf": "malaga cf",
    }

    for alias, target_key in aliases.items():
        if target_key in teams:
            teams[alias] = teams[
                target_key
            ]

    return teams


def remove_existing_rows(
    worksheet,
    headers: list[Any],
) -> int:
    indexes = {
        str(header): index
        for index, header in enumerate(
            headers
        )
        if header is not None
    }

    rows_to_delete: list[int] = []

    for row_number in range(
        worksheet.max_row,
        1,
        -1,
    ):
        target_league_id = worksheet.cell(
            row=row_number,
            column=(
                indexes[
                    "target_league_id"
                ]
                + 1
            ),
        ).value

        if (
            str(target_league_id or "")
            .strip()
            .upper()
            == TARGET_LEAGUE_ID
        ):
            rows_to_delete.append(
                row_number
            )

    for row_number in rows_to_delete:
        worksheet.delete_rows(
            row_number,
            1,
        )

    return len(rows_to_delete)


def update_promoted_sheet(
    workbook,
    promoted_records: list[dict[str, Any]],
) -> None:
    worksheet = workbook[
        PROMOTED_SHEET_NAME
    ]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    indexes = {
        str(header): index
        for index, header in enumerate(
            headers
        )
        if header is not None
    }

    by_team_id = {
        record["team_id"]: record
        for record in promoted_records
    }

    updated = 0

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        team_id = str(
            worksheet.cell(
                row=row_number,
                column=(
                    indexes["team_id"]
                    + 1
                ),
            ).value
            or ""
        ).strip()

        target_league_id = str(
            worksheet.cell(
                row=row_number,
                column=(
                    indexes[
                        "target_league_id"
                    ]
                    + 1
                ),
            ).value
            or ""
        ).strip().upper()

        if target_league_id != TARGET_LEAGUE_ID:
            continue

        record = by_team_id.get(
            team_id
        )

        if record is None:
            continue

        values = {
            "source_league_id": "ESP2",
            "source_position": record[
                "position"
            ],
            "promotion_method": record[
                "promotion_method"
            ],
            "played": record["played"],
            "points": record["points"],
            "goals_for": record[
                "goals_for"
            ],
            "goals_against": record[
                "goals_against"
            ],
            "goal_difference": record[
                "goal_difference"
            ],
            "source_status": "CONFIRMED",
            "data_confidence": 1.0,
            "source_url": ESP2_SOURCE_URL,
        }

        for field, value in values.items():
            if field not in indexes:
                continue

            worksheet.cell(
                row=row_number,
                column=indexes[field] + 1,
                value=value,
            )

        updated += 1

    if updated != 3:
        raise SpanishPerformanceCollectorError(
            "Esperavam-se 3 promovidas atualizadas, "
            f"mas foram atualizadas {updated}."
        )


def write_performance(
    esp1_standings: list[dict[str, Any]],
    esp2_standings: list[dict[str, Any]],
) -> None:
    if not DATASET_PATH.exists():
        raise SpanishPerformanceCollectorError(
            f"Dataset inexistente: {DATASET_PATH}"
        )

    backup_path = DATASET_PATH.with_name(
        f"{DATASET_PATH.stem}"
        "_BACKUP_BEFORE_ESP1"
        f"{DATASET_PATH.suffix}"
    )

    shutil.copy2(
        DATASET_PATH,
        backup_path,
    )

    workbook = load_workbook(
        DATASET_PATH
    )

    try:
        for required_sheet in (
            TEAMS_SHEET_NAME,
            PROMOTED_SHEET_NAME,
            PERFORMANCE_SHEET_NAME,
        ):
            if required_sheet not in workbook.sheetnames:
                raise SpanishPerformanceCollectorError(
                    f"Falta a folha {required_sheet}."
                )

        target_teams = load_target_teams(
            workbook
        )

        worksheet = workbook[
            PERFORMANCE_SHEET_NAME
        ]

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        required_headers = {
            "team_id",
            "source_league_id",
            "target_league_id",
            "season_label",
            "position",
            "played",
            "wins",
            "draws",
            "losses",
            "goals_for",
            "goals_against",
            "goal_difference",
            "points",
            "points_adjustment",
            "promoted",
            "promotion_method",
            "source_status",
            "data_confidence",
            "source_url",
            "accessed_at",
        }

        existing_headers = {
            str(header)
            for header in headers
            if header is not None
        }

        missing_headers = (
            required_headers
            - existing_headers
        )

        if missing_headers:
            raise SpanishPerformanceCollectorError(
                "Faltam colunas na folha de performance: "
                f"{sorted(missing_headers)}"
            )

        removed = remove_existing_rows(
            worksheet,
            headers,
        )

        accessed_at = datetime.now(
            timezone.utc
        ).replace(
            microsecond=0
        ).isoformat()

        permanent_written = 0
        relegated_skipped = 0
        promoted_written = 0
        missing_target: list[str] = []

        for performance in esp1_standings:
            lookup_name = performance[
                "lookup_name"
            ]

            if lookup_name in RELEGATED_ESP1_TEAMS:
                relegated_skipped += 1
                continue

            target_team = target_teams.get(
                lookup_name
            )

            if target_team is None:
                missing_target.append(
                    performance[
                        "source_team_name"
                    ]
                )
                continue

            if target_team["promoted"] != 0:
                raise SpanishPerformanceCollectorError(
                    "Uma equipa da ESP1 anterior está "
                    "marcada como promovida: "
                    f"{performance['source_team_name']}"
                )

            row_data = {
                "team_id": target_team[
                    "team_id"
                ],
                "source_league_id": "ESP1",
                "target_league_id": "ESP1",
                "season_label": SEASON_LABEL,
                "position": performance[
                    "position"
                ],
                "played": performance[
                    "played"
                ],
                "wins": performance["wins"],
                "draws": performance[
                    "draws"
                ],
                "losses": performance[
                    "losses"
                ],
                "goals_for": performance[
                    "goals_for"
                ],
                "goals_against": performance[
                    "goals_against"
                ],
                "goal_difference": performance[
                    "goal_difference"
                ],
                "points": performance[
                    "points"
                ],
                "points_adjustment": 0,
                "promoted": 0,
                "promotion_method": None,
                "source_status": "CONFIRMED",
                "data_confidence": 1.0,
                "source_url": ESP1_SOURCE_URL,
                "accessed_at": accessed_at,
            }

            worksheet.append(
                [
                    row_data.get(
                        str(header)
                    )
                    for header in headers
                ]
            )

            permanent_written += 1

        promoted_records: list[
            dict[str, Any]
        ] = []

        for performance in esp2_standings:
            promoted_config = (
                PROMOTED_TEAMS.get(
                    performance[
                        "lookup_name"
                    ]
                )
            )

            if promoted_config is None:
                continue

            target_team = target_teams.get(
                performance[
                    "lookup_name"
                ]
            )

            if target_team is None:
                raise SpanishPerformanceCollectorError(
                    "Não foi possível mapear a promovida "
                    f"{performance['source_team_name']}."
                )

            if (
                target_team["team_id"]
                != promoted_config["team_id"]
            ):
                raise SpanishPerformanceCollectorError(
                    "team_id inesperado para "
                    f"{performance['source_team_name']}."
                )

            if target_team["promoted"] != 1:
                raise SpanishPerformanceCollectorError(
                    "A equipa promovida não está marcada "
                    f"como promovida: "
                    f"{performance['source_team_name']}."
                )

            expected_method = promoted_config[
                "promotion_method"
            ]

            if (
                target_team["promotion_method"]
                != expected_method
            ):
                raise SpanishPerformanceCollectorError(
                    "Método de promoção inesperado para "
                    f"{performance['source_team_name']}."
                )

            row_data = {
                "team_id": target_team[
                    "team_id"
                ],
                "source_league_id": "ESP2",
                "target_league_id": "ESP1",
                "season_label": SEASON_LABEL,
                "position": performance[
                    "position"
                ],
                "played": performance[
                    "played"
                ],
                "wins": performance["wins"],
                "draws": performance[
                    "draws"
                ],
                "losses": performance[
                    "losses"
                ],
                "goals_for": performance[
                    "goals_for"
                ],
                "goals_against": performance[
                    "goals_against"
                ],
                "goal_difference": performance[
                    "goal_difference"
                ],
                "points": performance[
                    "points"
                ],
                "points_adjustment": 0,
                "promoted": 1,
                "promotion_method": expected_method,
                "source_status": "CONFIRMED",
                "data_confidence": 1.0,
                "source_url": ESP2_SOURCE_URL,
                "accessed_at": accessed_at,
            }

            worksheet.append(
                [
                    row_data.get(
                        str(header)
                    )
                    for header in headers
                ]
            )

            promoted_records.append(
                {
                    **performance,
                    "team_id": target_team[
                        "team_id"
                    ],
                    "promotion_method": (
                        expected_method
                    ),
                }
            )

            promoted_written += 1

        if missing_target:
            raise SpanishPerformanceCollectorError(
                "Não foi possível mapear estas equipas: "
                + ", ".join(
                    missing_target
                )
            )

        if permanent_written != 17:
            raise SpanishPerformanceCollectorError(
                "Esperavam-se 17 equipas permanentes "
                f"gravadas, mas foram gravadas "
                f"{permanent_written}."
            )

        if relegated_skipped != 3:
            raise SpanishPerformanceCollectorError(
                "Esperavam-se 3 despromovidas ignoradas, "
                f"mas foram ignoradas "
                f"{relegated_skipped}."
            )

        if promoted_written != 3:
            raise SpanishPerformanceCollectorError(
                "Esperavam-se 3 promovidas gravadas, "
                f"mas foram gravadas "
                f"{promoted_written}."
            )

        update_promoted_sheet(
            workbook,
            promoted_records,
        )

        workbook.save(
            DATASET_PATH
        )

        print()
        print("=" * 100)
        print(
            "ESP1 — PERFORMANCE GRAVADA NO DATASET"
        )
        print("=" * 100)
        print(
            f"Dataset: {DATASET_PATH.resolve()}"
        )
        print(
            f"Backup:  {backup_path.resolve()}"
        )
        print(
            f"Linhas ESP1 removidas: {removed}"
        )
        print(
            "Equipas permanentes gravadas: "
            f"{permanent_written}"
        )
        print(
            "Despromovidas ignoradas: "
            f"{relegated_skipped}"
        )
        print(
            "Promovidas ESP2 gravadas: "
            f"{promoted_written}"
        )
        print(
            "Total ESP1 2026/27 gravado: "
            f"{permanent_written + promoted_written}"
        )
        print("=" * 100)

    except Exception:
        workbook.close()

        shutil.copy2(
            backup_path,
            DATASET_PATH,
        )

        raise

    else:
        workbook.close()


def main() -> int:
    print()
    print("=" * 100)
    print(
        "FOOTWIN SPORTS — RECOLHA ESP1 2025/26"
    )
    print("=" * 100)

    esp1_standings = parse_standings(
        ESP1_JSON_PATH,
        expected_teams=20,
    )

    esp2_standings = parse_standings(
        ESP2_JSON_PATH,
        expected_teams=22,
    )

    print(
        "Equipas ESP1 recolhidas: "
        f"{len(esp1_standings)}"
    )

    print(
        "Equipas ESP2 recolhidas: "
        f"{len(esp2_standings)}"
    )

    write_performance(
        esp1_standings,
        esp2_standings,
    )

    print()
    print(
        "✅ ESP1 concluída com sucesso."
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(
        main()
    )
