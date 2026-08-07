# -*- coding: utf-8 -*-

from __future__ import annotations

import re
import shutil
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from bs4 import BeautifulSoup
from openpyxl import load_workbook


DATASET_PATH = Path(
    "data/input/FOOTWIN_Dataset_2026_27_V001.xlsx"
)

HTML_PATH = Path(
    "data/raw/performance_pages/GER1.html"
)

TEAMS_SHEET_NAME = "Equipas_2026_27"
PERFORMANCE_SHEET_NAME = "Desempenho_2025_26"

SOURCE_LEAGUE_ID = "GER1"
TARGET_LEAGUE_ID = "GER1"
SEASON_LABEL = "2025/26"

SOURCE_URL = (
    "https://www.bundesliga.com/en/bundesliga/table"
)

EXCLUDED_RELEGATED_TEAMS = {
    "wolfsburg",
    "heidenheim",
    "st pauli",
}

EXPECTED_PROMOTED_TEAMS = {
    "schalke",
    "elversberg",
    "paderborn",
}

NAME_ALIASES = {
    "bayern munich": "bayern",
    "borussia dortmund": "dortmund",
    "rb leipzig": "rb leipzig",
    "vfb stuttgart": "stuttgart",
    "hoffenheim": "hoffenheim",
    "bayer leverkusen": "leverkusen",
    "freiburg": "freiburg",
    "eintracht frankfurt": "frankfurt",
    "augsburg": "augsburg",
    "mainz": "mainz",
    "union berlin": "union berlin",
    "borussia monchengladbach": "m gladbach",
    "hamburg": "hamburg",
    "cologne": "koln",
    "werder bremen": "werder",
    "wolfsburg": "wolfsburg",
    "heidenheim": "heidenheim",
    "st pauli": "st pauli",
}


class GermanPerformanceCollectorError(
    RuntimeError
):
    """Erro na recolha ou escrita da performance GER1."""


def normalize_text(
    value: Any,
) -> str:
    text = str(value or "").strip().casefold()

    text = unicodedata.normalize(
        "NFKD",
        text,
    )

    text = "".join(
        character
        for character in text
        if not unicodedata.combining(character)
    )

    text = text.replace(
        "ß",
        "ss",
    )

    text = re.sub(
        r"[^a-z0-9]+",
        " ",
        text,
    )

    return re.sub(
        r"\s+",
        " ",
        text,
    ).strip()


def extract_standings() -> list[dict[str, Any]]:
    if not HTML_PATH.exists():
        raise GermanPerformanceCollectorError(
            f"HTML inexistente: {HTML_PATH}"
        )

    html = HTML_PATH.read_text(
        encoding="utf-8",
        errors="ignore",
    )

    soup = BeautifulSoup(
        html,
        "html.parser",
    )

    tables = soup.find_all(
        "table"
    )

    if not tables:
        raise GermanPerformanceCollectorError(
            "Não foram encontradas tabelas no HTML GER1."
        )

    target_table = None

    for table in tables:
        text = normalize_text(
            table.get_text(
                " ",
                strip=True,
            )
        )

        if (
            "won draw lost" in text
            and "pts" in text
        ):
            target_table = table
            break

    if target_table is None:
        raise GermanPerformanceCollectorError(
            "Não foi encontrada a tabela da Bundesliga."
        )

    standings: list[
        dict[str, Any]
    ] = []

    rows = target_table.find_all(
        "tr"
    )

    for row in rows:
        cells = row.find_all(
            ["th", "td"]
        )

        values = [
            re.sub(
                r"\s+",
                " ",
                cell.get_text(
                    " ",
                    strip=True,
                ),
            ).strip()
            for cell in cells
        ]

        values = [
            value
            for value in values
            if value
        ]

        if len(values) != 7:
            continue

        if not values[0].isdigit():
            continue

        position = int(
            values[0]
        )

        raw_team_name = values[1]

        played = int(
            values[2]
        )

        wins, draws, losses = (
            int(part)
            for part in values[3].split("-")
        )

        goals_for, goals_against = (
            int(part)
            for part in values[4].split(":")
        )

        goal_difference = int(
            values[5]
        )

        points = int(
            values[6]
        )

        normalized_raw_name = normalize_text(
            raw_team_name
        )

        matched_alias = None

        for alias in NAME_ALIASES:
            if normalize_text(alias) in normalized_raw_name:
                matched_alias = alias
                break

        if matched_alias is None:
            raise GermanPerformanceCollectorError(
                "Não foi possível identificar a equipa: "
                f"{raw_team_name}"
            )

        record = {
            "source_team_name": raw_team_name,
            "lookup_name": NAME_ALIASES[
                matched_alias
            ],
            "position": position,
            "played": played,
            "wins": wins,
            "draws": draws,
            "losses": losses,
            "goals_for": goals_for,
            "goals_against": goals_against,
            "goal_difference": goal_difference,
            "points": points,
        }

        if (
            played
            != wins
            + draws
            + losses
        ):
            raise GermanPerformanceCollectorError(
                "Totais de jogos incoerentes para "
                f"{raw_team_name}."
            )

        if (
            goal_difference
            != goals_for
            - goals_against
        ):
            raise GermanPerformanceCollectorError(
                "Diferença de golos incoerente para "
                f"{raw_team_name}."
            )

        expected_points = (
            3 * wins
            + draws
        )

        if points != expected_points:
            raise GermanPerformanceCollectorError(
                "Pontuação incoerente para "
                f"{raw_team_name}: esperados "
                f"{expected_points}, encontrados "
                f"{points}."
            )

        standings.append(
            record
        )

    if len(standings) != 18:
        raise GermanPerformanceCollectorError(
            "Esperavam-se 18 equipas GER1, "
            f"mas foram extraídas {len(standings)}."
        )

    return standings


def load_target_teams(
    workbook,
) -> dict[str, dict[str, Any]]:
    worksheet = workbook[
        TEAMS_SHEET_NAME
    ]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    indexes = {
        str(header): index
        for index, header in enumerate(
            headers
        )
        if header is not None
    }

    required = {
        "team_id",
        "team_name",
        "short_name",
        "normalized_name",
        "league_id",
        "promoted",
        "promotion_method",
    }

    missing = (
        required
        - set(indexes)
    )

    if missing:
        raise GermanPerformanceCollectorError(
            "Faltam colunas na folha de equipas: "
            f"{sorted(missing)}"
        )

    teams: dict[
        str,
        dict[str, Any]
    ] = {}

    for row in worksheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        league_id = normalize_text(
            row[indexes["league_id"]]
        ).upper()

        if league_id != TARGET_LEAGUE_ID:
            continue

        team_name = normalize_text(
            row[indexes["team_name"]]
        )

        short_name = normalize_text(
            row[indexes["short_name"]]
        )

        normalized_name = normalize_text(
            row[indexes["normalized_name"]]
        )

        keys = {
            team_name,
            short_name,
            normalized_name,
        }

        if "bayern" in short_name:
            keys.add("bayern")

        if "dortmund" in short_name:
            keys.add("dortmund")

        if "leipzig" in short_name:
            keys.add("rb leipzig")

        if "stuttgart" in short_name:
            keys.add("stuttgart")

        if "hoffenheim" in short_name:
            keys.add("hoffenheim")

        if "leverkusen" in short_name:
            keys.add("leverkusen")

        if "freiburg" in short_name:
            keys.add("freiburg")

        if "frankfurt" in short_name:
            keys.add("frankfurt")

        if "augsburg" in short_name:
            keys.add("augsburg")

        if "mainz" in short_name:
            keys.add("mainz")

        if "union berlin" in short_name:
            keys.add("union berlin")

        if "gladbach" in short_name:
            keys.add("m gladbach")

        if "hamburg" in short_name:
            keys.add("hamburg")

        if "koln" in short_name:
            keys.add("koln")

        if "werder" in short_name:
            keys.add("werder")

        if "schalke" in short_name:
            keys.add("schalke")

        if "elversberg" in short_name:
            keys.add("elversberg")

        if "paderborn" in short_name:
            keys.add("paderborn")

        record = {
            "team_id": row[
                indexes["team_id"]
            ],
            "promoted": int(
                row[indexes["promoted"]]
                or 0
            ),
            "promotion_method": (
                str(
                    row[
                        indexes[
                            "promotion_method"
                        ]
                    ]
                    or ""
                )
                .strip()
                .upper()
                or None
            ),
        }

        for key in keys:
            if key:
                teams[key] = record

    return teams


def remove_existing_rows(
    worksheet,
    headers: list[Any],
) -> int:
    indexes = {
        str(header): index
        for index, header in enumerate(
            headers
        )
        if header is not None
    }

    rows_to_delete: list[int] = []

    for row_number in range(
        worksheet.max_row,
        1,
        -1,
    ):
        target_league_id = worksheet.cell(
            row=row_number,
            column=(
                indexes["target_league_id"]
                + 1
            ),
        ).value

        if (
            str(target_league_id or "")
            .strip()
            .upper()
            == TARGET_LEAGUE_ID
        ):
            rows_to_delete.append(
                row_number
            )

    for row_number in rows_to_delete:
        worksheet.delete_rows(
            row_number,
            1,
        )

    return len(rows_to_delete)


def write_performance(
    standings: list[dict[str, Any]],
) -> None:
    if not DATASET_PATH.exists():
        raise GermanPerformanceCollectorError(
            f"Dataset inexistente: {DATASET_PATH}"
        )

    backup_path = DATASET_PATH.with_name(
        f"{DATASET_PATH.stem}"
        "_BACKUP_BEFORE_GER1"
        f"{DATASET_PATH.suffix}"
    )

    shutil.copy2(
        DATASET_PATH,
        backup_path,
    )

    workbook = load_workbook(
        DATASET_PATH
    )

    try:
        target_teams = load_target_teams(
            workbook
        )

        worksheet = workbook[
            PERFORMANCE_SHEET_NAME
        ]

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        required_headers = {
            "team_id",
            "source_league_id",
            "target_league_id",
            "season_label",
            "position",
            "played",
            "wins",
            "draws",
            "losses",
            "goals_for",
            "goals_against",
            "goal_difference",
            "points",
            "points_adjustment",
            "promoted",
            "promotion_method",
            "source_status",
            "data_confidence",
            "source_url",
            "accessed_at",
        }

        existing_headers = {
            str(header)
            for header in headers
            if header is not None
        }

        missing_headers = (
            required_headers
            - existing_headers
        )

        if missing_headers:
            raise GermanPerformanceCollectorError(
                "Faltam colunas na folha de performance: "
                f"{sorted(missing_headers)}"
            )

        removed = remove_existing_rows(
            worksheet,
            headers,
        )

        accessed_at = datetime.now(
            timezone.utc
        ).replace(
            microsecond=0
        ).isoformat()

        written = 0
        skipped_relegated = 0
        missing_target: list[str] = []

        for performance in standings:
            lookup_name = performance[
                "lookup_name"
            ]

            if lookup_name in EXCLUDED_RELEGATED_TEAMS:
                skipped_relegated += 1
                continue

            target_team = target_teams.get(
                lookup_name
            )

            if target_team is None:
                missing_target.append(
                    performance[
                        "source_team_name"
                    ]
                )
                continue

            if target_team["promoted"] != 0:
                raise GermanPerformanceCollectorError(
                    "Uma equipa GER1 recolhida está "
                    "marcada como promovida: "
                    f"{performance['source_team_name']}"
                )

            row_data = {
                "team_id": target_team[
                    "team_id"
                ],
                "source_league_id": (
                    SOURCE_LEAGUE_ID
                ),
                "target_league_id": (
                    TARGET_LEAGUE_ID
                ),
                "season_label": (
                    SEASON_LABEL
                ),
                "position": (
                    performance["position"]
                ),
                "played": (
                    performance["played"]
                ),
                "wins": (
                    performance["wins"]
                ),
                "draws": (
                    performance["draws"]
                ),
                "losses": (
                    performance["losses"]
                ),
                "goals_for": (
                    performance["goals_for"]
                ),
                "goals_against": (
                    performance[
                        "goals_against"
                    ]
                ),
                "goal_difference": (
                    performance[
                        "goal_difference"
                    ]
                ),
                "points": (
                    performance["points"]
                ),
                "points_adjustment": 0,
                "promoted": 0,
                "promotion_method": None,
                "source_status": "CONFIRMED",
                "data_confidence": 1.0,
                "source_url": SOURCE_URL,
                "accessed_at": accessed_at,
            }

            worksheet.append(
                [
                    row_data.get(
                        str(header)
                    )
                    for header in headers
                ]
            )

            written += 1

        if missing_target:
            raise GermanPerformanceCollectorError(
                "Não foi possível mapear: "
                + ", ".join(
                    missing_target
                )
            )

        if written != 15:
            raise GermanPerformanceCollectorError(
                "Esperavam-se 15 equipas GER1 "
                f"gravadas, mas foram gravadas {written}."
            )

        promoted_found = {
            name
            for name in EXPECTED_PROMOTED_TEAMS
            if name in target_teams
        }

        if promoted_found != EXPECTED_PROMOTED_TEAMS:
            raise GermanPerformanceCollectorError(
                "Não foram encontradas todas as "
                "promovidas GER1 no dataset."
            )

        workbook.save(
            DATASET_PATH
        )

        print()
        print("=" * 100)
        print(
            "GER1 — PERFORMANCE GRAVADA NO DATASET"
        )
        print("=" * 100)
        print(
            f"Dataset: {DATASET_PATH.resolve()}"
        )
        print(
            f"Backup:  {backup_path.resolve()}"
        )
        print(
            f"Linhas GER1 removidas: {removed}"
        )
        print(
            f"Equipas GER1 gravadas: {written}"
        )
        print(
            f"Relegadas ignoradas: {skipped_relegated}"
        )
        print(
            "Promovidas pendentes da GER2: "
            "Schalke, Elversberg, Paderborn"
        )
        print("=" * 100)

    except Exception:
        workbook.close()

        shutil.copy2(
            backup_path,
            DATASET_PATH,
        )

        raise

    else:
        workbook.close()


def main() -> int:
    print()
    print("=" * 100)
    print(
        "FOOTWIN SPORTS — RECOLHA GER1 2025/26"
    )
    print("=" * 100)

    standings = extract_standings()

    print(
        f"Equipas recolhidas: {len(standings)}"
    )

    write_performance(
        standings
    )

    print()
    print(
        "✅ GER1 concluída com sucesso."
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(
        main()
    )
