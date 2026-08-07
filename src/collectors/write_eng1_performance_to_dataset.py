# -*- coding: utf-8 -*-

from __future__ import annotations

import json
import re
import shutil
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DATASET_PATH = Path(
    "data/input/FOOTWIN_Dataset_2026_27_V001.xlsx"
)

JSON_PATH = Path(
    "data/raw/performance_pages/"
    "ENG1_2025_26_standings.json"
)

TEAMS_SHEET_NAME = "Equipas_2026_27"
PERFORMANCE_SHEET_NAME = "Desempenho_2025_26"

SOURCE_LEAGUE_ID = "ENG1"
TARGET_LEAGUE_ID = "ENG1"
SEASON_LABEL = "2025/26"

SOURCE_URL = (
    "https://footballapi.pulselive.com/"
    "football/standings"
    "?comp=1"
    "&compSeasons=777"
    "&altIds=true"
    "&page=0"
    "&pageSize=100"
)

EXCLUDED_RELEGATED_TEAMS = {
    "west ham united",
    "burnley",
    "wolverhampton wanderers",
}

EXPECTED_PROMOTED_TEAMS = {
    "coventry city",
    "ipswich town",
    "hull city",
}

NAME_ALIASES = {
    "arsenal": "arsenal",
    "manchester city": "manchester city",
    "manchester united": "manchester united",
    "aston villa": "aston villa",
    "liverpool": "liverpool",
    "bournemouth": "afc bournemouth",
    "sunderland": "sunderland",
    "brighton and hove albion": "brighton and hove albion",
    "brentford": "brentford",
    "chelsea": "chelsea",
    "fulham": "fulham",
    "newcastle united": "newcastle united",
    "everton": "everton",
    "leeds united": "leeds united",
    "crystal palace": "crystal palace",
    "nottingham forest": "nottingham forest",
    "tottenham hotspur": "tottenham hotspur",
    "west ham united": "west ham united",
    "burnley": "burnley",
    "wolverhampton wanderers": (
        "wolverhampton wanderers"
    ),
}


class EnglishPerformanceCollectorError(
    RuntimeError
):
    """Erro na recolha ou escrita da performance ENG1."""


def normalize_text(
    value: Any,
) -> str:
    text = str(value or "").strip().casefold()

    text = unicodedata.normalize(
        "NFKD",
        text,
    )

    text = "".join(
        character
        for character in text
        if not unicodedata.combining(character)
    )

    text = text.replace(
        "&",
        " and ",
    )

    text = re.sub(
        r"[^a-z0-9]+",
        " ",
        text,
    )

    return re.sub(
        r"\s+",
        " ",
        text,
    ).strip()


def extract_standings() -> list[dict[str, Any]]:
    if not JSON_PATH.exists():
        raise EnglishPerformanceCollectorError(
            f"JSON inexistente: {JSON_PATH}"
        )

    try:
        data = json.loads(
            JSON_PATH.read_text(
                encoding="utf-8",
            )
        )
    except json.JSONDecodeError as exc:
        raise EnglishPerformanceCollectorError(
            "O ficheiro ENG1 não contém JSON válido."
        ) from exc

    tables = data.get(
        "tables",
        []
    )

    if not isinstance(
        tables,
        list,
    ) or not tables:
        raise EnglishPerformanceCollectorError(
            "O JSON ENG1 não contém tabelas."
        )

    target_table = None

    for table in tables:
        entries = table.get(
            "entries",
            []
        )

        game_week = table.get(
            "gameWeek"
        )

        if (
            isinstance(entries, list)
            and entries
            and game_week == 38
        ):
            target_table = table
            break

    if target_table is None:
        raise EnglishPerformanceCollectorError(
            "Não foi encontrada a classificação "
            "final da jornada 38."
        )

    entries = target_table.get(
        "entries",
        []
    )

    standings: list[
        dict[str, Any]
    ] = []

    for entry in entries:
        team = entry.get(
            "team",
            {}
        )

        overall = entry.get(
            "overall",
            {}
        )

        raw_team_name = str(
            team.get("name")
            or team.get("shortName")
            or ""
        ).strip()

        if not raw_team_name:
            raise EnglishPerformanceCollectorError(
                "Foi encontrada uma equipa sem nome."
            )

        normalized_raw_name = normalize_text(
            raw_team_name
        )

        lookup_name = NAME_ALIASES.get(
            normalized_raw_name
        )

        if lookup_name is None:
            raise EnglishPerformanceCollectorError(
                "Não foi possível identificar a equipa: "
                f"{raw_team_name}"
            )

        try:
            position = int(
                entry["position"]
            )
            played = int(
                overall["played"]
            )
            wins = int(
                overall["won"]
            )
            draws = int(
                overall["drawn"]
            )
            losses = int(
                overall["lost"]
            )
            goals_for = int(
                overall["goalsFor"]
            )
            goals_against = int(
                overall["goalsAgainst"]
            )
            goal_difference = int(
                overall["goalsDifference"]
            )
            points = int(
                overall["points"]
            )
        except (
            KeyError,
            TypeError,
            ValueError,
        ) as exc:
            raise EnglishPerformanceCollectorError(
                "Dados incompletos ou inválidos para "
                f"{raw_team_name}."
            ) from exc

        if (
            played
            != wins
            + draws
            + losses
        ):
            raise EnglishPerformanceCollectorError(
                "Totais de jogos incoerentes para "
                f"{raw_team_name}."
            )

        if (
            goal_difference
            != goals_for
            - goals_against
        ):
            raise EnglishPerformanceCollectorError(
                "Diferença de golos incoerente para "
                f"{raw_team_name}."
            )

        expected_points = (
            3 * wins
            + draws
        )

        if points != expected_points:
            raise EnglishPerformanceCollectorError(
                "Pontuação incoerente para "
                f"{raw_team_name}: esperados "
                f"{expected_points}, encontrados "
                f"{points}."
            )

        standings.append(
            {
                "source_team_name": raw_team_name,
                "lookup_name": lookup_name,
                "position": position,
                "played": played,
                "wins": wins,
                "draws": draws,
                "losses": losses,
                "goals_for": goals_for,
                "goals_against": goals_against,
                "goal_difference": goal_difference,
                "points": points,
            }
        )

    if len(standings) != 20:
        raise EnglishPerformanceCollectorError(
            "Esperavam-se 20 equipas ENG1, "
            f"mas foram extraídas {len(standings)}."
        )

    positions = {
        record["position"]
        for record in standings
    }

    if positions != set(
        range(1, 21)
    ):
        raise EnglishPerformanceCollectorError(
            "As posições ENG1 não correspondem "
            "ao intervalo completo de 1 a 20."
        )

    return standings


def load_target_teams(
    workbook,
) -> dict[str, dict[str, Any]]:
    worksheet = workbook[
        TEAMS_SHEET_NAME
    ]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    indexes = {
        str(header): index
        for index, header in enumerate(
            headers
        )
        if header is not None
    }

    required = {
        "team_id",
        "team_name",
        "short_name",
        "normalized_name",
        "league_id",
        "promoted",
        "promotion_method",
    }

    missing = (
        required
        - set(indexes)
    )

    if missing:
        raise EnglishPerformanceCollectorError(
            "Faltam colunas na folha de equipas: "
            f"{sorted(missing)}"
        )

    teams: dict[
        str,
        dict[str, Any]
    ] = {}

    for row in worksheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        league_id = str(
            row[indexes["league_id"]]
            or ""
        ).strip().upper()

        if league_id != TARGET_LEAGUE_ID:
            continue

        team_name = normalize_text(
            row[indexes["team_name"]]
        )

        short_name = normalize_text(
            row[indexes["short_name"]]
        )

        normalized_name = normalize_text(
            row[indexes["normalized_name"]]
        )

        keys = {
            team_name,
            short_name,
            normalized_name,
        }

        if "bournemouth" in team_name:
            keys.add(
                "afc bournemouth"
            )
            keys.add(
                "bournemouth"
            )

        if "brighton" in team_name:
            keys.add(
                "brighton hove albion"
            )

        if "manchester city" in team_name:
            keys.add(
                "manchester city"
            )

        if "manchester united" in team_name:
            keys.add(
                "manchester united"
            )

        if "newcastle" in team_name:
            keys.add(
                "newcastle united"
            )

        if "nottingham forest" in team_name:
            keys.add(
                "nottingham forest"
            )

        if "tottenham" in team_name:
            keys.add(
                "tottenham hotspur"
            )

        if "coventry" in team_name:
            keys.add(
                "coventry city"
            )

        if "ipswich" in team_name:
            keys.add(
                "ipswich town"
            )

        if "hull" in team_name:
            keys.add(
                "hull city"
            )

        record = {
            "team_id": row[
                indexes["team_id"]
            ],
            "team_name": row[
                indexes["team_name"]
            ],
            "promoted": int(
                row[indexes["promoted"]]
                or 0
            ),
            "promotion_method": (
                str(
                    row[
                        indexes[
                            "promotion_method"
                        ]
                    ]
                    or ""
                )
                .strip()
                .upper()
                or None
            ),
        }

        for key in keys:
            if key:
                teams[key] = record

    return teams


def remove_existing_rows(
    worksheet,
    headers: list[Any],
) -> int:
    indexes = {
        str(header): index
        for index, header in enumerate(
            headers
        )
        if header is not None
    }

    rows_to_delete: list[int] = []

    for row_number in range(
        worksheet.max_row,
        1,
        -1,
    ):
        target_league_id = worksheet.cell(
            row=row_number,
            column=(
                indexes["target_league_id"]
                + 1
            ),
        ).value

        if (
            str(target_league_id or "")
            .strip()
            .upper()
            == TARGET_LEAGUE_ID
        ):
            rows_to_delete.append(
                row_number
            )

    for row_number in rows_to_delete:
        worksheet.delete_rows(
            row_number,
            1,
        )

    return len(rows_to_delete)


def write_performance(
    standings: list[dict[str, Any]],
) -> None:
    if not DATASET_PATH.exists():
        raise EnglishPerformanceCollectorError(
            f"Dataset inexistente: {DATASET_PATH}"
        )

    backup_path = DATASET_PATH.with_name(
        f"{DATASET_PATH.stem}"
        "_BACKUP_BEFORE_ENG1"
        f"{DATASET_PATH.suffix}"
    )

    shutil.copy2(
        DATASET_PATH,
        backup_path,
    )

    workbook = load_workbook(
        DATASET_PATH
    )

    try:
        target_teams = load_target_teams(
            workbook
        )

        worksheet = workbook[
            PERFORMANCE_SHEET_NAME
        ]

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        required_headers = {
            "team_id",
            "source_league_id",
            "target_league_id",
            "season_label",
            "position",
            "played",
            "wins",
            "draws",
            "losses",
            "goals_for",
            "goals_against",
            "goal_difference",
            "points",
            "points_adjustment",
            "promoted",
            "promotion_method",
            "source_status",
            "data_confidence",
            "source_url",
            "accessed_at",
        }

        existing_headers = {
            str(header)
            for header in headers
            if header is not None
        }

        missing_headers = (
            required_headers
            - existing_headers
        )

        if missing_headers:
            raise EnglishPerformanceCollectorError(
                "Faltam colunas na folha de performance: "
                f"{sorted(missing_headers)}"
            )

        removed = remove_existing_rows(
            worksheet,
            headers,
        )

        accessed_at = datetime.now(
            timezone.utc
        ).replace(
            microsecond=0
        ).isoformat()

        written = 0
        skipped_relegated = 0
        missing_target: list[str] = []

        for performance in standings:
            lookup_name = performance[
                "lookup_name"
            ]

            if (
                lookup_name
                in EXCLUDED_RELEGATED_TEAMS
            ):
                skipped_relegated += 1
                continue

            target_team = target_teams.get(
                lookup_name
            )

            if target_team is None:
                missing_target.append(
                    performance[
                        "source_team_name"
                    ]
                )
                continue

            if target_team["promoted"] != 0:
                raise EnglishPerformanceCollectorError(
                    "Uma equipa ENG1 recolhida está "
                    "marcada como promovida: "
                    f"{performance['source_team_name']}"
                )

            row_data = {
                "team_id": target_team[
                    "team_id"
                ],
                "source_league_id": (
                    SOURCE_LEAGUE_ID
                ),
                "target_league_id": (
                    TARGET_LEAGUE_ID
                ),
                "season_label": (
                    SEASON_LABEL
                ),
                "position": (
                    performance["position"]
                ),
                "played": (
                    performance["played"]
                ),
                "wins": (
                    performance["wins"]
                ),
                "draws": (
                    performance["draws"]
                ),
                "losses": (
                    performance["losses"]
                ),
                "goals_for": (
                    performance["goals_for"]
                ),
                "goals_against": (
                    performance[
                        "goals_against"
                    ]
                ),
                "goal_difference": (
                    performance[
                        "goal_difference"
                    ]
                ),
                "points": (
                    performance["points"]
                ),
                "points_adjustment": 0,
                "promoted": 0,
                "promotion_method": None,
                "source_status": "CONFIRMED",
                "data_confidence": 1.0,
                "source_url": SOURCE_URL,
                "accessed_at": accessed_at,
            }

            worksheet.append(
                [
                    row_data.get(
                        str(header)
                    )
                    for header in headers
                ]
            )

            written += 1

        if missing_target:
            raise EnglishPerformanceCollectorError(
                "Não foi possível mapear: "
                + ", ".join(
                    missing_target
                )
            )

        if written != 17:
            raise EnglishPerformanceCollectorError(
                "Esperavam-se 17 equipas ENG1 "
                f"gravadas, mas foram gravadas {written}."
            )

        if skipped_relegated != 3:
            raise EnglishPerformanceCollectorError(
                "Esperavam-se 3 equipas despromovidas "
                "ignoradas, mas foram ignoradas "
                f"{skipped_relegated}."
            )

        promoted_found = {
            name
            for name in EXPECTED_PROMOTED_TEAMS
            if (
                name in target_teams
                and target_teams[name]["promoted"] == 1
            )
        }

        if promoted_found != EXPECTED_PROMOTED_TEAMS:
            missing_promoted = (
                EXPECTED_PROMOTED_TEAMS
                - promoted_found
            )

            raise EnglishPerformanceCollectorError(
                "Não foram encontradas todas as "
                "promovidas ENG1 no dataset: "
                f"{sorted(missing_promoted)}"
            )

        workbook.save(
            DATASET_PATH
        )

        print()
        print("=" * 100)
        print(
            "ENG1 — PERFORMANCE GRAVADA NO DATASET"
        )
        print("=" * 100)
        print(
            f"Dataset: {DATASET_PATH.resolve()}"
        )
        print(
            f"Backup:  {backup_path.resolve()}"
        )
        print(
            f"Linhas ENG1 removidas: {removed}"
        )
        print(
            f"Equipas ENG1 gravadas: {written}"
        )
        print(
            f"Despromovidas ignoradas: "
            f"{skipped_relegated}"
        )
        print(
            "Promovidas pendentes da ENG2: "
            "Coventry City, Ipswich Town, Hull City"
        )
        print("=" * 100)

    except Exception:
        workbook.close()

        shutil.copy2(
            backup_path,
            DATASET_PATH,
        )

        raise

    else:
        workbook.close()


def main() -> int:
    print()
    print("=" * 100)
    print(
        "FOOTWIN SPORTS — RECOLHA ENG1 2025/26"
    )
    print("=" * 100)

    standings = extract_standings()

    print(
        f"Equipas recolhidas: {len(standings)}"
    )

    write_performance(
        standings
    )

    print()
    print(
        "✅ ENG1 concluída com sucesso."
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(
        main()
    )
