# -*- coding: utf-8 -*-

from __future__ import annotations

import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from src.collectors.team_catalog import (
    TEAM_CATALOG,
    TeamCatalogEntry,
    validate_team_catalog,
)


DEFAULT_DATASET_PATH = Path(
    "data/input/FOOTWIN_Dataset_2026_27_V001.xlsx"
)

SHEET_NAME = "Promovidas"

EXPECTED_HEADERS = [
    "team_id",
    "target_league_id",
    "source_league_id",
    "source_position",
    "promotion_method",
    "played",
    "points",
    "goals_for",
    "goals_against",
    "goal_difference",
    "promotion_factor",
    "source_status",
    "data_confidence",
    "source_url",
]

EXPECTED_PROMOTED_COUNTS = {
    "ENG1": 3,
    "ESP1": 3,
    "ITA1": 3,
    "GER1": 3,
    "FRA1": 2,
    "POR1": 2,
}

EXPECTED_TOTAL_PROMOTED = 16


@dataclass(frozen=True)
class PromotedWriteResult:
    dataset_path: Path
    backup_path: Path
    rows_deleted: int
    rows_written: int
    league_counts: dict[str, int]


class PromotedWriteError(RuntimeError):
    """Erro ao gravar as equipas promovidas no dataset."""


def get_promoted_teams() -> tuple[TeamCatalogEntry, ...]:
    """
    Obtém todas as equipas marcadas como promovidas.
    """

    league_order = (
        "ENG1",
        "ESP1",
        "ITA1",
        "GER1",
        "FRA1",
        "POR1",
    )

    promoted_teams: list[TeamCatalogEntry] = []

    for league_id in league_order:
        entries = TEAM_CATALOG[
            league_id
        ]

        for entry in entries:
            if entry.promoted == 1:
                promoted_teams.append(
                    entry
                )

    return tuple(
        promoted_teams
    )


def validate_promoted_catalog(
    promoted_teams: tuple[TeamCatalogEntry, ...],
) -> None:
    """
    Valida as equipas promovidas existentes no catálogo.
    """

    if len(promoted_teams) != EXPECTED_TOTAL_PROMOTED:
        raise PromotedWriteError(
            "Total de equipas promovidas incorreto: "
            f"{len(promoted_teams)}. "
            f"Esperadas: {EXPECTED_TOTAL_PROMOTED}."
        )

    team_ids = [
        entry.team_id
        for entry in promoted_teams
    ]

    if len(team_ids) != len(
        set(team_ids)
    ):
        raise PromotedWriteError(
            "Existem team_id duplicados "
            "entre as equipas promovidas."
        )

    league_counts: dict[str, int] = {}

    for entry in promoted_teams:
        league_counts[
            entry.league_id
        ] = (
            league_counts.get(
                entry.league_id,
                0,
            )
            + 1
        )

        if entry.promoted != 1:
            raise PromotedWriteError(
                f"{entry.team_id}: "
                "a equipa não está marcada "
                "como promovida."
            )

        if entry.promotion_method not in {
            "CHAMPION",
            "DIRECT",
            "PLAYOFF",
        }:
            raise PromotedWriteError(
                f"{entry.team_id}: "
                "método de promoção inválido: "
                f"{entry.promotion_method}"
            )

        if (
            not entry.previous_division
            or entry.previous_division
            == entry.league_id
        ):
            raise PromotedWriteError(
                f"{entry.team_id}: "
                "liga de origem inválida."
            )

    if league_counts != EXPECTED_PROMOTED_COUNTS:
        raise PromotedWriteError(
            "Contagens de promovidas incorretas.\n"
            f"Esperado: {EXPECTED_PROMOTED_COUNTS}\n"
            f"Encontrado: {league_counts}"
        )


def build_backup_path(
    dataset_path: Path,
) -> Path:
    """
    Cria um caminho único para o backup.
    """

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    return dataset_path.with_name(
        f"{dataset_path.stem}"
        f"_BEFORE_PROMOTED_"
        f"{timestamp}"
        f"{dataset_path.suffix}"
    )


def create_backup(
    dataset_path: Path,
) -> Path:
    """
    Cria uma cópia integral do dataset.
    """

    backup_path = build_backup_path(
        dataset_path
    )

    shutil.copy2(
        dataset_path,
        backup_path,
    )

    if not backup_path.exists():
        raise PromotedWriteError(
            "O backup não foi criado."
        )

    if backup_path.stat().st_size <= 0:
        raise PromotedWriteError(
            "O backup criado está vazio."
        )

    return backup_path


def validate_headers(
    worksheet,
) -> None:
    """
    Confirma os cabeçalhos da folha Promovidas.
    """

    headers = [
        worksheet.cell(
            row=1,
            column=column,
        ).value
        for column in range(
            1,
            len(EXPECTED_HEADERS) + 1,
        )
    ]

    if headers != EXPECTED_HEADERS:
        raise PromotedWriteError(
            "Os cabeçalhos da folha "
            f"{SHEET_NAME} não correspondem "
            "ao formato esperado.\n"
            f"Esperado: {EXPECTED_HEADERS}\n"
            f"Encontrado: {headers}"
        )


def clear_existing_rows(
    worksheet,
) -> int:
    """
    Remove todos os registos abaixo do cabeçalho.
    """

    if worksheet.max_row <= 1:
        return 0

    rows_to_delete = (
        worksheet.max_row - 1
    )

    worksheet.delete_rows(
        2,
        rows_to_delete,
    )

    return rows_to_delete


def build_promoted_row(
    entry: TeamCatalogEntry,
) -> list:
    """
    Converte uma equipa promovida numa linha do Excel.

    Os dados estatísticos permanecem vazios porque ainda
    não foram recolhidos e validados nas fontes oficiais.
    """

    return [
        entry.team_id,
        entry.league_id,
        entry.previous_division,
        None,
        entry.promotion_method,
        None,
        None,
        None,
        None,
        None,
        None,
        "MISSING",
        0,
        None,
    ]


def write_promoted_rows(
    worksheet,
    promoted_teams: tuple[TeamCatalogEntry, ...],
) -> dict[str, int]:
    """
    Grava as equipas promovidas na folha.
    """

    league_counts: dict[str, int] = {}

    for row_number, entry in enumerate(
        promoted_teams,
        start=2,
    ):
        values = build_promoted_row(
            entry
        )

        for column_number, value in enumerate(
            values,
            start=1,
        ):
            worksheet.cell(
                row=row_number,
                column=column_number,
                value=value,
            )

        league_counts[
            entry.league_id
        ] = (
            league_counts.get(
                entry.league_id,
                0,
            )
            + 1
        )

    return league_counts


def validate_written_rows(
    worksheet,
) -> None:
    """
    Valida os registos gravados na folha Promovidas.
    """

    rows = list(
        worksheet.iter_rows(
            min_row=2,
            max_col=len(
                EXPECTED_HEADERS
            ),
            values_only=True,
        )
    )

    non_empty_rows = [
        row
        for row in rows
        if any(
            value is not None
            for value in row
        )
    ]

    if len(non_empty_rows) != EXPECTED_TOTAL_PROMOTED:
        raise PromotedWriteError(
            "Quantidade de promovidas gravada "
            f"incorretamente: {len(non_empty_rows)}. "
            f"Esperadas: {EXPECTED_TOTAL_PROMOTED}."
        )

    team_ids = [
        str(row[0]).strip()
        for row in non_empty_rows
    ]

    if len(team_ids) != len(
        set(team_ids)
    ):
        raise PromotedWriteError(
            "Existem team_id duplicados "
            "na folha Promovidas."
        )

    catalog_promoted_ids = {
        entry.team_id
        for entry in get_promoted_teams()
    }

    written_ids = set(
        team_ids
    )

    if written_ids != catalog_promoted_ids:
        missing = (
            catalog_promoted_ids
            - written_ids
        )

        extra = (
            written_ids
            - catalog_promoted_ids
        )

        raise PromotedWriteError(
            "Os IDs gravados não correspondem "
            "ao catálogo.\n"
            f"Em falta: {sorted(missing)}\n"
            f"Adicionais: {sorted(extra)}"
        )

    league_counts: dict[str, int] = {}

    for row in non_empty_rows:
        team_id = str(
            row[0]
        ).strip()

        target_league_id = str(
            row[1]
        ).strip()

        source_league_id = str(
            row[2]
        ).strip()

        promotion_method = str(
            row[4]
        ).strip()

        source_status = str(
            row[11]
        ).strip()

        data_confidence = row[12]

        if promotion_method not in {
            "CHAMPION",
            "DIRECT",
            "PLAYOFF",
        }:
            raise PromotedWriteError(
                f"{team_id}: método de "
                "promoção inválido."
            )

        if (
            not source_league_id
            or source_league_id
            == target_league_id
        ):
            raise PromotedWriteError(
                f"{team_id}: liga de origem "
                "inválida."
            )

        if source_status != "MISSING":
            raise PromotedWriteError(
                f"{team_id}: source_status "
                "deveria ser MISSING."
            )

        if data_confidence != 0:
            raise PromotedWriteError(
                f"{team_id}: data_confidence "
                "deveria ser 0."
            )

        league_counts[
            target_league_id
        ] = (
            league_counts.get(
                target_league_id,
                0,
            )
            + 1
        )

    if league_counts != EXPECTED_PROMOTED_COUNTS:
        raise PromotedWriteError(
            "Contagens gravadas por liga "
            "incorretas.\n"
            f"Esperado: {EXPECTED_PROMOTED_COUNTS}\n"
            f"Encontrado: {league_counts}"
        )


def write_promoted_teams_to_dataset(
    dataset_path: str | Path = (
        DEFAULT_DATASET_PATH
    ),
) -> PromotedWriteResult:
    """
    Grava as 16 equipas promovidas no dataset oficial.
    """

    validate_team_catalog()

    promoted_teams = (
        get_promoted_teams()
    )

    validate_promoted_catalog(
        promoted_teams
    )

    final_dataset_path = Path(
        dataset_path
    ).expanduser().resolve()

    if not final_dataset_path.exists():
        raise PromotedWriteError(
            "O dataset não existe: "
            f"{final_dataset_path}"
        )

    if not final_dataset_path.is_file():
        raise PromotedWriteError(
            "O caminho não corresponde "
            "a um ficheiro: "
            f"{final_dataset_path}"
        )

    backup_path = create_backup(
        final_dataset_path
    )

    workbook = load_workbook(
        final_dataset_path
    )

    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise PromotedWriteError(
                "Falta a folha obrigatória: "
                f"{SHEET_NAME}"
            )

        worksheet = workbook[
            SHEET_NAME
        ]

        validate_headers(
            worksheet
        )

        rows_deleted = (
            clear_existing_rows(
                worksheet
            )
        )

        league_counts = (
            write_promoted_rows(
                worksheet,
                promoted_teams,
            )
        )

        validate_written_rows(
            worksheet
        )

        workbook.save(
            final_dataset_path
        )

    except Exception:
        workbook.close()

        shutil.copy2(
            backup_path,
            final_dataset_path,
        )

        raise

    finally:
        try:
            workbook.close()
        except Exception:
            pass

    return PromotedWriteResult(
        dataset_path=final_dataset_path,
        backup_path=backup_path,
        rows_deleted=rows_deleted,
        rows_written=(
            EXPECTED_TOTAL_PROMOTED
        ),
        league_counts=league_counts,
    )


def print_result(
    result: PromotedWriteResult,
) -> None:
    print()
    print("=" * 100)
    print(
        "✅ EQUIPAS PROMOVIDAS "
        "GRAVADAS NO DATASET"
    )
    print("=" * 100)

    print(
        f"Dataset:          "
        f"{result.dataset_path}"
    )

    print(
        f"Backup:           "
        f"{result.backup_path}"
    )

    print(
        f"Linhas removidas: "
        f"{result.rows_deleted}"
    )

    print(
        f"Linhas gravadas:  "
        f"{result.rows_written}"
    )

    print("-" * 100)

    for league_id in (
        "ENG1",
        "ESP1",
        "ITA1",
        "GER1",
        "FRA1",
        "POR1",
    ):
        print(
            f"{league_id}: "
            f"{result.league_counts[league_id]} "
            "promovidas"
        )

    print("=" * 100)


def main() -> int:
    try:
        result = (
            write_promoted_teams_to_dataset()
        )

        print_result(
            result
        )

        return 0

    except Exception as exc:
        print()
        print("=" * 100)
        print(
            "❌ ERRO AO GRAVAR "
            "AS EQUIPAS PROMOVIDAS"
        )
        print("=" * 100)
        print(exc)
        print("=" * 100)

        return 1


if __name__ == "__main__":
    raise SystemExit(
        main()
    )
