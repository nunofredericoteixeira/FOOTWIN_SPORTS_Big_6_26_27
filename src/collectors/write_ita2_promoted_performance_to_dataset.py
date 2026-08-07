# -*- coding: utf-8 -*-

from __future__ import annotations

import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DATASET_PATH = Path(
    "data/input/FOOTWIN_Dataset_2026_27_V001.xlsx"
)

TEAMS_SHEET_NAME = "Equipas_2026_27"
PERFORMANCE_SHEET_NAME = "Desempenho_2025_26"

SOURCE_LEAGUE_ID = "ITA2"
TARGET_LEAGUE_ID = "ITA1"
SEASON_LABEL = "2025/26"

SOURCE_URL = (
    "https://sportstats365.com/football/"
    "serie-b/2025-2026/teams/frosinone"
)

EXPECTED_TEAMS = {
    "ITA1_VENEZIA": {
        "position": 1,
        "played": 38,
        "wins": 24,
        "draws": 10,
        "losses": 4,
        "goals_for": 77,
        "goals_against": 31,
        "goal_difference": 46,
        "points": 82,
        "promoted": 1,
        "promotion_method": "CHAMPION",
    },
    "ITA1_FROSINONE": {
        "position": 2,
        "played": 38,
        "wins": 23,
        "draws": 12,
        "losses": 3,
        "goals_for": 76,
        "goals_against": 34,
        "goal_difference": 42,
        "points": 81,
        "promoted": 1,
        "promotion_method": "DIRECT",
    },
    "ITA1_MONZA": {
        "position": 3,
        "played": 38,
        "wins": 22,
        "draws": 10,
        "losses": 6,
        "goals_for": 61,
        "goals_against": 32,
        "goal_difference": 29,
        "points": 76,
        "promoted": 1,
        "promotion_method": "PLAYOFF",
    },
}


class Ita2PromotedPerformanceError(
    RuntimeError
):
    """Erro na escrita da performance das promovidas ITA2."""


def validate_performance_record(
    team_id: str,
    record: dict[str, Any],
) -> None:
    if (
        record["played"]
        != record["wins"]
        + record["draws"]
        + record["losses"]
    ):
        raise Ita2PromotedPerformanceError(
            "Total de jogos incoerente para "
            f"{team_id}."
        )

    if (
        record["goal_difference"]
        != record["goals_for"]
        - record["goals_against"]
    ):
        raise Ita2PromotedPerformanceError(
            "Diferença de golos incoerente para "
            f"{team_id}."
        )

    expected_points = (
        3 * record["wins"]
        + record["draws"]
    )

    if record["points"] != expected_points:
        raise Ita2PromotedPerformanceError(
            "Pontuação incoerente para "
            f"{team_id}: esperados "
            f"{expected_points}, encontrados "
            f"{record['points']}."
        )


def load_team_catalog(
    workbook,
) -> dict[str, dict[str, Any]]:
    worksheet = workbook[
        TEAMS_SHEET_NAME
    ]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    indexes = {
        str(header): index
        for index, header in enumerate(
            headers
        )
        if header is not None
    }

    required_headers = {
        "team_id",
        "league_id",
        "promoted",
        "promotion_method",
        "previous_division",
    }

    missing = (
        required_headers
        - set(indexes)
    )

    if missing:
        raise Ita2PromotedPerformanceError(
            "Faltam colunas na folha de equipas: "
            f"{sorted(missing)}"
        )

    teams: dict[
        str,
        dict[str, Any]
    ] = {}

    for row in worksheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        team_id = str(
            row[indexes["team_id"]]
            or ""
        ).strip()

        if team_id not in EXPECTED_TEAMS:
            continue

        teams[team_id] = {
            "league_id": str(
                row[indexes["league_id"]]
                or ""
            ).strip().upper(),
            "promoted": int(
                row[indexes["promoted"]]
                or 0
            ),
            "promotion_method": (
                str(
                    row[
                        indexes[
                            "promotion_method"
                        ]
                    ]
                    or ""
                )
                .strip()
                .upper()
            ),
            "previous_division": (
                str(
                    row[
                        indexes[
                            "previous_division"
                        ]
                    ]
                    or ""
                )
                .strip()
                .upper()
            ),
        }

    return teams


def remove_existing_rows(
    worksheet,
    headers: list[Any],
) -> int:
    indexes = {
        str(header): index
        for index, header in enumerate(
            headers
        )
        if header is not None
    }

    rows_to_delete: list[int] = []

    for row_number in range(
        worksheet.max_row,
        1,
        -1,
    ):
        team_id = worksheet.cell(
            row=row_number,
            column=(
                indexes["team_id"]
                + 1
            ),
        ).value

        if str(team_id or "").strip() in (
            EXPECTED_TEAMS
        ):
            rows_to_delete.append(
                row_number
            )

    for row_number in rows_to_delete:
        worksheet.delete_rows(
            row_number,
            1,
        )

    return len(rows_to_delete)


def write_records() -> None:
    if not DATASET_PATH.exists():
        raise Ita2PromotedPerformanceError(
            f"Dataset inexistente: {DATASET_PATH}"
        )

    backup_path = DATASET_PATH.with_name(
        f"{DATASET_PATH.stem}"
        "_BACKUP_BEFORE_ITA2_PROMOTED"
        f"{DATASET_PATH.suffix}"
    )

    shutil.copy2(
        DATASET_PATH,
        backup_path,
    )

    workbook = load_workbook(
        DATASET_PATH
    )

    try:
        teams = load_team_catalog(
            workbook
        )

        missing_teams = (
            set(EXPECTED_TEAMS)
            - set(teams)
        )

        if missing_teams:
            raise Ita2PromotedPerformanceError(
                "Faltam promovidas no catálogo: "
                f"{sorted(missing_teams)}"
            )

        for team_id, expected in (
            EXPECTED_TEAMS.items()
        ):
            catalog = teams[team_id]

            if (
                catalog["league_id"]
                != TARGET_LEAGUE_ID
            ):
                raise Ita2PromotedPerformanceError(
                    f"{team_id} não pertence à "
                    f"{TARGET_LEAGUE_ID}."
                )

            if catalog["promoted"] != 1:
                raise Ita2PromotedPerformanceError(
                    f"{team_id} não está marcada "
                    "como promovida."
                )

            if (
                catalog["promotion_method"]
                != expected[
                    "promotion_method"
                ]
            ):
                raise Ita2PromotedPerformanceError(
                    "Método de promoção diferente "
                    f"para {team_id}: catálogo="
                    f"{catalog['promotion_method']}, "
                    "esperado="
                    f"{expected['promotion_method']}."
                )

            if (
                catalog["previous_division"]
                != SOURCE_LEAGUE_ID
            ):
                raise Ita2PromotedPerformanceError(
                    f"{team_id} não tem "
                    f"{SOURCE_LEAGUE_ID} como "
                    "divisão anterior."
                )

            validate_performance_record(
                team_id,
                expected,
            )

        worksheet = workbook[
            PERFORMANCE_SHEET_NAME
        ]

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        required_headers = {
            "team_id",
            "source_league_id",
            "target_league_id",
            "season_label",
            "position",
            "played",
            "wins",
            "draws",
            "losses",
            "goals_for",
            "goals_against",
            "goal_difference",
            "points",
            "points_adjustment",
            "promoted",
            "promotion_method",
            "source_status",
            "data_confidence",
            "source_url",
            "accessed_at",
        }

        existing_headers = {
            str(header)
            for header in headers
            if header is not None
        }

        missing_headers = (
            required_headers
            - existing_headers
        )

        if missing_headers:
            raise Ita2PromotedPerformanceError(
                "Faltam colunas na folha de "
                "performance: "
                f"{sorted(missing_headers)}"
            )

        removed = remove_existing_rows(
            worksheet,
            headers,
        )

        accessed_at = datetime.now(
            timezone.utc
        ).replace(
            microsecond=0
        ).isoformat()

        written = 0

        for team_id, performance in (
            EXPECTED_TEAMS.items()
        ):
            row_data = {
                "team_id": team_id,
                "source_league_id": (
                    SOURCE_LEAGUE_ID
                ),
                "target_league_id": (
                    TARGET_LEAGUE_ID
                ),
                "season_label": (
                    SEASON_LABEL
                ),
                "position": (
                    performance["position"]
                ),
                "played": (
                    performance["played"]
                ),
                "wins": (
                    performance["wins"]
                ),
                "draws": (
                    performance["draws"]
                ),
                "losses": (
                    performance["losses"]
                ),
                "goals_for": (
                    performance["goals_for"]
                ),
                "goals_against": (
                    performance[
                        "goals_against"
                    ]
                ),
                "goal_difference": (
                    performance[
                        "goal_difference"
                    ]
                ),
                "points": (
                    performance["points"]
                ),
                "points_adjustment": 0,
                "promoted": 1,
                "promotion_method": (
                    performance[
                        "promotion_method"
                    ]
                ),
                "source_status": (
                    "MANUAL_VALIDATED"
                ),
                "data_confidence": 1.0,
                "source_url": SOURCE_URL,
                "accessed_at": accessed_at,
            }

            worksheet.append(
                [
                    row_data.get(
                        str(header)
                    )
                    for header in headers
                ]
            )

            written += 1

        workbook.save(
            DATASET_PATH
        )

        print()
        print("=" * 100)
        print(
            "ITA2 — PROMOVIDAS GRAVADAS NO DATASET"
        )
        print("=" * 100)
        print(
            f"Dataset: {DATASET_PATH.resolve()}"
        )
        print(
            f"Backup:  {backup_path.resolve()}"
        )
        print(
            f"Linhas anteriores removidas: {removed}"
        )
        print(
            f"Equipas gravadas: {written}"
        )
        print(
            "Venezia:    1.º | 82 pontos | CHAMPION"
        )
        print(
            "Frosinone:  2.º | 81 pontos | DIRECT"
        )
        print(
            "Monza:      3.º | 76 pontos | PLAYOFF"
        )
        print("=" * 100)

    except Exception:
        workbook.close()

        shutil.copy2(
            backup_path,
            DATASET_PATH,
        )

        raise

    else:
        workbook.close()


def main() -> int:
    print()
    print("=" * 100)
    print(
        "FOOTWIN SPORTS — PERFORMANCE "
        "PROMOVIDAS ITA2 2025/26"
    )
    print("=" * 100)

    write_records()

    print()
    print(
        "✅ Promovidas da ITA2 concluídas."
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(
        main()
    )
