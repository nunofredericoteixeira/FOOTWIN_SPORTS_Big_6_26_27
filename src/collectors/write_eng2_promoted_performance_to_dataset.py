# -*- coding: utf-8 -*-

from __future__ import annotations

import json
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DATASET_PATH = Path(
    "data/input/FOOTWIN_Dataset_2026_27_V001.xlsx"
)

JSON_PATH = Path(
    "data/raw/performance_pages/"
    "ENG2_2025_26_standings.json"
)

TEAMS_SHEET_NAME = "Equipas_2026_27"
PERFORMANCE_SHEET_NAME = "Desempenho_2025_26"

SOURCE_LEAGUE_ID = "ENG2"
TARGET_LEAGUE_ID = "ENG1"
SEASON_LABEL = "2025/26"

SOURCE_URL = (
    "https://footballapi.pulselive.com/"
    "football/standings"
    "?comp=12"
    "&compSeasons=778"
    "&altIds=true"
    "&page=0"
    "&pageSize=100"
)

EXPECTED_TEAMS = {
    "Coventry City": {
        "team_id": "ENG1_COVENTRY",
        "promotion_method": "CHAMPION",
    },
    "Ipswich Town": {
        "team_id": "ENG1_IPSWICH",
        "promotion_method": "DIRECT",
    },
    "Hull City": {
        "team_id": "ENG1_HULL_CITY",
        "promotion_method": "PLAYOFF",
    },
}


class Eng2PromotedPerformanceError(
    RuntimeError
):
    """Erro na escrita da performance das promovidas ENG2."""


def validate_record(
    team_id: str,
    record: dict[str, Any],
) -> None:
    if (
        record["played"]
        != record["wins"]
        + record["draws"]
        + record["losses"]
    ):
        raise Eng2PromotedPerformanceError(
            f"Total de jogos incoerente para {team_id}."
        )

    if (
        record["goal_difference"]
        != record["goals_for"]
        - record["goals_against"]
    ):
        raise Eng2PromotedPerformanceError(
            f"Diferença de golos incoerente para {team_id}."
        )

    expected_points = (
        3 * record["wins"]
        + record["draws"]
    )

    if record["points"] != expected_points:
        raise Eng2PromotedPerformanceError(
            f"Pontuação incoerente para {team_id}: "
            f"esperados {expected_points}, "
            f"encontrados {record['points']}."
        )


def extract_promoted_performance() -> dict[str, dict[str, Any]]:
    if not JSON_PATH.exists():
        raise Eng2PromotedPerformanceError(
            f"JSON inexistente: {JSON_PATH}"
        )

    try:
        data = json.loads(
            JSON_PATH.read_text(
                encoding="utf-8",
            )
        )
    except json.JSONDecodeError as exc:
        raise Eng2PromotedPerformanceError(
            "O ficheiro ENG2 não contém JSON válido."
        ) from exc

    tables = data.get(
        "tables",
        []
    )

    if not isinstance(
        tables,
        list,
    ) or not tables:
        raise Eng2PromotedPerformanceError(
            "O JSON ENG2 não contém tabelas."
        )

    target_table = None

    for table in tables:
        entries = table.get(
            "entries",
            []
        )

        if (
            table.get("gameWeek") == 46
            and isinstance(entries, list)
            and len(entries) == 24
        ):
            target_table = table
            break

    if target_table is None:
        raise Eng2PromotedPerformanceError(
            "Não foi encontrada a classificação "
            "final da Championship na jornada 46."
        )

    records: dict[str, dict[str, Any]] = {}

    for entry in target_table.get(
        "entries",
        []
    ):
        team_name = str(
            entry.get(
                "team",
                {},
            ).get(
                "name",
                "",
            )
        ).strip()

        if team_name not in EXPECTED_TEAMS:
            continue

        overall = entry.get(
            "overall",
            {}
        )

        expected = EXPECTED_TEAMS[
            team_name
        ]

        try:
            record = {
                "team_id": expected["team_id"],
                "team_name": team_name,
                "position": int(
                    entry["position"]
                ),
                "played": int(
                    overall["played"]
                ),
                "wins": int(
                    overall["won"]
                ),
                "draws": int(
                    overall["drawn"]
                ),
                "losses": int(
                    overall["lost"]
                ),
                "goals_for": int(
                    overall["goalsFor"]
                ),
                "goals_against": int(
                    overall["goalsAgainst"]
                ),
                "goal_difference": int(
                    overall["goalsDifference"]
                ),
                "points": int(
                    overall["points"]
                ),
                "promoted": 1,
                "promotion_method": expected[
                    "promotion_method"
                ],
            }
        except (
            KeyError,
            TypeError,
            ValueError,
        ) as exc:
            raise Eng2PromotedPerformanceError(
                "Dados incompletos ou inválidos para "
                f"{team_name}."
            ) from exc

        validate_record(
            record["team_id"],
            record,
        )

        records[
            record["team_id"]
        ] = record

    expected_team_ids = {
        configuration["team_id"]
        for configuration in EXPECTED_TEAMS.values()
    }

    missing_team_ids = (
        expected_team_ids
        - set(records)
    )

    if missing_team_ids:
        raise Eng2PromotedPerformanceError(
            "Faltam promovidas na classificação ENG2: "
            f"{sorted(missing_team_ids)}"
        )

    if records["ENG1_COVENTRY"]["position"] != 1:
        raise Eng2PromotedPerformanceError(
            "Coventry City não aparece na posição 1."
        )

    if records["ENG1_IPSWICH"]["position"] != 2:
        raise Eng2PromotedPerformanceError(
            "Ipswich Town não aparece na posição 2."
        )

    if records["ENG1_HULL_CITY"]["position"] != 6:
        raise Eng2PromotedPerformanceError(
            "Hull City não aparece na posição 6."
        )

    return records


def load_team_catalog(
    workbook,
) -> dict[str, dict[str, Any]]:
    worksheet = workbook[
        TEAMS_SHEET_NAME
    ]

    headers = [
        cell.value
        for cell in worksheet[1]
    ]

    indexes = {
        str(header): index
        for index, header in enumerate(headers)
        if header is not None
    }

    required = {
        "team_id",
        "league_id",
        "promoted",
        "promotion_method",
        "previous_division",
    }

    missing = required - set(indexes)

    if missing:
        raise Eng2PromotedPerformanceError(
            "Faltam colunas na folha de equipas: "
            f"{sorted(missing)}"
        )

    expected_team_ids = {
        configuration["team_id"]
        for configuration in EXPECTED_TEAMS.values()
    }

    teams: dict[str, dict[str, Any]] = {}

    for row in worksheet.iter_rows(
        min_row=2,
        values_only=True,
    ):
        team_id = str(
            row[indexes["team_id"]]
            or ""
        ).strip()

        if team_id not in expected_team_ids:
            continue

        teams[team_id] = {
            "league_id": str(
                row[indexes["league_id"]]
                or ""
            ).strip().upper(),
            "promoted": int(
                row[indexes["promoted"]]
                or 0
            ),
            "promotion_method": str(
                row[indexes["promotion_method"]]
                or ""
            ).strip().upper(),
            "previous_division": str(
                row[indexes["previous_division"]]
                or ""
            ).strip().upper(),
        }

    return teams


def remove_existing_rows(
    worksheet,
    headers: list[Any],
    expected_team_ids: set[str],
) -> int:
    indexes = {
        str(header): index
        for index, header in enumerate(headers)
        if header is not None
    }

    rows_to_delete: list[int] = []

    for row_number in range(
        worksheet.max_row,
        1,
        -1,
    ):
        team_id = worksheet.cell(
            row=row_number,
            column=indexes["team_id"] + 1,
        ).value

        if (
            str(team_id or "").strip()
            in expected_team_ids
        ):
            rows_to_delete.append(
                row_number
            )

    for row_number in rows_to_delete:
        worksheet.delete_rows(
            row_number,
            1,
        )

    return len(rows_to_delete)


def write_records(
    performances: dict[str, dict[str, Any]],
) -> None:
    if not DATASET_PATH.exists():
        raise Eng2PromotedPerformanceError(
            f"Dataset inexistente: {DATASET_PATH}"
        )

    backup_path = DATASET_PATH.with_name(
        f"{DATASET_PATH.stem}"
        "_BACKUP_BEFORE_ENG2_PROMOTED"
        f"{DATASET_PATH.suffix}"
    )

    shutil.copy2(
        DATASET_PATH,
        backup_path,
    )

    workbook = load_workbook(
        DATASET_PATH
    )

    try:
        teams = load_team_catalog(
            workbook
        )

        expected_team_ids = set(
            performances
        )

        missing_teams = (
            expected_team_ids
            - set(teams)
        )

        if missing_teams:
            raise Eng2PromotedPerformanceError(
                "Faltam promovidas no catálogo: "
                f"{sorted(missing_teams)}"
            )

        for team_id, performance in performances.items():
            catalog = teams[
                team_id
            ]

            if catalog["league_id"] != TARGET_LEAGUE_ID:
                raise Eng2PromotedPerformanceError(
                    f"{team_id} não pertence à "
                    f"{TARGET_LEAGUE_ID}."
                )

            if catalog["promoted"] != 1:
                raise Eng2PromotedPerformanceError(
                    f"{team_id} não está marcada como promovida."
                )

            if (
                catalog["promotion_method"]
                != performance["promotion_method"]
            ):
                raise Eng2PromotedPerformanceError(
                    "Método de promoção incorreto para "
                    f"{team_id}: "
                    f"{catalog['promotion_method']}."
                )

            if (
                catalog["previous_division"]
                != SOURCE_LEAGUE_ID
            ):
                raise Eng2PromotedPerformanceError(
                    f"{team_id} não tem {SOURCE_LEAGUE_ID} "
                    "como divisão anterior."
                )

        worksheet = workbook[
            PERFORMANCE_SHEET_NAME
        ]

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        required_headers = {
            "team_id",
            "source_league_id",
            "target_league_id",
            "season_label",
            "position",
            "played",
            "wins",
            "draws",
            "losses",
            "goals_for",
            "goals_against",
            "goal_difference",
            "points",
            "points_adjustment",
            "promoted",
            "promotion_method",
            "source_status",
            "data_confidence",
            "source_url",
            "accessed_at",
        }

        existing_headers = {
            str(header)
            for header in headers
            if header is not None
        }

        missing_headers = (
            required_headers
            - existing_headers
        )

        if missing_headers:
            raise Eng2PromotedPerformanceError(
                "Faltam colunas na folha de performance: "
                f"{sorted(missing_headers)}"
            )

        removed = remove_existing_rows(
            worksheet,
            headers,
            expected_team_ids,
        )

        accessed_at = datetime.now(
            timezone.utc
        ).replace(
            microsecond=0
        ).isoformat()

        written = 0

        for team_id, performance in performances.items():
            row_data = {
                "team_id": team_id,
                "source_league_id": SOURCE_LEAGUE_ID,
                "target_league_id": TARGET_LEAGUE_ID,
                "season_label": SEASON_LABEL,
                "position": performance["position"],
                "played": performance["played"],
                "wins": performance["wins"],
                "draws": performance["draws"],
                "losses": performance["losses"],
                "goals_for": performance["goals_for"],
                "goals_against": performance["goals_against"],
                "goal_difference": performance["goal_difference"],
                "points": performance["points"],
                "points_adjustment": 0,
                "promoted": 1,
                "promotion_method": performance[
                    "promotion_method"
                ],
                "source_status": "CONFIRMED",
                "data_confidence": 1.0,
                "source_url": SOURCE_URL,
                "accessed_at": accessed_at,
            }

            worksheet.append(
                [
                    row_data.get(
                        str(header)
                    )
                    for header in headers
                ]
            )

            written += 1

        if written != 3:
            raise Eng2PromotedPerformanceError(
                "Esperavam-se 3 promovidas ENG2 "
                f"gravadas, mas foram gravadas {written}."
            )

        workbook.save(
            DATASET_PATH
        )

        print()
        print("=" * 100)
        print(
            "ENG2 — PROMOVIDAS GRAVADAS NO DATASET"
        )
        print("=" * 100)
        print(
            f"Dataset: {DATASET_PATH.resolve()}"
        )
        print(
            f"Backup:  {backup_path.resolve()}"
        )
        print(
            f"Linhas anteriores removidas: {removed}"
        )
        print(
            f"Equipas gravadas: {written}"
        )
        print(
            "Coventry City: 1.º | 95 pontos | CHAMPION"
        )
        print(
            "Ipswich Town:  2.º | 84 pontos | DIRECT"
        )
        print(
            "Hull City:     6.º | 73 pontos | PLAYOFF"
        )
        print("=" * 100)

    except Exception:
        workbook.close()

        shutil.copy2(
            backup_path,
            DATASET_PATH,
        )

        raise

    else:
        workbook.close()


def main() -> int:
    print()
    print("=" * 100)
    print(
        "FOOTWIN SPORTS — PERFORMANCE "
        "PROMOVIDAS ENG2 2025/26"
    )
    print("=" * 100)

    performances = extract_promoted_performance()

    print(
        f"Promovidas recolhidas: {len(performances)}"
    )

    write_records(
        performances
    )

    print()
    print(
        "✅ Promovidas da ENG2 concluídas."
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(
        main()
    )
