# -*- coding: utf-8 -*-

from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


DATASET_PATH = Path("data/input/FOOTWIN_Dataset_2026_27_V001.xlsx")
ICS_PATH = Path(
    "data/raw/performance_pages/portugal/api/"
    "POR1_2026_27_CALENDAR.ics"
)
BACKUP_PATH = Path(
    "data/input/"
    "FOOTWIN_Dataset_2026_27_V001_BACKUP_BEFORE_POR1_CALENDAR.xlsx"
)

SHEET_NAME = "Calendario_2026_27"
LEAGUE_ID = "POR1"
SEASON_LABEL = "2026/27"

EXPECTED_HEADERS = [
    "match_id",
    "league_id",
    "season_label",
    "round_number",
    "match_date",
    "home_team_id",
    "away_team_id",
    "status",
    "home_goals",
    "away_goals",
    "schedule_type",
    "source_url",
]

TEAM_NAME_TO_ID = {
    "Académico": "POR1_ACADEMICO",
    "CD Nacional": "POR1_NACIONAL",
    "Santa Clara": "POR1_SANTA_CLARA",
    "Casa Pia AC": "POR1_CASA_PIA",
    "Estoril Praia": "POR1_ESTORIL",
    "Estrela Amadora": "POR1_ESTRELA_AMADORA",
    "FC Alverca": "POR1_ALVERCA",
    "FC Arouca": "POR1_AROUCA",
    "FC Famalicão": "POR1_FAMALICAO",
    "FC Porto": "POR1_PORTO",
    "Gil Vicente FC": "POR1_GIL_VICENTE",
    "Marítimo M.": "POR1_MARITIMO",
    "Moreirense FC": "POR1_MOREIRENSE",
    "Rio Ave FC": "POR1_RIO_AVE",
    "SC Braga": "POR1_BRAGA",
    "SL Benfica": "POR1_BENFICA",
    "Sporting CP": "POR1_SPORTING",
    "Vitória SC": "POR1_VITORIA_SC",
}


def unfold_ics(text: str) -> str:
    return re.sub(r"\r?\n[ \t]", "", text)


def parse_event(block: str) -> dict[str, object]:
    fields: dict[str, str] = {}

    for line in block.splitlines():
        if ":" not in line:
            continue

        key, value = line.split(":", 1)
        key = key.split(";", 1)[0]
        fields[key] = value.strip()

    summary = fields.get("SUMMARY", "")
    description = fields.get("DESCRIPTION", "")
    source_url = fields.get("URL", "")

    if " - " not in summary:
        raise ValueError(f"SUMMARY inválido: {summary!r}")

    home_name, away_name = [
        value.strip()
        for value in summary.split(" - ", 1)
    ]

    round_match = re.search(
        r"jornada\s+(\d+)",
        description,
        flags=re.IGNORECASE,
    )
    if not round_match:
        raise ValueError(
            f"Jornada não encontrada no evento: {summary!r}"
        )

    date_raw = fields.get("DTSTART", "")
    if not date_raw:
        raise ValueError(
            f"DTSTART não encontrado no evento: {summary!r}"
        )

    match_date = datetime.strptime(
        date_raw,
        "%Y%m%dT%H%M%SZ",
    )

    if home_name not in TEAM_NAME_TO_ID:
        raise ValueError(
            f"Equipa da casa sem mapeamento: {home_name!r}"
        )

    if away_name not in TEAM_NAME_TO_ID:
        raise ValueError(
            f"Equipa visitante sem mapeamento: {away_name!r}"
        )

    round_number = int(round_match.group(1))

    url_match = re.search(
        r"/match/20262027/ligaportugalbetclic/(\d+)/(\d+)",
        source_url,
    )

    if url_match:
        official_round = int(url_match.group(1))
        official_match_number = int(url_match.group(2))
    else:
        official_round = round_number
        official_match_number = 0

    match_id = (
        f"POR1_2026_27_R{official_round:02d}_"
        f"M{official_match_number:02d}_"
        f"{TEAM_NAME_TO_ID[home_name]}_"
        f"{TEAM_NAME_TO_ID[away_name]}"
    )

    return {
        "match_id": match_id,
        "league_id": LEAGUE_ID,
        "season_label": SEASON_LABEL,
        "round_number": round_number,
        "match_date": match_date,
        "home_team_id": TEAM_NAME_TO_ID[home_name],
        "away_team_id": TEAM_NAME_TO_ID[away_name],
        "status": "SCHEDULED",
        "home_goals": None,
        "away_goals": None,
        "schedule_type": "OFFICIAL",
        "source_url": source_url,
    }


def read_calendar() -> list[dict[str, object]]:
    text = ICS_PATH.read_text(
        encoding="utf-8",
        errors="replace",
    )
    text = unfold_ics(text)

    blocks = re.findall(
        r"BEGIN:VEVENT\n(.*?)\nEND:VEVENT",
        text,
        flags=re.DOTALL,
    )

    matches = [parse_event(block) for block in blocks]

    matches.sort(
        key=lambda item: (
            int(item["round_number"]),
            item["match_date"],
            str(item["match_id"]),
        )
    )

    return matches


def validate_matches(
    matches: list[dict[str, object]],
) -> None:
    if len(matches) != 306:
        raise ValueError(
            f"Esperados 306 jogos, encontrados {len(matches)}."
        )

    rounds = {
        int(match["round_number"])
        for match in matches
    }
    if rounds != set(range(1, 35)):
        raise ValueError(
            f"Jornadas inválidas: {sorted(rounds)}"
        )

    match_ids = [
        str(match["match_id"])
        for match in matches
    ]
    if len(match_ids) != len(set(match_ids)):
        raise ValueError("Existem match_id duplicados.")

    directed_pairs = [
        (
            str(match["home_team_id"]),
            str(match["away_team_id"]),
        )
        for match in matches
    ]
    if len(directed_pairs) != len(set(directed_pairs)):
        raise ValueError(
            "Existem confrontos casa/fora duplicados."
        )

    teams = {
        str(match["home_team_id"])
        for match in matches
    } | {
        str(match["away_team_id"])
        for match in matches
    }

    if teams != set(TEAM_NAME_TO_ID.values()):
        missing = set(TEAM_NAME_TO_ID.values()) - teams
        unexpected = teams - set(TEAM_NAME_TO_ID.values())
        raise ValueError(
            "Conjunto de equipas inválido. "
            f"Em falta={sorted(missing)}; "
            f"inesperadas={sorted(unexpected)}"
        )

    for round_number in range(1, 35):
        round_matches = [
            match
            for match in matches
            if int(match["round_number"]) == round_number
        ]

        if len(round_matches) != 9:
            raise ValueError(
                f"Jornada {round_number}: "
                f"esperados 9 jogos, encontrados "
                f"{len(round_matches)}."
            )

        round_teams: list[str] = []
        for match in round_matches:
            round_teams.extend(
                [
                    str(match["home_team_id"]),
                    str(match["away_team_id"]),
                ]
            )

        if len(round_teams) != len(set(round_teams)):
            raise ValueError(
                f"Jornada {round_number}: "
                "há equipas repetidas."
            )


def write_dataset(
    matches: list[dict[str, object]],
) -> None:
    if not DATASET_PATH.exists():
        raise FileNotFoundError(
            f"Dataset não encontrado: {DATASET_PATH}"
        )

    if not ICS_PATH.exists():
        raise FileNotFoundError(
            f"Calendário ICS não encontrado: {ICS_PATH}"
        )

    shutil.copy2(DATASET_PATH, BACKUP_PATH)

    workbook = load_workbook(DATASET_PATH)

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(
            f"Folha não encontrada: {SHEET_NAME}"
        )

    worksheet = workbook[SHEET_NAME]

    headers = [
        worksheet.cell(row=1, column=column).value
        for column in range(1, len(EXPECTED_HEADERS) + 1)
    ]

    if headers != EXPECTED_HEADERS:
        raise ValueError(
            "Cabeçalhos inesperados.\n"
            f"Esperados: {EXPECTED_HEADERS}\n"
            f"Encontrados: {headers}"
        )

    rows_to_delete: list[int] = []

    for row in range(2, worksheet.max_row + 1):
        league_id = worksheet.cell(
            row=row,
            column=2,
        ).value
        season_label = worksheet.cell(
            row=row,
            column=3,
        ).value

        if (
            str(league_id or "").strip() == LEAGUE_ID
            and str(season_label or "").strip()
            == SEASON_LABEL
        ):
            rows_to_delete.append(row)

    for row in reversed(rows_to_delete):
        worksheet.delete_rows(row)

    first_output_row = worksheet.max_row + 1

    for match in matches:
        worksheet.append(
            [
                match[header]
                for header in EXPECTED_HEADERS
            ]
        )

    for row in range(
        first_output_row,
        first_output_row + len(matches),
    ):
        worksheet.cell(
            row=row,
            column=5,
        ).number_format = "yyyy-mm-dd hh:mm:ss"

    workbook.save(DATASET_PATH)


def main() -> None:
    matches = read_calendar()
    validate_matches(matches)
    write_dataset(matches)

    print("=" * 90)
    print("POR1 — CALENDÁRIO OFICIAL 2026/27")
    print("=" * 90)
    print(f"Fonte ICS: {ICS_PATH}")
    print(f"Dataset: {DATASET_PATH}")
    print(f"Backup: {BACKUP_PATH}")
    print(f"Jogos gravados: {len(matches)}")
    print(
        "Jornadas:",
        min(int(match["round_number"]) for match in matches),
        "a",
        max(int(match["round_number"]) for match in matches),
    )
    print(
        "Equipas:",
        len(
            {
                str(match["home_team_id"])
                for match in matches
            }
            | {
                str(match["away_team_id"])
                for match in matches
            }
        ),
    )
    print("Estado: SCHEDULED")
    print("Tipo: OFFICIAL")
    print("RESULTADO: SUCCESS")


if __name__ == "__main__":
    main()
