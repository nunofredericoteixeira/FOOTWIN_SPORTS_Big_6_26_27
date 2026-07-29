# -*- coding: utf-8 -*-

from __future__ import annotations

import math
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo,
)

from src.models.simulation_query_service import (
    SimulationSummary,
    get_latest_simulation,
    get_position_probability,
    get_simulation_by_id,
)
from src.utils.logger import get_logger


logger = get_logger(
    "exports.simulation_excel_exporter"
)


DEFAULT_OUTPUT_DIRECTORY = Path(
    "outputs/simulations"
)

TITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor="163A5F",
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="245B85",
)

SUBHEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

CHAMPION_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAD3",
)

RELEGATION_FILL = PatternFill(
    fill_type="solid",
    fgColor="F4CCCC",
)

WHITE_FONT = Font(
    color="FFFFFF",
    bold=True,
)

TITLE_FONT = Font(
    color="FFFFFF",
    bold=True,
    size=16,
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

BOLD_FONT = Font(
    bold=True,
)

THIN_BORDER = Border(
    left=Side(
        style="thin",
        color="D9D9D9",
    ),
    right=Side(
        style="thin",
        color="D9D9D9",
    ),
    top=Side(
        style="thin",
        color="D9D9D9",
    ),
    bottom=Side(
        style="thin",
        color="D9D9D9",
    ),
)


@dataclass(frozen=True)
class SimulationExcelExportResult:
    output_path: Path
    simulation_id: str
    team_count: int
    sheet_names: tuple[str, ...]


class SimulationExcelExportError(RuntimeError):
    """Erro ao exportar a simulação para Excel."""


def export_latest_simulation_to_excel(
    league_id: str,
    season_label: str = "2026/27",
    model_version: str | None = None,
    output_directory: str | Path | None = None,
    filename: str | None = None,
) -> SimulationExcelExportResult:
    """
    Exporta a simulação mais recente de uma liga.
    """

    simulation = get_latest_simulation(
        league_id=league_id,
        season_label=season_label,
        model_version=model_version,
    )

    return export_simulation_to_excel(
        simulation=simulation,
        output_directory=output_directory,
        filename=filename,
    )


def export_simulation_by_id_to_excel(
    simulation_id: str,
    output_directory: str | Path | None = None,
    filename: str | None = None,
) -> SimulationExcelExportResult:
    """
    Exporta uma simulação específica.
    """

    simulation = get_simulation_by_id(
        simulation_id
    )

    return export_simulation_to_excel(
        simulation=simulation,
        output_directory=output_directory,
        filename=filename,
    )


def export_simulation_to_excel(
    simulation: SimulationSummary,
    output_directory: str | Path | None = None,
    filename: str | None = None,
) -> SimulationExcelExportResult:
    """
    Cria o ficheiro Excel da simulação.
    """

    validate_simulation_for_export(
        simulation
    )

    final_output_directory = Path(
        output_directory
        if output_directory is not None
        else DEFAULT_OUTPUT_DIRECTORY
    )

    final_output_directory.mkdir(
        parents=True,
        exist_ok=True,
    )

    final_filename = (
        clean_filename(filename)
        if filename
        else build_default_filename(
            simulation
        )
    )

    output_path = (
        final_output_directory
        / final_filename
    )

    workbook = Workbook()

    default_sheet = workbook.active

    if default_sheet is not None:
        workbook.remove(
            default_sheet
        )

    summary_sheet = workbook.create_sheet(
        "Resumo"
    )

    positions_sheet = workbook.create_sheet(
        "Probabilidades_Posicao"
    )

    metadata_sheet = workbook.create_sheet(
        "Metadados"
    )

    build_summary_sheet(
        sheet=summary_sheet,
        simulation=simulation,
    )

    build_position_probabilities_sheet(
        sheet=positions_sheet,
        simulation=simulation,
    )

    build_metadata_sheet(
        sheet=metadata_sheet,
        simulation=simulation,
    )

    workbook.save(
        output_path
    )

    if not output_path.exists():
        raise SimulationExcelExportError(
            "O ficheiro Excel não foi criado."
        )

    if output_path.stat().st_size <= 0:
        raise SimulationExcelExportError(
            "O ficheiro Excel criado está vazio."
        )

    logger.info(
        "Simulação exportada para Excel | "
        "simulation_id=%s | ficheiro=%s",
        simulation.simulation_id,
        output_path.resolve(),
    )

    return SimulationExcelExportResult(
        output_path=output_path.resolve(),
        simulation_id=(
            simulation.simulation_id
        ),
        team_count=len(
            simulation.teams
        ),
        sheet_names=(
            "Resumo",
            "Probabilidades_Posicao",
            "Metadados",
        ),
    )


def build_summary_sheet(
    sheet: Any,
    simulation: SimulationSummary,
) -> None:
    """
    Cria a folha principal da simulação.
    """

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A8"

    sheet.merge_cells(
        "A1:Q1"
    )

    title_cell = sheet["A1"]

    title_cell.value = (
        f"FOOTWIN SPORTS — "
        f"SIMULAÇÃO {simulation.league_name}"
    )

    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    sheet.row_dimensions[1].height = 28

    metadata = (
        (
            "Simulation ID",
            simulation.simulation_id,
        ),
        (
            "Liga",
            simulation.league_name,
        ),
        (
            "Época",
            simulation.season_label,
        ),
        (
            "Modelo",
            simulation.model_version,
        ),
        (
            "Simulações",
            simulation.simulation_count,
        ),
        (
            "Seed",
            simulation.random_seed,
        ),
        (
            "Estado",
            simulation.status,
        ),
        (
            "Finalizada em",
            simulation.finished_at or "",
        ),
    )

    metadata_positions = (
        ("A3", "B3"),
        ("D3", "E3"),
        ("G3", "H3"),
        ("J3", "K3"),
        ("A4", "B4"),
        ("D4", "E4"),
        ("G4", "H4"),
        ("J4", "K4"),
    )

    for (
        label,
        value,
    ), (
        label_cell,
        value_cell,
    ) in zip(
        metadata,
        metadata_positions,
    ):
        sheet[label_cell] = label
        sheet[value_cell] = value

        sheet[label_cell].font = BOLD_FONT
        sheet[label_cell].fill = SUBHEADER_FILL

        sheet[label_cell].border = THIN_BORDER
        sheet[value_cell].border = THIN_BORDER

    headers = [
        "Posição",
        "Equipa",
        "Posição média",
        "Posição mediana",
        "Pontos médios",
        "Golos marcados",
        "Golos sofridos",
        "Diferença de golos",
        "Título",
        "Europa",
        "Descida",
        "Playoff",
        "P10",
        "P25",
        "P50",
        "P75",
        "P90",
    ]

    header_row = 7

    for column_index, header in enumerate(
        headers,
        start=1,
    ):
        cell = sheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = THIN_BORDER

    for ranking_position, team in enumerate(
        simulation.teams,
        start=1,
    ):
        row_number = (
            header_row
            + ranking_position
        )

        values = [
            ranking_position,
            team.team_name,
            team.average_position,
            team.median_position,
            team.average_points,
            team.average_goals_for,
            team.average_goals_against,
            team.average_goal_difference,
            team.title_probability,
            team.europe_probability,
            team.relegation_probability,
            team.playoff_probability,
            team.points_p10,
            team.points_p25,
            team.points_p50,
            team.points_p75,
            team.points_p90,
        ]

        for column_index, value in enumerate(
            values,
            start=1,
        ):
            cell = sheet.cell(
                row=row_number,
                column=column_index,
                value=value,
            )

            cell.border = THIN_BORDER

            if column_index == 2:
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                )
            else:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

        for probability_column in range(
            9,
            13,
        ):
            sheet.cell(
                row=row_number,
                column=probability_column,
            ).number_format = "0.00%"

        for decimal_column in range(
            3,
            18,
        ):
            if decimal_column not in range(
                9,
                13,
            ):
                sheet.cell(
                    row=row_number,
                    column=decimal_column,
                ).number_format = "0.00"

        if ranking_position == 1:
            for column_index in range(
                1,
                len(headers) + 1,
            ):
                sheet.cell(
                    row=row_number,
                    column=column_index,
                ).fill = CHAMPION_FILL

        if ranking_position == len(
            simulation.teams
        ):
            for column_index in range(
                1,
                len(headers) + 1,
            ):
                sheet.cell(
                    row=row_number,
                    column=column_index,
                ).fill = RELEGATION_FILL

    final_row = (
        header_row
        + len(simulation.teams)
    )

    table_reference = (
        f"A{header_row}:"
        f"Q{final_row}"
    )

    table = Table(
        displayName=(
            "TabelaResumoSimulacao"
        ),
        ref=table_reference,
    )

    table_style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    table.tableStyleInfo = (
        table_style
    )

    sheet.add_table(
        table
    )

    widths = {
        "A": 11,
        "B": 28,
        "C": 15,
        "D": 17,
        "E": 15,
        "F": 16,
        "G": 15,
        "H": 19,
        "I": 12,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 10,
        "N": 10,
        "O": 10,
        "P": 10,
        "Q": 10,
    }

    for column_letter, width in widths.items():
        sheet.column_dimensions[
            column_letter
        ].width = width

    sheet.auto_filter.ref = (
        table_reference
    )


def build_position_probabilities_sheet(
    sheet: Any,
    simulation: SimulationSummary,
) -> None:
    """
    Cria a matriz de probabilidades por posição.
    """

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "B4"

    team_count = len(
        simulation.teams
    )

    final_column = get_column_letter(
        team_count + 2
    )

    sheet.merge_cells(
        f"A1:{final_column}1"
    )

    title_cell = sheet["A1"]

    title_cell.value = (
        "PROBABILIDADES DE POSIÇÃO"
    )

    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    headers = [
        "Equipa",
        "Posição média",
    ] + [
        f"{position}.º"
        for position in range(
            1,
            team_count + 1,
        )
    ]

    header_row = 3

    for column_index, header in enumerate(
        headers,
        start=1,
    ):
        cell = sheet.cell(
            row=header_row,
            column=column_index,
            value=header,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for row_offset, team in enumerate(
        simulation.teams,
        start=1,
    ):
        row_number = (
            header_row
            + row_offset
        )

        sheet.cell(
            row=row_number,
            column=1,
            value=team.team_name,
        )

        sheet.cell(
            row=row_number,
            column=2,
            value=team.average_position,
        )

        sheet.cell(
            row=row_number,
            column=2,
        ).number_format = "0.00"

        for position in range(
            1,
            team_count + 1,
        ):
            probability = (
                get_position_probability(
                    team=team,
                    position=position,
                )
            )

            probability_cell = sheet.cell(
                row=row_number,
                column=position + 2,
                value=probability,
            )

            probability_cell.number_format = (
                "0.00%"
            )

        for column_index in range(
            1,
            team_count + 3,
        ):
            cell = sheet.cell(
                row=row_number,
                column=column_index,
            )

            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal=(
                    "left"
                    if column_index == 1
                    else "center"
                ),
                vertical="center",
            )

    total_row = (
        header_row
        + team_count
        + 1
    )

    sheet.cell(
        row=total_row,
        column=1,
        value="TOTAL",
    )

    sheet.cell(
        row=total_row,
        column=1,
    ).font = BOLD_FONT

    sheet.cell(
        row=total_row,
        column=1,
    ).fill = SUBHEADER_FILL

    sheet.cell(
        row=total_row,
        column=2,
        value="",
    )

    for position in range(
        1,
        team_count + 1,
    ):
        column_index = (
            position + 2
        )

        column_letter = get_column_letter(
            column_index
        )

        first_data_row = (
            header_row + 1
        )

        last_data_row = (
            header_row + team_count
        )

        total_cell = sheet.cell(
            row=total_row,
            column=column_index,
            value=(
                f"=SUM("
                f"{column_letter}"
                f"{first_data_row}:"
                f"{column_letter}"
                f"{last_data_row})"
            ),
        )

        total_cell.number_format = "0.00%"
        total_cell.font = BOLD_FONT
        total_cell.fill = SUBHEADER_FILL

    for column_index in range(
        1,
        team_count + 3,
    ):
        sheet.cell(
            row=total_row,
            column=column_index,
        ).border = THIN_BORDER

        sheet.cell(
            row=total_row,
            column=column_index,
        ).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    table_reference = (
        f"A{header_row}:"
        f"{final_column}"
        f"{header_row + team_count}"
    )

    table = Table(
        displayName=(
            "TabelaProbabilidadesPosicao"
        ),
        ref=table_reference,
    )

    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    sheet.add_table(
        table
    )

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 16

    for column_index in range(
        3,
        team_count + 3,
    ):
        sheet.column_dimensions[
            get_column_letter(
                column_index
            )
        ].width = 12


def build_metadata_sheet(
    sheet: Any,
    simulation: SimulationSummary,
) -> None:
    """
    Cria a folha de metadados.
    """

    sheet.sheet_view.showGridLines = False

    sheet.merge_cells(
        "A1:D1"
    )

    title_cell = sheet["A1"]

    title_cell.value = (
        "METADADOS DA SIMULAÇÃO"
    )

    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    metadata = [
        (
            "Simulation ID",
            simulation.simulation_id,
        ),
        (
            "Liga ID",
            simulation.league_id,
        ),
        (
            "Liga",
            simulation.league_name,
        ),
        (
            "Época",
            simulation.season_label,
        ),
        (
            "Modelo",
            simulation.model_version,
        ),
        (
            "Run ID",
            simulation.run_id or "",
        ),
        (
            "Número de simulações",
            simulation.simulation_count,
        ),
        (
            "Random seed",
            simulation.random_seed,
        ),
        (
            "Estado",
            simulation.status,
        ),
        (
            "Início",
            simulation.started_at,
        ),
        (
            "Fim",
            simulation.finished_at or "",
        ),
        (
            "Número de equipas",
            len(simulation.teams),
        ),
        (
            "Exportado em",
            datetime.now().isoformat(
                timespec="seconds"
            ),
        ),
    ]

    header_row = 3

    sheet.cell(
        row=header_row,
        column=1,
        value="Campo",
    )

    sheet.cell(
        row=header_row,
        column=2,
        value="Valor",
    )

    for column_index in (
        1,
        2,
    ):
        cell = sheet.cell(
            row=header_row,
            column=column_index,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for row_offset, (
        label,
        value,
    ) in enumerate(
        metadata,
        start=1,
    ):
        row_number = (
            header_row
            + row_offset
        )

        sheet.cell(
            row=row_number,
            column=1,
            value=label,
        )

        sheet.cell(
            row=row_number,
            column=2,
            value=value,
        )

        sheet.cell(
            row=row_number,
            column=1,
        ).font = BOLD_FONT

        sheet.cell(
            row=row_number,
            column=1,
        ).fill = SUBHEADER_FILL

        for column_index in (
            1,
            2,
        ):
            sheet.cell(
                row=row_number,
                column=column_index,
            ).border = THIN_BORDER

            sheet.cell(
                row=row_number,
                column=column_index,
            ).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 65


def validate_simulation_for_export(
    simulation: SimulationSummary,
) -> None:
    """
    Confirma que a simulação pode ser exportada.
    """

    if not simulation.simulation_id:
        raise SimulationExcelExportError(
            "A simulação não possui simulation_id."
        )

    if not simulation.teams:
        raise SimulationExcelExportError(
            "A simulação não possui equipas."
        )

    if simulation.simulation_count <= 0:
        raise SimulationExcelExportError(
            "simulation_count deve ser "
            "superior a zero."
        )

    team_ids: set[str] = set()

    for team in simulation.teams:
        if team.team_id in team_ids:
            raise SimulationExcelExportError(
                f"Equipa duplicada: {team.team_id}"
            )

        team_ids.add(
            team.team_id
        )

        probability_total = sum(
            item.probability
            for item
            in team.position_probabilities
        )

        if not math.isclose(
            probability_total,
            1.0,
            abs_tol=0.000001,
        ):
            raise SimulationExcelExportError(
                "As probabilidades de posição "
                f"da equipa {team.team_id} "
                "não totalizam 1."
            )


def build_default_filename(
    simulation: SimulationSummary,
) -> str:
    """
    Cria o nome padrão do ficheiro.
    """

    league_part = sanitize_filename_part(
        simulation.league_id
    )

    season_part = sanitize_filename_part(
        simulation.season_label
    )

    model_part = sanitize_filename_part(
        simulation.model_version
    )

    return (
        f"Simulacao_{league_part}_"
        f"{season_part}_{model_part}.xlsx"
    )


def clean_filename(
    filename: str,
) -> str:
    """
    Valida e normaliza o nome do ficheiro.
    """

    cleaned = str(
        filename
    ).strip()

    if not cleaned:
        raise SimulationExcelExportError(
            "O nome do ficheiro está vazio."
        )

    if not cleaned.lower().endswith(
        ".xlsx"
    ):
        cleaned += ".xlsx"

    path = Path(
        cleaned
    )

    if path.name != cleaned:
        raise SimulationExcelExportError(
            "O nome do ficheiro não pode "
            "conter diretórios."
        )

    return cleaned


def sanitize_filename_part(
    value: Any,
) -> str:
    """
    Torna uma parte segura para o nome do ficheiro.
    """

    text = str(
        value
    ).strip()

    replacements = {
        "/": "-",
        "\\": "-",
        ":": "-",
        "*": "",
        "?": "",
        '"': "",
        "<": "",
        ">": "",
        "|": "",
        " ": "_",
    }

    for old_value, new_value in replacements.items():
        text = text.replace(
            old_value,
            new_value,
        )

    while "__" in text:
        text = text.replace(
            "__",
            "_",
        )

    final_text = text.strip(
        "._-"
    )

    if not final_text:
        raise SimulationExcelExportError(
            "Não foi possível criar uma parte "
            "válida para o nome do ficheiro."
        )

    return final_text
