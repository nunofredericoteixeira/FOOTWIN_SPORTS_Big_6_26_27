# -*- coding: utf-8 -*-

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.config.league_config import get_active_leagues
from src.config.path_config import load_paths_config
from src.utils.logger import get_logger


logger = get_logger("templates.dataset")


DATASET_FILENAME = "FOOTWIN_Dataset_2026_27_V001.xlsx"


SHEET_HEADERS: dict[str, list[str]] = {
    "Resumo": [
        "Indicador",
        "Valor",
        "Estado",
        "Observações",
    ],
    "Ligas": [
        "league_id",
        "league_name",
        "country",
        "country_code",
        "season_label",
        "team_count",
        "matches_per_team",
        "total_matches",
        "league_strength_factor",
        "relegation_places",
        "playoff_places",
        "active",
    ],
    "Equipas_2026_27": [
        "team_id",
        "team_name",
        "short_name",
        "normalized_name",
        "league_id",
        "country",
        "season_label",
        "promoted",
        "promotion_method",
        "previous_division",
        "active",
    ],
    "Desempenho_2025_26": [
        "team_id",
        "source_league_id",
        "target_league_id",
        "season_label",
        "position",
        "played",
        "wins",
        "draws",
        "losses",
        "goals_for",
        "goals_against",
        "goal_difference",
        "points",
        "points_adjustment",
        "promoted",
        "promotion_method",
        "source_status",
        "data_confidence",
        "source_url",
        "accessed_at",
    ],
    "Promovidas": [
        "team_id",
        "target_league_id",
        "source_league_id",
        "source_position",
        "promotion_method",
        "played",
        "points",
        "goals_for",
        "goals_against",
        "goal_difference",
        "promotion_factor",
        "source_status",
        "data_confidence",
        "source_url",
    ],
    "Calendario_2026_27": [
        "match_id",
        "league_id",
        "season_label",
        "round_number",
        "match_date",
        "home_team_id",
        "away_team_id",
        "status",
        "home_goals",
        "away_goals",
        "schedule_type",
        "source_url",
    ],
    "Fontes": [
        "source_record_id",
        "entity_type",
        "entity_id",
        "data_type",
        "source_name",
        "source_url",
        "season_label",
        "accessed_at",
        "source_status",
        "notes",
    ],
    "Mapeamento_IDs": [
        "source_name",
        "source_entity_id",
        "source_entity_name",
        "internal_entity_type",
        "internal_entity_id",
        "match_status",
        "confidence",
        "notes",
    ],
    "Validacao": [
        "issue_id",
        "severity",
        "entity_type",
        "entity_id",
        "field_name",
        "expected_value",
        "actual_value",
        "message",
        "resolved",
        "resolution_note",
    ],
}


HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

SUBHEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)

SUCCESS_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE",
)

WARNING_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFEB9C",
)

ERROR_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE",
)


def create_dataset_template(
    output_path: str | Path | None = None,
    overwrite: bool = False,
) -> Path:
    """
    Cria o template Excel oficial do dataset FOOTWIN SPORTS.
    """

    paths = load_paths_config()

    if output_path is None:
        output_file = paths["data"]["input"] / DATASET_FILENAME
    else:
        output_file = Path(output_path).expanduser().resolve()

    output_file.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    if output_file.exists() and not overwrite:
        raise FileExistsError(
            f"O ficheiro já existe: {output_file}"
        )

    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, headers in SHEET_HEADERS.items():
        worksheet = workbook.create_sheet(
            title=sheet_name,
        )

        _write_headers(
            worksheet=worksheet,
            headers=headers,
        )

        _apply_base_layout(
            worksheet=worksheet,
            headers=headers,
        )

    _populate_summary_sheet(workbook["Resumo"])
    _populate_leagues_sheet(workbook["Ligas"])
    _add_data_validations(workbook)
    _apply_number_formats(workbook)

    workbook.save(output_file)

    logger.info(
        "Template Excel criado | ficheiro=%s",
        output_file,
    )

    return output_file


def _write_headers(
    worksheet,
    headers: Iterable[str],
) -> None:
    for column_index, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )


def _apply_base_layout(
    worksheet,
    headers: list[str],
) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(headers))}1"
    )

    worksheet.row_dimensions[1].height = 24

    for column_index, header in enumerate(
        headers,
        start=1,
    ):
        width = max(
            14,
            min(
                35,
                len(header) + 4,
            ),
        )

        worksheet.column_dimensions[
            get_column_letter(column_index)
        ].width = width


def _populate_summary_sheet(worksheet) -> None:
    rows = [
        ("Dataset", "DATASET_2026_27_V001", "PENDING", ""),
        ("Época", "2026/27", "CONFIGURED", ""),
        ("Ligas esperadas", 6, "CONFIGURED", ""),
        ("Equipas esperadas", 114, "CONFIGURED", ""),
        ("Jogos esperados", 2058, "CONFIGURED", ""),
        ("Desempenhos esperados", 114, "PENDING", ""),
        ("Duplicados", 0, "PENDING", ""),
        ("Erros críticos", 0, "PENDING", ""),
        ("Estado final", "", "PENDING", ""),
    ]

    for row_index, row_values in enumerate(
        rows,
        start=2,
    ):
        for column_index, value in enumerate(
            row_values,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 25
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 40


def _populate_leagues_sheet(worksheet) -> None:
    leagues = get_active_leagues()

    for row_index, (
        league_id,
        league,
    ) in enumerate(
        leagues.items(),
        start=2,
    ):
        values = [
            league_id,
            league["name"],
            league["country"],
            league["country_code"],
            league["season_label"],
            league["team_count"],
            league["matches_per_team"],
            league["total_matches"],
            league["league_strength_factor"],
            league["relegation_places"],
            league["playoff_places"],
            int(bool(league["active"])),
        ]

        for column_index, value in enumerate(
            values,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )


def _add_data_validations(workbook: Workbook) -> None:
    league_ids = ",".join(
        get_active_leagues().keys()
    )

    league_validation = DataValidation(
        type="list",
        formula1=f'"{league_ids}"',
        allow_blank=False,
    )

    promoted_validation = DataValidation(
        type="list",
        formula1='"0,1"',
        allow_blank=False,
    )

    active_validation = DataValidation(
        type="list",
        formula1='"0,1"',
        allow_blank=False,
    )

    promotion_method_validation = DataValidation(
        type="list",
        formula1='"CHAMPION,DIRECT,PLAYOFF"',
        allow_blank=True,
    )

    source_status_validation = DataValidation(
        type="list",
        formula1=(
            '"CONFIRMED,COMPLETE,PARTIAL,CACHED,'
            'MISSING,CONFLICTING,OUTDATED,MANUAL_VALIDATED"'
        ),
        allow_blank=False,
    )

    match_status_validation = DataValidation(
        type="list",
        formula1=(
            '"SCHEDULED,PLAYED,POSTPONED,'
            'CANCELLED,ABANDONED,AWARDED"'
        ),
        allow_blank=False,
    )

    schedule_type_validation = DataValidation(
        type="list",
        formula1='"OFFICIAL,SYNTHETIC"',
        allow_blank=False,
    )

    severity_validation = DataValidation(
        type="list",
        formula1='"ERROR,WARNING,INFO"',
        allow_blank=False,
    )

    resolved_validation = DataValidation(
        type="list",
        formula1='"0,1"',
        allow_blank=False,
    )

    teams_sheet = workbook["Equipas_2026_27"]

    for validation in (
        league_validation,
        promoted_validation,
        active_validation,
        promotion_method_validation,
    ):
        teams_sheet.add_data_validation(validation)

    league_validation.add("E2:E500")
    promoted_validation.add("H2:H500")
    promotion_method_validation.add("I2:I500")
    active_validation.add("K2:K500")

    performance_sheet = workbook["Desempenho_2025_26"]

    performance_league_source = DataValidation(
        type="list",
        formula1=f'"{league_ids},ENG2,ESP2,ITA2,GER2,FRA2,POR2"',
        allow_blank=False,
    )

    performance_league_target = DataValidation(
        type="list",
        formula1=f'"{league_ids}"',
        allow_blank=False,
    )

    performance_promoted = DataValidation(
        type="list",
        formula1='"0,1"',
        allow_blank=False,
    )

    performance_promotion_method = DataValidation(
        type="list",
        formula1='"CHAMPION,DIRECT,PLAYOFF"',
        allow_blank=True,
    )

    performance_sheet.add_data_validation(
        performance_league_source
    )
    performance_sheet.add_data_validation(
        performance_league_target
    )
    performance_sheet.add_data_validation(
        performance_promoted
    )
    performance_sheet.add_data_validation(
        performance_promotion_method
    )
    performance_sheet.add_data_validation(
        source_status_validation
    )

    performance_league_source.add("B2:B500")
    performance_league_target.add("C2:C500")
    performance_promoted.add("O2:O500")
    performance_promotion_method.add("P2:P500")
    source_status_validation.add("Q2:Q500")

    promoted_sheet = workbook["Promovidas"]

    promoted_method = DataValidation(
        type="list",
        formula1='"CHAMPION,DIRECT,PLAYOFF"',
        allow_blank=False,
    )

    promoted_status = DataValidation(
        type="list",
        formula1=(
            '"CONFIRMED,COMPLETE,PARTIAL,CACHED,'
            'MISSING,CONFLICTING,OUTDATED,MANUAL_VALIDATED"'
        ),
        allow_blank=False,
    )

    promoted_sheet.add_data_validation(promoted_method)
    promoted_sheet.add_data_validation(promoted_status)

    promoted_method.add("E2:E100")
    promoted_status.add("L2:L100")

    fixtures_sheet = workbook["Calendario_2026_27"]

    fixture_league_validation = DataValidation(
        type="list",
        formula1=f'"{league_ids}"',
        allow_blank=False,
    )

    fixtures_sheet.add_data_validation(
        fixture_league_validation
    )
    fixtures_sheet.add_data_validation(
        match_status_validation
    )
    fixtures_sheet.add_data_validation(
        schedule_type_validation
    )

    fixture_league_validation.add("B2:B3000")
    match_status_validation.add("H2:H3000")
    schedule_type_validation.add("K2:K3000")

    sources_sheet = workbook["Fontes"]

    source_status_validation_copy = DataValidation(
        type="list",
        formula1=(
            '"CONFIRMED,COMPLETE,PARTIAL,CACHED,'
            'MISSING,CONFLICTING,OUTDATED,MANUAL_VALIDATED"'
        ),
        allow_blank=False,
    )

    sources_sheet.add_data_validation(
        source_status_validation_copy
    )
    source_status_validation_copy.add("I2:I5000")

    validation_sheet = workbook["Validacao"]
    validation_sheet.add_data_validation(
        severity_validation
    )
    validation_sheet.add_data_validation(
        resolved_validation
    )

    severity_validation.add("B2:B5000")
    resolved_validation.add("I2:I5000")


def _apply_number_formats(workbook: Workbook) -> None:
    leagues_sheet = workbook["Ligas"]

    for row in range(2, 100):
        leagues_sheet.cell(
            row=row,
            column=9,
        ).number_format = "0.00"

    performance_sheet = workbook["Desempenho_2025_26"]

    for row in range(2, 500):
        performance_sheet.cell(
            row=row,
            column=18,
        ).number_format = "0.00"

        performance_sheet.cell(
            row=row,
            column=20,
        ).number_format = "yyyy-mm-dd hh:mm"

    promoted_sheet = workbook["Promovidas"]

    for row in range(2, 100):
        promoted_sheet.cell(
            row=row,
            column=11,
        ).number_format = "0.00"

        promoted_sheet.cell(
            row=row,
            column=13,
        ).number_format = "0.00"

    fixtures_sheet = workbook["Calendario_2026_27"]

    for row in range(2, 3000):
        fixtures_sheet.cell(
            row=row,
            column=5,
        ).number_format = "yyyy-mm-dd hh:mm"

    sources_sheet = workbook["Fontes"]

    for row in range(2, 5000):
        sources_sheet.cell(
            row=row,
            column=8,
        ).number_format = "yyyy-mm-dd hh:mm"


if __name__ == "__main__":
    created_file = create_dataset_template()

    print(f"✅ Template criado: {created_file}")
